#!/usr/bin/env python3
"""Importa a "Planilha de Controle Geral" (abas Fina-R / Fina-D) para financial_entries.

Loader REPRODUZÍVEL e idempotente (provenance — substitui o import ad-hoc de 2026-06-03).
NÃO contém dados nem credenciais: lê o .xlsm local e EMITE SQL no stdout (ou arquivo),
que deve ser revisado e aplicado no DB do tenant correspondente. Dado sensível (PF/PJ):
aplicar só no tenant dono, com a UI gestor-only (/reports/financeiro).

Uso:
    python3 scripts/import_financeiro_planilha.py <caminho.xlsm> <org_id> [saida.sql]

Mapeamento (header na linha 6 de cada aba):
  Fina-R: Data Prevista Receb. | Valor | Tipo | Processo Ref. | Tipo Cliente | Cliente | Parcela | Recebido? | Data Receb.
  Fina-D: Data Prevista Pag.   | Valor | Despesa | Categoria | Parcela | Pago? | Data Efetiva Pag.
Idempotência: DELETE WHERE org_id=:org AND source='planilha' antes do INSERT.
"""
import sys
import datetime
import re


def _parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        d = v.date() if isinstance(v, datetime.datetime) else v
        return d if d.year >= 2000 else None
    return None


def _q(s):
    if s is None or str(s).strip() == "":
        return "NULL"
    return "'" + str(s).strip().replace("'", "''")[:300] + "'"


def _d(dt):
    return "DATE '%s'" % dt.isoformat() if dt else "NULL"


def _num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _yn(v):
    return "true" if str(v or "").strip().lower() in ("sim", "yes", "true", "1") else "false"


def build_sql(xlsm_path: str, org_id: int) -> str:
    import openpyxl  # lazy: só quem roda o loader precisa da lib
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    out = ["BEGIN;", "DELETE FROM financial_entries WHERE org_id=%d AND source='planilha';" % org_id]

    for r in wb["Fina-R"].iter_rows(min_row=7, values_only=True):
        valor, dp = _num(r[3]) if len(r) > 3 else None, _parse_date(r[2]) if len(r) > 2 else None
        if valor is None or dp is None:
            continue
        g = lambda i: r[i] if len(r) > i else None
        out.append(
            "INSERT INTO financial_entries (org_id,kind,valor,data_prevista,data_efetiva,settled,tipo,descricao,processo_ref,cliente,tipo_cliente,parcela,source) "
            "VALUES (%d,'receita',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'planilha');"
            % (org_id, valor, _d(dp), _d(_parse_date(g(10))), _yn(g(9)), _q(g(4)), _q(g(5)), _q(g(5)), _q(g(7)), _q(g(6)), _q(g(8)))
        )

    for r in wb["Fina-D"].iter_rows(min_row=7, values_only=True):
        valor, dp = _num(r[3]) if len(r) > 3 else None, _parse_date(r[2]) if len(r) > 2 else None
        if valor is None or dp is None:
            continue
        g = lambda i: r[i] if len(r) > i else None
        out.append(
            "INSERT INTO financial_entries (org_id,kind,valor,data_prevista,data_efetiva,settled,tipo,descricao,parcela,source) "
            "VALUES (%d,'despesa',%s,%s,%s,%s,%s,%s,%s,'planilha');"
            % (org_id, valor, _d(dp), _d(_parse_date(g(8))), _yn(g(7)), _q(g(5)), _q(g(4)), _q(g(6)))
        )

    out.append("COMMIT;")
    return "\n".join(out) + "\n"


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    sql = build_sql(sys.argv[1], int(sys.argv[2]))
    if len(sys.argv) > 3:
        open(sys.argv[3], "w").write(sql)
        print("SQL escrito em %s (%d statements)" % (sys.argv[3], sql.count(";")))
    else:
        sys.stdout.write(sql)
