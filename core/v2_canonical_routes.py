"""
UI Remake — Canonical route handlers (rich templates only).

REGRAS:
1. Sem /v2/* — só canonical paths (/casehub/X direto).
2. Override APENAS rotas com template rico portado em templates/app/.
3. Para qualquer rota sem template rico, NÃO registramos override aqui —
   o legacy router include responde normalmente com a UI legacy intacta.
4. Funcionalidade legacy (WhatsApp, Emails, Tools BR, Letters, Questionnaires,
   USCIS, ILC Tools, Onboarding, Bulk, etc) continua funcionando exatamente
   como antes nas rotas canônicas tradicionais.

Registered BEFORE app.include_router() loop in app_factory.create_app() so
FastAPI first-match-wins makes these win over the legacy router includes
for the specific paths we cover.
"""
from __future__ import annotations
from datetime import date
from typing import Any, Callable, Optional

from fastapi import FastAPI, Depends, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.orm import Session

from models import get_db
from auth import get_current_user


def register_canonical_routes(app: FastAPI, templates: Any, get_context: Callable, PREFIX: str) -> None:
    """Wire canonical handlers for rich (fully ported) templates only.

    Called from create_app() right after the dashboard handler and BEFORE
    the legacy router include loop.
    """

    def _redirect_login(target: str) -> RedirectResponse:
        return RedirectResponse(url=f"{PREFIX}/login?next={target}", status_code=302)

    async def _delegate_async(legacy_fn, **kwargs) -> dict:
        """Call legacy handler, extract its TemplateResponse.context dict."""
        try:
            import inspect
            coro = legacy_fn(**kwargs)
            response = await coro if inspect.iscoroutine(coro) else coro
            return getattr(response, "context", None) or {}
        except Exception:
            return {}

    # ─────────── Financeiro: gate reutilizável (dado financeiro sensível) ───────────
    # RED LINE (Equipe CaseHub 03/06): /reports/financeiro e seus endpoints de mutação são
    # restritos a SUPERADMIN OU aos sócios explícitos da org na allowlist
    # organizations.settings.financeiro_user_ids. NÃO usar has_permission(cases.edit)
    # nem 'admin' genérico (QA/admins operacionais NÃO veem dado financeiro). Sócio
    # NÃO vira superadmin (superadmin é cross-tenant). Org-scoped sempre.
    def _financeiro_user_allowed(user, org_id, db) -> bool:
        """True se o usuário pode ver/editar o financeiro DESTA org. Pura (sem HTTP)."""
        from sqlalchemy import text as _sql_text
        _utype = (getattr(user, "user_type", None) or "").lower()
        if _utype == "superadmin":
            return True
        if org_id is None:
            return False
        try:
            import json as _json
            _row = db.execute(
                _sql_text("SELECT settings FROM organizations WHERE id=:o"),
                {"o": org_id},
            ).first()
            _st = (_row[0] if _row else None) or {}
            if isinstance(_st, str):
                _st = _json.loads(_st or "{}")
            return user.id in (_st.get("financeiro_user_ids") or [])
        except Exception:
            return False

    def _require_financeiro_access(request, db):
        """Gate central p/ TODO endpoint financeiro.

        Retorna (user, org_id) quando autorizado. Caso contrário retorna uma
        Response (redirect de login OU 403) que o handler deve devolver
        imediatamente. Chamadores: `result = _require_financeiro_access(...);`
        `if not isinstance(result, tuple): return result; user, org_id = result`.
        """
        from fastapi.responses import JSONResponse as _JSONResponse
        user = get_current_user(request, db)
        if not user:
            # GET → redirect; mutações/JSON → 401 estruturado.
            if request.method == "GET":
                return _redirect_login(request.url.path)
            return _JSONResponse({"error": "Nao autenticado"}, status_code=401)

        org_id = getattr(request.state, "org_id", None)
        if org_id is None:
            org_id = getattr(user, "org_id", None)

        if not _financeiro_user_allowed(user, org_id, db):
            msg = "Relatórios financeiros são restritos aos sócios do escritório."
            if request.method == "GET":
                return templates.TemplateResponse(
                    "app/errors/403.html",
                    {**get_context(request, db, user=user), "user": user, "message": msg},
                    status_code=403,
                )
            return _JSONResponse({"error": msg}, status_code=403)
        return (user, org_id)

    # ─────────── Workspace utility surfaces ───────────
    @app.get(f"{PREFIX}/casehub-md", response_class=HTMLResponse)
    async def casehub_md_workspace_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/casehub-md")
        ctx = {
            **get_context(request, db, user=user),
            "user": user,
            "today": date.today(),
            "active_module": "md",
        }
        return templates.TemplateResponse("app/casehub_md/index.html", ctx)

    @app.get(f"{PREFIX}/whatsapp-chat", response_class=HTMLResponse)
    async def whatsapp_chat_workspace_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/whatsapp-chat")
        try:
            from routes.whatsapp_chat import get_bot_status as _get_bot_status
            bot_status = await _get_bot_status()
        except Exception:
            bot_status = {"connected": False, "status": "offline", "ok": False}
        ctx = {
            **get_context(request, db, user=user),
            "user": user,
            "today": date.today(),
            "active_module": "whatsapp",
            "bot_status": bot_status,
        }
        return templates.TemplateResponse("app/whatsapp/chat.html", ctx)

    @app.get(f"{PREFIX}/route-map", response_class=HTMLResponse)
    async def route_map_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/route-map")
        groups = [
            ("Produto", [
                ("Painel", f"{PREFIX}/dashboard"),
                ("Controladoria", f"{PREFIX}/controladoria"),
                ("Agenda", f"{PREFIX}/calendar/agenda"),
                ("Kanban", f"{PREFIX}/tasks/kanban"),
                ("Processos", f"{PREFIX}/cases"),
                ("Clientes", f"{PREFIX}/clients"),
                ("WhatsApp Chat", f"{PREFIX}/whatsapp-chat"),
                ("Maestro", f"{PREFIX}/assistente"),
                ("Documentos / Drive", f"{PREFIX}/documents"),
                ("CaseHub.md", f"{PREFIX}/casehub-md"),
            ]),
            ("Operação", [
                ("Financeiro", f"{PREFIX}/reports/financeiro"),
                ("Faturas", f"{PREFIX}/invoices"),
                ("Pagamentos", f"{PREFIX}/payments"),
                ("Modelos de documentos", f"{PREFIX}/doc-templates"),
                ("Relatórios", f"{PREFIX}/reports"),
                ("Notificações", f"{PREFIX}/notifications"),
            ] + (
                # Emails é gestor-only: só aparece no mapa de rotas para
                # admin/superadmin (gate server-side vive em routes/_email_gate.py).
                [("Emails", f"{PREFIX}/emails")]
                if (getattr(user, "user_type", "") or "").lower() in ("admin", "superadmin")
                else []
            )),
            ("Conta e Admin", [
                ("Perfil", f"{PREFIX}/profile"),
                ("Configurações", f"{PREFIX}/settings"),
                ("Numeração", f"{PREFIX}/settings/numbering"),
                ("Integrações", f"{PREFIX}/integrations"),
                ("Assinatura", f"{PREFIX}/subscription"),
                ("Admin", f"{PREFIX}/admin"),
                ("Usuários", f"{PREFIX}/admin/users"),
                ("Customização", f"{PREFIX}/admin/customizacao"),
                ("Branding", f"{PREFIX}/admin/branding"),
                ("Design editor", f"{PREFIX}/admin/design-editor"),
            ]),
        ]
        ctx = {
            **get_context(request, db, user=user),
            "user": user,
            "today": date.today(),
            "active_module": "route-map",
            "route_groups": groups,
        }
        return templates.TemplateResponse("app/route_map/index.html", ctx)

    # ─────────── Clients ───────────
    @app.get(f"{PREFIX}/clients/new", response_class=HTMLResponse)
    async def clients_new_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/clients/new")
        try:
            from routes.clients import new_client as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/clients/form.html", ctx)

    @app.get(f"{PREFIX}/clients/{{client_id}}", response_class=HTMLResponse)
    async def clients_detail_canon(client_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/clients/{client_id}")
        try:
            from routes.clients import view_client as _fn
            ctx = await _delegate_async(_fn, client_id=client_id, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/clients/detail.html", ctx)

    @app.get(f"{PREFIX}/clients/{{client_id}}/edit", response_class=HTMLResponse)
    async def clients_edit_canon(client_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/clients/{client_id}/edit")
        try:
            # Legacy fn is exported as `edit_client_form` (not `edit_client`).
            # Importing the wrong name silently dropped ctx → template UndefinedError.
            from routes.clients import edit_client_form as _fn
            ctx = await _delegate_async(_fn, client_id=client_id, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/clients/form.html", ctx)

    # ─────────── Cases ───────────
    @app.get(f"{PREFIX}/cases/new", response_class=HTMLResponse)
    async def cases_new_canon(request: Request, client_id: Optional[int] = None, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/cases/new")
        try:
            from routes.cases import new_case as _fn
            ctx = await _delegate_async(_fn, request=request, client_id=client_id, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/cases/form.html", ctx)

    @app.get(f"{PREFIX}/cases/{{case_id}}", response_class=HTMLResponse)
    async def cases_detail_canon(case_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/cases/{case_id}")
        try:
            from routes.cases import view_case as _fn
            ctx = await _delegate_async(_fn, case_id=case_id, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/cases/detail.html", ctx)

    @app.get(f"{PREFIX}/cases/{{case_id}}/edit", response_class=HTMLResponse)
    async def cases_edit_canon(case_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/cases/{case_id}/edit")
        try:
            # Legacy fn is `edit_case_form` (parity com edit_client_form fix BL-1).
            from routes.cases import edit_case_form as _fn
            ctx = await _delegate_async(_fn, case_id=case_id, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/cases/form.html", ctx)

    # ─────────── Tasks ───────────
    @app.get(f"{PREFIX}/tasks/new", response_class=HTMLResponse)
    async def tasks_new_canon(
        request: Request,
        client_id: Optional[int] = None,
        case_id: Optional[int] = None,
        db: Session = Depends(get_db),
    ):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/tasks/new")
        try:
            from routes.tasks import new_task as _fn
            ctx = await _delegate_async(_fn, request=request, client_id=client_id, case_id=case_id, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/tasks/form.html", ctx)

    @app.get(f"{PREFIX}/tasks/calendar", response_class=HTMLResponse)
    async def tasks_calendar_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/tasks/calendar")
        from models import Task as _Task
        from models.tenant import tenant_query as _tenant_query
        today_ = date.today()
        tasks = _tenant_query(db, _Task, request.state.org_id).filter(
            _Task.due_date.isnot(None), _Task.status != "completed",
        ).order_by(_Task.due_date.asc()).limit(200).all()
        return templates.TemplateResponse("app/tasks/calendar_view.html", {
            **get_context(request, db, user=user),
            "user": user, "today": today_, "tasks": tasks,
        })

    # ─────────── Invoices ───────────
    @app.get(f"{PREFIX}/invoices", response_class=HTMLResponse)
    async def invoices_list_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/invoices")
        try:
            # Legacy fn is `invoice_list`, not `list_invoices`.
            from routes.invoices import invoice_list as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/invoices/list.html", ctx)

    @app.get(f"{PREFIX}/invoices/new", response_class=HTMLResponse)
    async def invoices_new_canon(request: Request, case_id: Optional[int] = None, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/invoices/new")
        try:
            # Legacy fn is `new_invoice_form`, not `new_invoice`.
            from routes.invoices import new_invoice_form as _fn
            ctx = await _delegate_async(_fn, request=request, case_id=case_id, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/invoices/form.html", ctx)

    @app.get(f"{PREFIX}/invoices/{{invoice_number}}", response_class=HTMLResponse)
    async def invoices_detail_canon(invoice_number: str, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/invoices/{invoice_number}")
        try:
            from routes.invoices import view_invoice as _fn
            ctx = await _delegate_async(_fn, invoice_number=invoice_number, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/invoices/detail.html", ctx)

    @app.get(f"{PREFIX}/invoices/{{invoice_number}}/print", response_class=HTMLResponse)
    async def invoices_print_canon(invoice_number: str, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/invoices/{invoice_number}/print")
        try:
            from routes.invoices import print_invoice as _fn
            ctx = await _delegate_async(_fn, invoice_number=invoice_number, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        return templates.TemplateResponse("app/invoices/print.html", ctx)

    # ─────────── Billing sub-routes ───────────
    @app.get(f"{PREFIX}/billing/items/new", response_class=HTMLResponse)
    async def billing_items_new_canon(
        request: Request,
        case_id: Optional[int] = None,
        db: Session = Depends(get_db),
    ):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/billing/items/new")
        try:
            from routes.billing import new_billing_item as _fn
            ctx = await _delegate_async(_fn, request=request, case_id=case_id, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/billing/items_form.html", ctx)

    @app.get(f"{PREFIX}/billing/time/new", response_class=HTMLResponse)
    async def billing_time_new_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/billing/time/new")
        try:
            from routes.billing import new_time_entry as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/billing/time_form.html", ctx)

    # ─────────── Documents sub-routes ───────────
    @app.get(f"{PREFIX}/documents/upload", response_class=HTMLResponse)
    async def documents_upload_canon(
        request: Request,
        client_id: Optional[int] = None,
        case_id: Optional[int] = None,
        db: Session = Depends(get_db),
    ):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/documents/upload")
        try:
            from routes.documents import upload_form as _fn
            ctx = await _delegate_async(_fn, request=request, client_id=client_id, case_id=case_id, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/documents/upload.html", ctx)

    @app.get(f"{PREFIX}/documents/{{doc_id}}", response_class=HTMLResponse)
    async def documents_detail_canon(doc_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/documents/{doc_id}")
        try:
            from routes.documents import view_document as _fn
            ctx = await _delegate_async(_fn, doc_id=doc_id, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/documents/detail.html", ctx)

    # ─────────── Doc templates ───────────
    @app.get(f"{PREFIX}/doc-templates", response_class=HTMLResponse)
    async def doc_templates_list_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/doc-templates")
        try:
            # Legacy fn is `template_list`, not `list_templates`.
            from routes.doc_templates import template_list as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/doc_templates/list.html", ctx)

    @app.get(f"{PREFIX}/doc-templates/new", response_class=HTMLResponse)
    async def doc_templates_new_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/doc-templates/new")
        try:
            # Legacy fn is `new_template_form`.
            from routes.doc_templates import new_template_form as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/doc_templates/form.html", ctx)

    @app.get(f"{PREFIX}/doc-templates/{{template_id}}", response_class=HTMLResponse)
    async def doc_templates_detail_canon(template_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/doc-templates/{template_id}")
        try:
            # `view_template` now exists in routes/doc_templates.py (added
            # alongside this fix) — it never had a legacy equivalent before.
            from routes.doc_templates import view_template as _fn
            ctx = await _delegate_async(_fn, template_id=template_id, request=request, db=db)
        except Exception:
            ctx = {"template_data": None}
        # `_delegate_async` swallows the 404 `view_template` raises for an
        # unknown/cross-tenant id (see its own bare `except Exception:
        # return {}`), so `ctx` can come back without a `template_data` key
        # at all in that case — NOT just when the whole try/except above
        # fires. The template unconditionally accesses `template_data.*`,
        # and Jinja's default Undefined raises on attribute access (not a
        # silent None), so this default is required, not cosmetic.
        ctx.setdefault("template_data", None)
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/doc_templates/detail.html", ctx)

    @app.get(f"{PREFIX}/doc-templates/{{template_id}}/edit", response_class=HTMLResponse)
    async def doc_templates_edit_canon(template_id: int, request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/doc-templates/{template_id}/edit")
        try:
            # Legacy fn is `edit_template_form`, not `edit_template`.
            from routes.doc_templates import edit_template_form as _fn
            ctx = await _delegate_async(_fn, template_id=template_id, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/doc_templates/form.html", ctx)

    # ─────────── Payments (Stripe flow rich) ───────────
    @app.get(f"{PREFIX}/payments", response_class=HTMLResponse)
    async def payments_list_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/payments")
        ctx = {**get_context(request, db, user=user), "user": user, "today": date.today()}
        try:
            # Legacy fn is `payment_overview`, not `overview`.
            from routes.payments import payment_overview as _fn  # type: ignore
            extra = await _delegate_async(_fn, request=request, db=db)
            ctx.update(extra)
        except Exception:
            pass
        return templates.TemplateResponse("app/payments/list.html", ctx)

    @app.get(f"{PREFIX}/payments/success", response_class=HTMLResponse)
    async def payments_success_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/payments/success")
        ctx = {**get_context(request, db, user=user), "user": user, "today": date.today()}
        return templates.TemplateResponse("app/payments/success.html", ctx)

    @app.get(f"{PREFIX}/payments/cancel", response_class=HTMLResponse)
    async def payments_cancel_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/payments/cancel")
        ctx = {**get_context(request, db, user=user), "user": user, "today": date.today()}
        return templates.TemplateResponse("app/payments/cancel.html", ctx)

    @app.get(f"{PREFIX}/payments/error", response_class=HTMLResponse)
    async def payments_error_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/payments/error")
        ctx = {**get_context(request, db, user=user), "user": user, "today": date.today()}
        return templates.TemplateResponse("app/payments/error.html", ctx)

    # ─────────── Admin (rich) ───────────
    def _admin_canon(template_name: str, handler_name: str, url_suffix: str):
        async def _impl(request: Request, db: Session = Depends(get_db)):
            user = get_current_user(request, db)
            if not user:
                return _redirect_login(f"{PREFIX}{url_suffix}")
            try:
                from routes import admin as _admin_mod
                fn = getattr(_admin_mod, handler_name, None)
                if fn:
                    ctx = await _delegate_async(fn, request=request, db=db)
                else:
                    ctx = {}
            except Exception:
                ctx = {}
            ctx.setdefault("user", user); ctx.setdefault("today", date.today())
            ctx.update(get_context(request, db, user=user))
            return templates.TemplateResponse(template_name, ctx)
        return _impl

    app.get(f"{PREFIX}/admin", response_class=HTMLResponse)(_admin_canon("app/admin/home.html", "admin_home", "/admin"))
    app.get(f"{PREFIX}/admin/users", response_class=HTMLResponse)(_admin_canon("app/admin/users.html", "list_users", "/admin/users"))
    app.get(f"{PREFIX}/admin/users/new", response_class=HTMLResponse)(_admin_canon("app/admin/user_form.html", "new_user_form", "/admin/users/new"))
    # Legacy fn is `settings`, not `admin_settings`.
    app.get(f"{PREFIX}/admin/settings", response_class=HTMLResponse)(_admin_canon("app/admin/settings.html", "settings", "/admin/settings"))

    @app.get(f"{PREFIX}/admin/branding", response_class=HTMLResponse)
    async def admin_branding_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/admin/branding")
        try:
            from routes.branding import branding_page as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/admin/branding.html", ctx)

    @app.get(f"{PREFIX}/admin/customizacao", response_class=HTMLResponse)
    async def admin_customizacao_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/admin/customizacao")
        try:
            from routes.customizacao import customizacao_page as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/admin/customizacao.html", ctx)

    @app.get(f"{PREFIX}/admin/design-editor", response_class=HTMLResponse)
    async def admin_design_editor_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/admin/design-editor")
        try:
            # Legacy fn is `design_editor_page`.
            from routes.design_editor import design_editor_page as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/admin/design_editor.html", ctx)

    # ─────────── Profile + 2FA + Integrations + Subscription ───────────
    @app.get(f"{PREFIX}/profile", response_class=HTMLResponse)
    async def profile_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/profile")
        try:
            from routes.profile import profile_page as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/profile/index.html", ctx)

    @app.get(f"{PREFIX}/2fa/setup", response_class=HTMLResponse)
    async def two_factor_setup_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/2fa/setup")
        try:
            from routes.two_factor import setup_2fa_page as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/two_factor/setup.html", ctx)

    @app.get(f"{PREFIX}/integrations", response_class=HTMLResponse)
    async def integrations_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/integrations")
        try:
            from routes.integrations import _integration_cards
            org_id = getattr(request.state, "org_id", None)
            ctx = {"integrations": _integration_cards(org_id, db)}
        except Exception:
            ctx = {"integrations": []}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.setdefault("active_module", "settings")
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/integrations/index.html", ctx)

    @app.get(f"{PREFIX}/subscription", response_class=HTMLResponse)
    async def subscription_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/subscription")
        try:
            from routes.subscription import subscription_dashboard as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/subscription/index.html", ctx)

    def _customizacao_redirect(tab: str):
        def _impl():
            return RedirectResponse(url=f"{PREFIX}/admin/customizacao?tab={tab}", status_code=302)
        return _impl

    # `integracoes` tab → canonical /integrations hub (audit P0.3:
    # eliminate 3 concurrent integration screens). 301 permanent so search
    # engines and bookmarks update.
    @app.get(f"{PREFIX}/admin/customizacao/integracoes", include_in_schema=False)
    async def _customizacao_integracoes_to_hub():
        return RedirectResponse(url=f"{PREFIX}/integrations", status_code=301)

    for _customizacao_tab in ("notificacoes", "aparencia", "sidebar", "widgets"):
        app.get(f"{PREFIX}/admin/customizacao/{_customizacao_tab}")(_customizacao_redirect(_customizacao_tab))

    # ─────────── Settings sub ───────────
    @app.get(f"{PREFIX}/settings/numbering", response_class=HTMLResponse)
    async def settings_numbering_canon(request: Request, db: Session = Depends(get_db)):
        user = get_current_user(request, db)
        if not user:
            return _redirect_login(f"{PREFIX}/settings/numbering")
        try:
            from routes.settings import numbering_settings_page as _fn
            ctx = await _delegate_async(_fn, request=request, db=db)
        except Exception:
            ctx = {}
        ctx.setdefault("user", user); ctx.setdefault("today", date.today())
        ctx.update(get_context(request, db, user=user))
        return templates.TemplateResponse("app/settings/numbering.html", ctx)

    # ─────────── Relatórios · Inteligência Financeira (gestor-only) ───────────
    # Reunião 03/06: cruzar a planilha financeira (financial_entries) com a
    # produtividade da Controladoria (prazos_processuais) por PERÍODO.
    # GATING: dado financeiro é sensível → só admin/superadmin. Advogado/estagiário 403.
    # Org-scoped sempre (WHERE org_id). Agregação no banco (SUM/GROUP BY), nunca puxa
    # as ~4000 linhas pro front. Charts SVG/CSS leves (sem libs).
    @app.get(f"{PREFIX}/reports/financeiro", response_class=HTMLResponse)
    async def reports_financeiro_canon(
        request: Request,
        ano: Optional[int] = None,
        page: int = 1,
        sort: str = "previsto",
        direction: str = "desc",
        db: Session = Depends(get_db),
    ):
        from sqlalchemy import text as _sql_text

        # Gate central (superadmin OU financeiro_user_ids da org). Mesmo gate dos
        # endpoints de mutação — extraído em _require_financeiro_access.
        _gate = _require_financeiro_access(request, db)
        if not isinstance(_gate, tuple):
            return _gate
        user, org_id = _gate

        is_pg = (db.get_bind().dialect.name == "postgresql") if db.get_bind() is not None else False

        # Soft-delete: agregações ignoram linhas com ativo=FALSE. 'ativo IS NOT FALSE'
        # mantém NULL legado + TRUE (sem precisar de backfill). Dialect-aware como o
        # padrão de `settled` no resto do route (SQLite não tem keyword FALSE confiável
        # em todas as versões → comparamos com 0).
        active_clause = "ativo IS NOT FALSE" if is_pg else "(ativo IS NULL OR ativo <> 0)"

        # Anos disponíveis na base financeira da org (para o seletor).
        anos_rows = db.execute(
            _sql_text(f"""
                SELECT DISTINCT EXTRACT(YEAR FROM COALESCE(data_efetiva, data_prevista))::int AS ano
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause}
                  AND COALESCE(data_efetiva, data_prevista) IS NOT NULL
                ORDER BY ano DESC
            """) if is_pg else _sql_text(f"""
                SELECT DISTINCT CAST(strftime('%Y', COALESCE(data_efetiva, data_prevista)) AS INTEGER) AS ano
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause}
                  AND COALESCE(data_efetiva, data_prevista) IS NOT NULL
                ORDER BY ano DESC
            """),
            {"org_id": org_id},
        ).fetchall()
        anos_disponiveis = [int(r.ano) for r in anos_rows if r.ano is not None]
        if not anos_disponiveis:
            anos_disponiveis = [date.today().year]
        if ano not in anos_disponiveis:
            # Default = ano corrente se tiver lançamentos (é o que o sócio quer ver ao
            # abrir a tela); senão o ano mais recente com dados. Evita cair num ano
            # futuro esparso (ex.: parcelas de 2027) e parecer que a planilha sumiu.
            _yr = date.today().year
            ano = _yr if _yr in anos_disponiveis else anos_disponiveis[0]

        # COALESCE(data_efetiva, data_prevista) = data de competência do lançamento.
        mexpr = ("EXTRACT(MONTH FROM COALESCE(data_efetiva, data_prevista))::int"
                 if is_pg else
                 "CAST(strftime('%m', COALESCE(data_efetiva, data_prevista)) AS INTEGER)")
        yexpr = ("EXTRACT(YEAR FROM COALESCE(data_efetiva, data_prevista))::int"
                 if is_pg else
                 "CAST(strftime('%Y', COALESCE(data_efetiva, data_prevista)) AS INTEGER)")
        params = {"org_id": org_id, "ano": ano}

        # (a) Resumo mensal: receita / despesa por mês (12 meses) — espelha aba F-Resu.
        mensal_rows = db.execute(
            _sql_text(f"""
                SELECT {mexpr} AS mes, kind,
                       SUM(valor) AS total,
                       SUM(CASE WHEN settled THEN valor ELSE 0 END) AS total_liquidado
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause} AND {yexpr} = :ano
                GROUP BY {mexpr}, kind
            """),
            params,
        ).fetchall()
        meses = {m: {"receita": 0.0, "despesa": 0.0,
                     "receita_liq": 0.0, "despesa_liq": 0.0} for m in range(1, 13)}
        for r in mensal_rows:
            m = int(r.mes or 0)
            if m not in meses:
                continue
            k = "receita" if r.kind == "receita" else "despesa"
            meses[m][k] = float(r.total or 0)
            meses[m][f"{k}_liq"] = float(r.total_liquidado or 0)
        resumo_mensal = []
        for m in range(1, 13):
            rec = meses[m]["receita"]; desp = meses[m]["despesa"]
            resumo_mensal.append({
                "mes": m, "receita": rec, "despesa": desp, "resultado": rec - desp,
                "receita_liq": meses[m]["receita_liq"], "despesa_liq": meses[m]["despesa_liq"],
                "resultado_liq": meses[m]["receita_liq"] - meses[m]["despesa_liq"],
            })
        tot_receita = sum(x["receita"] for x in resumo_mensal)
        tot_despesa = sum(x["despesa"] for x in resumo_mensal)
        # Liquidado (settled) vs em aberto — a leitura mais valiosa que a página escondia.
        tot_receita_liq = sum(x["receita_liq"] for x in resumo_mensal)
        tot_despesa_liq = sum(x["despesa_liq"] for x in resumo_mensal)
        a_receber = tot_receita - tot_receita_liq   # receita não-settled
        a_pagar = tot_despesa - tot_despesa_liq      # despesa não-settled

        # (a2) Contas a receber/pagar por mês × settled (F-DasR / F-DasD).
        # Competência = data_prevista (é o vencimento); "em atraso" = não-settled e
        # data_prevista < hoje. Agregação no banco; uma linha por (mês,kind,settled,vencido).
        hoje = date.today()
        cmexpr = ("EXTRACT(MONTH FROM data_prevista)::int" if is_pg
                  else "CAST(strftime('%m', data_prevista) AS INTEGER)")
        cyexpr = ("EXTRACT(YEAR FROM data_prevista)::int" if is_pg
                  else "CAST(strftime('%Y', data_prevista) AS INTEGER)")
        venc_expr = ("(NOT settled AND data_prevista < :hoje)" if is_pg
                     else "(settled = 0 AND data_prevista < :hoje)")
        contas_params = {"org_id": org_id, "ano": ano, "hoje": hoje}
        contas_rows = db.execute(
            _sql_text(f"""
                SELECT {cmexpr} AS mes, kind, settled,
                       SUM(valor) AS total,
                       SUM(CASE WHEN {venc_expr} THEN valor ELSE 0 END) AS total_vencido
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause}
                  AND {cyexpr} = :ano AND data_prevista IS NOT NULL
                GROUP BY {cmexpr}, kind, settled
            """),
            contas_params,
        ).fetchall()

        def _empty_contas():
            return {m: {"liquidado": 0.0, "aberto": 0.0, "vencido": 0.0} for m in range(1, 13)}
        cr_meses, cp_meses = _empty_contas(), _empty_contas()
        for r in contas_rows:
            m = int(r.mes or 0)
            if m not in cr_meses:
                continue
            bucket = cr_meses if r.kind == "receita" else cp_meses
            # settled pode vir como bool (PG) ou 0/1 (SQLite).
            is_settled = bool(r.settled)
            if is_settled:
                bucket[m]["liquidado"] += float(r.total or 0)
            else:
                bucket[m]["aberto"] += float(r.total or 0)
            bucket[m]["vencido"] += float(r.total_vencido or 0)

        def _build_contas(meses_map):
            linhas, tot = [], {"liquidado": 0.0, "aberto": 0.0, "vencido": 0.0, "total": 0.0}
            for m in range(1, 13):
                liq = meses_map[m]["liquidado"]; ab = meses_map[m]["aberto"]; venc = meses_map[m]["vencido"]
                tot_mes = liq + ab
                if tot_mes == 0 and venc == 0:
                    continue
                linhas.append({"mes": m, "liquidado": liq, "aberto": ab,
                               "vencido": venc, "total": tot_mes})
                tot["liquidado"] += liq; tot["aberto"] += ab
                tot["vencido"] += venc; tot["total"] += tot_mes
            return {"linhas": linhas, "tot": tot}
        contas_receber = _build_contas(cr_meses)
        contas_pagar = _build_contas(cp_meses)

        # (a3) Parcelamento em aberto — usa o campo 'parcela' ('X de Y') já armazenado.
        # Honesto: só conta lançamentos parcelados ainda não liquidados (futuro a entrar/sair).
        parc_aberto_rows = db.execute(
            _sql_text(f"""
                SELECT kind, COUNT(*) AS qtd, SUM(valor) AS total
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause} AND {cyexpr} = :ano
                  AND data_prevista IS NOT NULL
                  AND COALESCE(NULLIF(parcela,''), '') <> ''
                  AND {"NOT settled" if is_pg else "settled = 0"}
                GROUP BY kind
            """),
            {"org_id": org_id, "ano": ano},
        ).fetchall()
        parcelamento = {"receita": {"qtd": 0, "total": 0.0}, "despesa": {"qtd": 0, "total": 0.0}}
        for r in parc_aberto_rows:
            k = "receita" if r.kind == "receita" else "despesa"
            parcelamento[k] = {"qtd": int(r.qtd or 0), "total": float(r.total or 0)}

        # (b) Receita por tipo + PF vs PJ + top 8 clientes.
        receita_tipo = [
            {"label": (r.tipo or "Sem categoria"), "valor": float(r.total or 0)}
            for r in db.execute(_sql_text(f"""
                SELECT COALESCE(NULLIF(tipo,''),'Sem categoria') AS tipo, SUM(valor) AS total
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause} AND {yexpr} = :ano AND kind = 'receita'
                GROUP BY COALESCE(NULLIF(tipo,''),'Sem categoria') ORDER BY total DESC
            """), params).fetchall()
        ]
        pf_pj = [
            {"label": (r.tc or "N/D"), "valor": float(r.total or 0)}
            for r in db.execute(_sql_text(f"""
                SELECT COALESCE(NULLIF(tipo_cliente,''),'N/D') AS tc, SUM(valor) AS total
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause} AND {yexpr} = :ano AND kind = 'receita'
                GROUP BY COALESCE(NULLIF(tipo_cliente,''),'N/D') ORDER BY total DESC
            """), params).fetchall()
        ]
        top_clientes = [
            {"label": (r.cli or "Sem cliente"), "valor": float(r.total or 0)}
            for r in db.execute(_sql_text(f"""
                SELECT COALESCE(NULLIF(cliente,''),'Sem cliente') AS cli, SUM(valor) AS total
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause} AND {yexpr} = :ano AND kind = 'receita'
                GROUP BY COALESCE(NULLIF(cliente,''),'Sem cliente') ORDER BY total DESC LIMIT 8
            """), params).fetchall()
        ]

        # (c) Despesa por categoria (top).
        despesa_categoria = [
            {"label": (r.tipo or "Sem categoria"), "valor": float(r.total or 0)}
            for r in db.execute(_sql_text(f"""
                SELECT COALESCE(NULLIF(tipo,''),'Sem categoria') AS tipo, SUM(valor) AS total
                FROM financial_entries
                WHERE org_id = :org_id AND {active_clause} AND {yexpr} = :ano AND kind = 'despesa'
                GROUP BY COALESCE(NULLIF(tipo,''),'Sem categoria') ORDER BY total DESC LIMIT 10
            """), params).fetchall()
        ]

        # (d) CRUZAMENTO produtividade × financeiro — por PERÍODO (mês), que é o vínculo
        # CONFIÁVEL. processo_ref da planilha é texto livre e NÃO casa 1-1 com o CNJ dos
        # prazos; por isso NÃO inventamos vínculo processo→receita. Cruzamos por mês.
        pmexpr = ("EXTRACT(MONTH FROM COALESCE(data_conclusao, data_vencimento))::int"
                  if is_pg else
                  "CAST(strftime('%m', COALESCE(data_conclusao, data_vencimento)) AS INTEGER)")
        pyexpr = ("EXTRACT(YEAR FROM COALESCE(data_conclusao, data_vencimento))::int"
                  if is_pg else
                  "CAST(strftime('%Y', COALESCE(data_conclusao, data_vencimento)) AS INTEGER)")
        prazos_mes_rows = db.execute(
            _sql_text(f"""
                SELECT {pmexpr} AS mes, COUNT(*) AS qtd
                FROM prazos_processuais
                WHERE org_id = :org_id AND status = 'concluido' AND {pyexpr} = :ano
                GROUP BY {pmexpr}
            """),
            params,
        ).fetchall()
        prazos_por_mes = {int(r.mes or 0): int(r.qtd or 0) for r in prazos_mes_rows if r.mes}
        cruzamento = []
        for x in resumo_mensal:
            concl = prazos_por_mes.get(x["mes"], 0)
            cruzamento.append({
                "mes": x["mes"], "resultado": x["resultado"], "receita": x["receita"],
                "concluidos": concl,
                "resultado_por_prazo": (x["resultado"] / concl) if concl else None,
            })
        tot_concluidos = sum(prazos_por_mes.values())

        # Produtividade por responsável no ano (accountability).
        prazos_resp = [
            {"nome": (r.nome or "Sem responsável"), "qtd": int(r.qtd or 0)}
            for r in db.execute(_sql_text(f"""
                SELECT COALESCE(NULLIF(responsavel,''),'Sem responsável') AS nome, COUNT(*) AS qtd
                FROM prazos_processuais
                WHERE org_id = :org_id AND status = 'concluido' AND {pyexpr} = :ano
                GROUP BY COALESCE(NULLIF(responsavel,''),'Sem responsável')
                ORDER BY qtd DESC LIMIT 8
            """), params).fetchall()
        ]

        # ── (e) Lançamentos editáveis (tabela paginada, org-scoped) ─────────────
        # Editar dado financeiro real exige ver a linha. Paginado p/ não puxar as
        # ~4000 linhas. INCLUI soft-deleted (ativo=FALSE) p/ permitir restaurar — a
        # linha vem marcada `ativo` e a UI a exibe esmaecida. Filtra pelo ano
        # selecionado (mesma competência das agregações). org_id SEMPRE no WHERE.
        per_page = 50
        try:
            page = max(1, int(page))
        except (TypeError, ValueError):
            page = 1
        sort = (sort or "previsto").strip().lower()
        direction = "desc" if (direction or "").strip().lower() == "desc" else "asc"
        sort_exprs = {
            "tipo": "kind",
            "categoria": "COALESCE(tipo,'')",
            "cliente": "COALESCE(cliente,'')",
            "pf_pj": "COALESCE(tipo_cliente,'')",
            "descricao": "COALESCE(descricao,'')",
            "previsto": "data_prevista",
            "efetivo": "data_efetiva",
            "liquidado": "settled",
            "valor": "valor",
        }
        if sort not in sort_exprs:
            sort = "previsto"
        sort_expr = sort_exprs[sort]
        order_dir = "DESC" if direction == "desc" else "ASC"
        if is_pg:
            order_sql = f"{sort_expr} {order_dir} NULLS LAST, id {order_dir}"
        else:
            order_sql = f"CASE WHEN {sort_expr} IS NULL THEN 1 ELSE 0 END, {sort_expr} {order_dir}, id {order_dir}"
        lanc_params = {"org_id": org_id, "ano": ano,
                       "limit": per_page, "offset": (page - 1) * per_page}
        lanc_total = int(db.execute(
            _sql_text(f"""
                SELECT COUNT(*) FROM financial_entries
                WHERE org_id = :org_id AND {yexpr} = :ano
            """),
            {"org_id": org_id, "ano": ano},
        ).scalar() or 0)
        lanc_rows = db.execute(
            _sql_text(f"""
                SELECT id, kind, valor, data_prevista, data_efetiva, settled,
                       tipo, descricao, processo_ref, cliente, tipo_cliente, parcela,
                       COALESCE(ativo, {'TRUE' if is_pg else '1'}) AS ativo
                FROM financial_entries
                WHERE org_id = :org_id AND {yexpr} = :ano
                ORDER BY {order_sql}
                LIMIT :limit OFFSET :offset
            """) if is_pg else _sql_text(f"""
                SELECT id, kind, valor, data_prevista, data_efetiva, settled,
                       tipo, descricao, processo_ref, cliente, tipo_cliente, parcela,
                       COALESCE(ativo, 1) AS ativo
                FROM financial_entries
                WHERE org_id = :org_id AND {yexpr} = :ano
                ORDER BY {order_sql}
                LIMIT :limit OFFSET :offset
            """),
            lanc_params,
        ).fetchall()
        lancamentos = []
        for r in lanc_rows:
            def _d(v):
                return v.isoformat() if hasattr(v, "isoformat") else (str(v) if v else "")
            lancamentos.append({
                "id": int(r.id),
                "kind": r.kind or "",
                "valor": float(r.valor or 0),
                "data_prevista": _d(r.data_prevista),
                "data_efetiva": _d(r.data_efetiva),
                "settled": bool(r.settled),
                "tipo": r.tipo or "",
                "descricao": r.descricao or "",
                "processo_ref": r.processo_ref or "",
                "cliente": r.cliente or "",
                "tipo_cliente": r.tipo_cliente or "",
                "parcela": r.parcela or "",
                "ativo": bool(r.ativo),
            })
        lanc_pages = max(1, (lanc_total + per_page - 1) // per_page)

        ctx = {
            **get_context(request, db, user=user),
            "user": user,
            "today": date.today(),
            "active_module": "reports",
            "ano": ano,
            "anos_disponiveis": anos_disponiveis,
            "resumo_mensal": resumo_mensal,
            "tot_receita": tot_receita,
            "tot_despesa": tot_despesa,
            "tot_resultado": tot_receita - tot_despesa,
            "tot_receita_liq": tot_receita_liq,
            "tot_despesa_liq": tot_despesa_liq,
            "a_receber": a_receber,
            "a_pagar": a_pagar,
            "contas_receber": contas_receber,
            "contas_pagar": contas_pagar,
            "parcelamento": parcelamento,
            "receita_tipo": receita_tipo,
            "pf_pj": pf_pj,
            "top_clientes": top_clientes,
            "despesa_categoria": despesa_categoria,
            "cruzamento": cruzamento,
            "prazos_resp": prazos_resp,
            "tot_concluidos": tot_concluidos,
            "resultado_por_prazo_ano": ((tot_receita - tot_despesa) / tot_concluidos) if tot_concluidos else None,
            # Edição inline + export. Esta página inteira já passou pelo mesmo gate
            # dos endpoints de mutação, então quem a vê pode editar.
            "can_edit_financeiro": True,
            "lancamentos": lancamentos,
            "lanc_total": lanc_total,
            "lanc_page": page,
            "lanc_pages": lanc_pages,
            "lanc_per_page": per_page,
            "lanc_sort": sort,
            "lanc_direction": direction,
        }
        return templates.TemplateResponse("app/reports/financeiro.html", ctx)

    # ─────────── Financeiro · Export .xlsx (gated, org-scoped, read-only) ──────────
    @app.get(f"{PREFIX}/reports/financeiro/export.xlsx")
    async def reports_financeiro_export_canon(
        request: Request,
        ano: Optional[int] = None,
        db: Session = Depends(get_db),
    ):
        import io as _io
        from sqlalchemy import text as _sql_text
        from fastapi.responses import JSONResponse as _JSONResponse, StreamingResponse as _StreamingResponse
        from services.audit import log_action as _log_action

        _gate = _require_financeiro_access(request, db)
        if not isinstance(_gate, tuple):
            return _gate
        user, org_id = _gate

        is_pg = (db.get_bind().dialect.name == "postgresql") if db.get_bind() is not None else False
        active_clause = "ativo IS NOT FALSE" if is_pg else "(ativo IS NULL OR ativo <> 0)"
        yexpr = ("EXTRACT(YEAR FROM COALESCE(data_efetiva, data_prevista))::int"
                 if is_pg else
                 "CAST(strftime('%Y', COALESCE(data_efetiva, data_prevista)) AS INTEGER)")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return _JSONResponse(
                {"error": "openpyxl nao instalado. Execute: pip install openpyxl"},
                status_code=500,
            )

        # Filtro de ano opcional. org_id SEMPRE no WHERE; soft-deleted excluídos do
        # export (espelha as agregações). Bind params — nunca concatena input.
        where = "org_id = :org_id AND " + active_clause
        params = {"org_id": org_id}
        if ano is not None:
            try:
                params["ano"] = int(ano)
                where += f" AND {yexpr} = :ano"
            except (TypeError, ValueError):
                ano = None

        rows = db.execute(
            _sql_text(f"""
                SELECT kind, valor, data_prevista, data_efetiva, settled,
                       tipo, descricao, processo_ref, cliente, tipo_cliente, parcela
                FROM financial_entries
                WHERE {where}
                ORDER BY kind, COALESCE(data_efetiva, data_prevista)
            """),
            params,
        ).fetchall()

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E4890", end_color="1E4890", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ["Tipo", "Categoria", "Cliente", "PF/PJ", "Processo",
                   "Descrição", "Previsto", "Efetivo", "Liquidado", "Parcela", "Valor (R$)"]

        def _fmt_date(v):
            if v is None:
                return ""
            if hasattr(v, "strftime"):
                return v.strftime("%d/%m/%Y")
            return str(v)

        sheets = {
            "receita": wb.active,
            "despesa": wb.create_sheet(title="Despesas"),
        }
        sheets["receita"].title = "Receitas"

        for ws in sheets.values():
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_alignment
                c.border = thin_border
            for col, w in enumerate([12, 22, 28, 8, 18, 40, 12, 12, 12, 10, 16], 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        next_row = {"receita": 2, "despesa": 2}
        for r in rows:
            kind = "receita" if r.kind == "receita" else "despesa"
            ws = sheets[kind]
            ri = next_row[kind]
            values = [
                "Receita" if kind == "receita" else "Despesa",
                r.tipo or "",
                r.cliente or "",
                r.tipo_cliente or "",
                r.processo_ref or "",
                r.descricao or "",
                _fmt_date(r.data_prevista),
                _fmt_date(r.data_efetiva),
                "Sim" if r.settled else "Não",
                r.parcela or "",
                float(r.valor or 0),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=ri, column=col, value=v)
                cell.border = thin_border
                if col == 11:
                    cell.number_format = '#,##0.00'
            next_row[kind] = ri + 1

        output = _io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Audit: registra o export (sem despejar PII — só contagem e escopo).
        _log_action(
            db=db, action="export", entity_type="financial_entries",
            user_id=user.id, user_email=getattr(user, "email", None),
            description=f"Exportou financeiro (.xlsx) org={org_id} ano={ano}",
            details={"rows": len(rows), "ano": ano}, request=request,
        )

        ano_str = str(ano) if ano else date.today().year
        filename = f"financeiro_{ano_str}.xlsx"
        return _StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ─────────── Financeiro · Inline update (gated, whitelist, org-scoped) ─────────
    @app.post(f"{PREFIX}/reports/financeiro/items")
    async def reports_financeiro_create_canon(
        request: Request,
        db: Session = Depends(get_db),
    ):
        from decimal import Decimal, InvalidOperation
        from datetime import datetime as _dt
        from sqlalchemy import text as _sql_text
        from fastapi.responses import JSONResponse as _JSONResponse
        from services.audit import log_action as _log_action

        _gate = _require_financeiro_access(request, db)
        if not isinstance(_gate, tuple):
            return _gate
        user, org_id = _gate

        try:
            data = await request.json()
        except Exception:
            return _JSONResponse({"error": "JSON invalido"}, status_code=400)

        def _txt(key: str, limit: int = 4000):
            value = data.get(key)
            if value is None:
                return None
            value = str(value).strip()
            return (value[:limit] if value else None)

        def _date_value(key: str):
            value = data.get(key)
            if value in (None, ""):
                return None
            return _dt.strptime(str(value).strip(), "%Y-%m-%d").date()

        try:
            kind = str(data.get("kind") or "").strip().lower()
            if kind not in ("receita", "despesa"):
                return _JSONResponse({"error": "Tipo deve ser receita ou despesa"}, status_code=400)

            raw_valor = str(data.get("valor") or "").strip().replace(" ", "")
            if not raw_valor:
                return _JSONResponse({"error": "Valor obrigatorio"}, status_code=400)
            if "," in raw_valor and "." in raw_valor:
                raw_valor = raw_valor.replace(".", "").replace(",", ".")
            elif "," in raw_valor:
                raw_valor = raw_valor.replace(",", ".")
            valor = Decimal(raw_valor).quantize(Decimal("0.01"))
            if valor < 0:
                return _JSONResponse({"error": "Valor nao pode ser negativo"}, status_code=400)

            data_prevista = _date_value("data_prevista")
            data_efetiva = _date_value("data_efetiva")
            tipo_cliente = (_txt("tipo_cliente", 10) or "").upper()
            if tipo_cliente not in ("", "PF", "PJ"):
                return _JSONResponse({"error": "PF/PJ deve ser PF, PJ ou vazio"}, status_code=400)
            tipo_cliente = tipo_cliente or None
            settled = str(data.get("settled") or "").strip().lower() in ("1", "true", "sim", "on", "yes", "t")
        except (InvalidOperation, TypeError, ValueError):
            return _JSONResponse({"error": "Dados invalidos para o lancamento"}, status_code=400)

        params = {
            "org_id": org_id,
            "kind": kind,
            "valor": valor,
            "data_prevista": data_prevista,
            "data_efetiva": data_efetiva,
            "settled": settled,
            "tipo": _txt("tipo", 120),
            "descricao": _txt("descricao"),
            "processo_ref": _txt("processo_ref", 160),
            "cliente": _txt("cliente", 200),
            "tipo_cliente": tipo_cliente,
            "parcela": _txt("parcela", 20),
            "source": "manual",
        }
        is_pg = (db.get_bind().dialect.name == "postgresql") if db.get_bind() is not None else False

        try:
            if is_pg:
                entry_id = db.execute(
                    _sql_text("""
                        INSERT INTO financial_entries
                            (org_id, kind, valor, data_prevista, data_efetiva, settled,
                             tipo, descricao, processo_ref, cliente, tipo_cliente, parcela, source)
                        VALUES
                            (:org_id, :kind, :valor, :data_prevista, :data_efetiva, :settled,
                             :tipo, :descricao, :processo_ref, :cliente, :tipo_cliente, :parcela, :source)
                        RETURNING id
                    """),
                    params,
                ).scalar()
            else:
                db.execute(
                    _sql_text("""
                        INSERT INTO financial_entries
                            (org_id, kind, valor, data_prevista, data_efetiva, settled,
                             tipo, descricao, processo_ref, cliente, tipo_cliente, parcela, source)
                        VALUES
                            (:org_id, :kind, :valor, :data_prevista, :data_efetiva, :settled,
                             :tipo, :descricao, :processo_ref, :cliente, :tipo_cliente, :parcela, :source)
                    """),
                    params,
                )
                entry_id = db.execute(_sql_text("SELECT last_insert_rowid()")).scalar()
            db.commit()
        except Exception as e:
            db.rollback()
            return _JSONResponse({"error": f"Erro ao criar lancamento: {str(e)}"}, status_code=500)

        item = {
            "id": int(entry_id),
            "kind": kind,
            "valor": float(valor),
            "data_prevista": data_prevista.isoformat() if data_prevista else "",
            "data_efetiva": data_efetiva.isoformat() if data_efetiva else "",
            "settled": settled,
            "tipo": params["tipo"] or "",
            "descricao": params["descricao"] or "",
            "processo_ref": params["processo_ref"] or "",
            "cliente": params["cliente"] or "",
            "tipo_cliente": tipo_cliente or "",
            "parcela": params["parcela"] or "",
            "ativo": True,
        }
        _log_action(
            db=db, action="create", entity_type="financial_entries", entity_id=int(entry_id),
            user_id=user.id, user_email=getattr(user, "email", None),
            description=f"Criou lançamento financeiro #{entry_id} org={org_id}",
            details={"fields": [k for k, v in params.items() if k not in ("org_id", "source") and v not in (None, "")]},
            request=request,
        )
        return _JSONResponse({"success": True, "item": item})

    @app.post(f"{PREFIX}/reports/financeiro/{{entry_id}}/update")
    async def reports_financeiro_update_canon(
        entry_id: int,
        request: Request,
        db: Session = Depends(get_db),
    ):
        from sqlalchemy import text as _sql_text
        from fastapi.responses import JSONResponse as _JSONResponse
        from services.audit import log_action as _log_action

        _gate = _require_financeiro_access(request, db)
        if not isinstance(_gate, tuple):
            return _gate
        user, org_id = _gate

        try:
            data = await request.json()
        except Exception:
            return _JSONResponse({"error": "JSON invalido"}, status_code=400)

        field = data.get("field")
        value = data.get("value")

        # Whitelist: nome lógico → coluna física real (schema 2026-06-03_financial_entries).
        # 'org_id', 'id', 'ativo', 'source', 'created_at' NUNCA editáveis por aqui.
        allowed = {
            "descricao": "descricao",
            "valor": "valor",
            "data_prevista": "data_prevista",
            "data_efetiva": "data_efetiva",
            "settled": "settled",
            "tipo": "tipo",
            "kind": "kind",
            "cliente": "cliente",
            "tipo_cliente": "tipo_cliente",
            "processo_ref": "processo_ref",
            "parcela": "parcela",
        }
        if field not in allowed:
            return _JSONResponse(
                {"error": f"Campo '{field}' nao editavel. Permitidos: {', '.join(allowed.keys())}"},
                status_code=400,
            )
        col = allowed[field]  # vem da whitelist, NUNCA do body diretamente.

        # ── Validação/normalização de tipos por campo ──
        try:
            if field == "valor":
                if value in (None, ""):
                    return _JSONResponse({"error": "Valor obrigatorio"}, status_code=400)
                # Aceita "1.234,56" (BR) ou "1234.56"; normaliza p/ float.
                s = str(value).strip().replace(" ", "")
                if "," in s and "." in s:
                    s = s.replace(".", "").replace(",", ".")
                elif "," in s:
                    s = s.replace(",", ".")
                value = round(float(s), 2)
                if value < 0:
                    return _JSONResponse({"error": "Valor nao pode ser negativo"}, status_code=400)
            elif field in ("data_prevista", "data_efetiva"):
                if value in (None, ""):
                    value = None
                else:
                    from datetime import datetime as _dt
                    # ISO (yyyy-mm-dd) do <input type=date>.
                    value = _dt.strptime(str(value).strip(), "%Y-%m-%d").date()
            elif field == "settled":
                value = str(value).strip().lower() in ("1", "true", "sim", "on", "yes", "t")
            elif field == "kind":
                v = str(value).strip().lower()
                if v not in ("receita", "despesa"):
                    return _JSONResponse({"error": "kind deve ser 'receita' ou 'despesa'"}, status_code=400)
                value = v
            elif field == "tipo_cliente":
                v = str(value or "").strip().upper()
                if v not in ("", "PF", "PJ"):
                    return _JSONResponse({"error": "tipo_cliente deve ser PF, PJ ou vazio"}, status_code=400)
                value = v or None
            else:
                # Campos texto: trim + cap defensivo de tamanho.
                value = (str(value).strip() if value is not None else None)
                if value is not None and len(value) > 4000:
                    value = value[:4000]
        except (TypeError, ValueError):
            return _JSONResponse({"error": f"Valor invalido para {field}"}, status_code=400)

        # updated_at é additive (migração 2026-06-04). Pode não existir num banco
        # antigo → checa antes de incluir no SET (dialect-aware).
        is_pg = (db.get_bind().dialect.name == "postgresql") if db.get_bind() is not None else False
        _has_upd = False
        try:
            if is_pg:
                _has_upd = db.execute(_sql_text(
                    "SELECT 1 FROM information_schema.columns "
                    "WHERE table_name='financial_entries' AND column_name='updated_at'"
                )).first() is not None
            else:
                _cols = db.execute(_sql_text("PRAGMA table_info(financial_entries)")).fetchall()
                _has_upd = any((c[1] == "updated_at") for c in _cols)
        except Exception:
            _has_upd = False
        set_upd = ", updated_at = CURRENT_TIMESTAMP" if _has_upd else ""

        try:
            # org_id SEMPRE no WHERE → sem cross-tenant. Coluna vem da whitelist;
            # valor é bind param (nunca concatenado).
            result = db.execute(
                _sql_text(
                    f"UPDATE financial_entries SET {col} = :value{set_upd} "
                    "WHERE id = :id AND org_id = :org_id"
                ),
                {"value": value, "id": entry_id, "org_id": org_id},
            )
            db.commit()
        except Exception as e:
            db.rollback()
            return _JSONResponse({"error": f"Erro ao salvar: {str(e)}"}, status_code=500)

        if result.rowcount == 0:
            # Não existe OU pertence a outra org → 404 (não vaza existência cross-tenant).
            return _JSONResponse({"error": "Lancamento nao encontrado"}, status_code=404)

        # Audit: registra campo alterado, SEM despejar valores sensíveis em massa.
        _log_action(
            db=db, action="update", entity_type="financial_entries", entity_id=entry_id,
            user_id=user.id, user_email=getattr(user, "email", None),
            description=f"Editou lançamento financeiro #{entry_id} (campo {field}) org={org_id}",
            details={"field": field}, request=request,
        )
        # Retorna o valor normalizado p/ a UI re-render sem reload.
        out = value.isoformat() if hasattr(value, "isoformat") else value
        return _JSONResponse({"success": True, "field": field, "value": out})

    # ─────────── Financeiro · Soft-delete / restore (gated, org-scoped) ────────────
    async def _financeiro_set_ativo(entry_id: int, ativo: bool, request, db):
        from sqlalchemy import text as _sql_text
        from fastapi.responses import JSONResponse as _JSONResponse
        from services.audit import log_action as _log_action

        _gate = _require_financeiro_access(request, db)
        if not isinstance(_gate, tuple):
            return _gate
        user, org_id = _gate

        try:
            # Soft-delete: NUNCA DELETE em dado financeiro real. org_id no WHERE.
            result = db.execute(
                _sql_text(
                    "UPDATE financial_entries SET ativo = :ativo "
                    "WHERE id = :id AND org_id = :org_id"
                ),
                {"ativo": ativo, "id": entry_id, "org_id": org_id},
            )
            db.commit()
        except Exception as e:
            db.rollback()
            return _JSONResponse({"error": f"Erro ao salvar: {str(e)}"}, status_code=500)

        if result.rowcount == 0:
            return _JSONResponse({"error": "Lancamento nao encontrado"}, status_code=404)

        _log_action(
            db=db,
            action="soft_delete" if not ativo else "restore",
            entity_type="financial_entries", entity_id=entry_id,
            user_id=user.id, user_email=getattr(user, "email", None),
            description=(f"{'Arquivou' if not ativo else 'Restaurou'} lançamento "
                         f"financeiro #{entry_id} org={org_id}"),
            request=request,
        )
        return _JSONResponse({"success": True, "id": entry_id, "ativo": ativo})

    @app.post(f"{PREFIX}/reports/financeiro/{{entry_id}}/delete")
    async def reports_financeiro_delete_canon(
        entry_id: int, request: Request, db: Session = Depends(get_db),
    ):
        return await _financeiro_set_ativo(entry_id, False, request, db)

    @app.post(f"{PREFIX}/reports/financeiro/{{entry_id}}/restore")
    async def reports_financeiro_restore_canon(
        entry_id: int, request: Request, db: Session = Depends(get_db),
    ):
        return await _financeiro_set_ativo(entry_id, True, request, db)
