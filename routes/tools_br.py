"""
CaseHub Lite - Brazilian Legal Tools
Calculators and utilities for Brazilian law firms.

Routes:
    GET  /tools                  — Tools landing page
    GET  /tools/rescisao         — Rescisão calculator form
    POST /tools/rescisao/calcular — Calculate rescisão and return results
    POST /tools/rescisao/pdf     — Generate printable HTML/PDF for rescisão
    GET  /tools/export/prazos    — Export prazos to Excel/CSV
    GET  /tools/export/processos — Export processos to Excel/CSV
    GET  /tools/export/clientes  — Export clientes to Excel/CSV

    Previdenciário:
    GET  /tools/tempo-contribuicao         — Simulação de tempo de contribuição faltante
    POST /tools/tempo-contribuicao/calcular
    GET  /tools/bpc-loas                   — BPC/LOAS análise
    POST /tools/bpc-loas/calcular
    GET  /tools/aposentadoria-idade        — Aposentadoria por idade
    POST /tools/aposentadoria-idade/calcular
    GET  /tools/pensao-morte               — Pensão por morte
    POST /tools/pensao-morte/calcular
"""
from fastapi import APIRouter, Depends, Request, HTTPException, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging
import math
import asyncio
import io
import csv
import json

logger = logging.getLogger(__name__)

from auth import get_current_user
from models import get_db, Client, Case
from models.tenant import tenant_query
from i18n import get_translations
from core.template_config import templates, PREFIX

router = APIRouter(prefix="/tools", tags=["tools_br"])


def get_context(request: Request, db: Session, **kwargs):
    """Build template context."""
    lang = request.cookies.get("lang", "pt")
    user = get_current_user(request, db)
    return {
        "request": request,
        "PREFIX": PREFIX,
        "lang": lang,
        "t": get_translations(lang),
        "user": user,
        **kwargs,
    }


# ---------------------------------------------------------------------------
# Rescisão Calculation Logic
# ---------------------------------------------------------------------------
def calcular_rescisao(
    salario: float,
    data_admissao: date,
    data_demissao: date,
    tipo_rescisao: str,
    horas_extras: float = 0,
    adicional_noturno: float = 0,
    comissoes: float = 0,
    insalubridade: float = 0,
    periculosidade: float = 0,
    vale_transporte: float = 0,
    vale_refeicao: float = 0,
    ferias_vencidas_dias: int = 0,
    aviso_previo_trabalhado: bool = False,
) -> dict:
    """
    Calculate Brazilian CLT rescisão trabalhista.

    Fundamentação legal:
    - CLT Art. 477 (prazo de pagamento)
    - CLT Art. 484-A (rescisão por acordo)
    - CLT Art. 487 (aviso prévio proporcional)
    - Lei 8.036/90 (FGTS)
    - Lei 12.506/2011 (aviso prévio proporcional)
    """
    # Tempo de serviço
    dias_totais = (data_demissao - data_admissao).days
    meses_trabalhados = dias_totais / 30
    anos = int(meses_trabalhados / 12)
    meses_frac = int(meses_trabalhados % 12)

    # Saldo de salário (dias trabalhados no mês da demissão)
    dias_mes = data_demissao.day
    saldo_salario = (salario / 30) * dias_mes

    # Aviso prévio proporcional (Lei 12.506/2011: 30 dias + 3 por ano, max 90)
    dias_aviso = min(30 + (anos * 3), 90)
    aviso_previo = 0.0 if aviso_previo_trabalhado else (salario / 30) * dias_aviso

    # 13º proporcional (meses trabalhados no ano corrente)
    meses_13 = data_demissao.month
    decimo_terceiro = (salario / 12) * meses_13

    # Férias proporcionais + 1/3 constitucional
    ferias_prop = (salario / 12) * meses_frac
    terco_ferias = ferias_prop / 3

    # Férias vencidas + 1/3
    ferias_vencidas = (salario / 30) * ferias_vencidas_dias if ferias_vencidas_dias > 0 else 0.0
    terco_vencidas = ferias_vencidas / 3

    # FGTS
    remuneracao_total = salario + horas_extras + adicional_noturno + comissoes + insalubridade + periculosidade
    fgts_mensal = remuneracao_total * 0.08
    fgts_saldo = fgts_mensal * meses_trabalhados

    # Multa FGTS e ajustes por tipo de rescisão
    multa_fgts = 0.0

    if tipo_rescisao == "sem_justa_causa":
        multa_fgts = fgts_saldo * 0.40

    elif tipo_rescisao == "acordo":
        # Art. 484-A CLT: 50% do aviso prévio, 20% multa FGTS
        multa_fgts = fgts_saldo * 0.20
        aviso_previo = aviso_previo * 0.5

    elif tipo_rescisao == "justa_causa":
        multa_fgts = 0.0
        aviso_previo = 0.0
        decimo_terceiro = 0.0
        ferias_prop = 0.0
        terco_ferias = 0.0

    elif tipo_rescisao == "pedido_demissao":
        multa_fgts = 0.0
        aviso_previo = 0.0  # empregado deve cumprir ou ter descontado

    elif tipo_rescisao == "termino_contrato":
        multa_fgts = 0.0
        aviso_previo = 0.0

    # Descontos
    descontos = vale_transporte + vale_refeicao

    # Adicionais
    adicionais = horas_extras + adicional_noturno + comissoes + insalubridade + periculosidade

    # Totais
    total_bruto = (
        saldo_salario + aviso_previo + decimo_terceiro
        + ferias_prop + terco_ferias
        + ferias_vencidas + terco_vencidas
        + fgts_saldo + multa_fgts + adicionais
    )
    total_liquido = total_bruto - descontos

    # Label do tipo de rescisão
    tipos_label = {
        "sem_justa_causa": "Sem Justa Causa (empregador)",
        "pedido_demissao": "Pedido de Demissão",
        "justa_causa": "Justa Causa",
        "acordo": "Rescisão por Acordo (Art. 484-A CLT)",
        "termino_contrato": "Término de Contrato",
    }

    return {
        "saldo_salario": round(saldo_salario, 2),
        "aviso_previo": round(aviso_previo, 2),
        "dias_aviso": dias_aviso,
        "decimo_terceiro": round(decimo_terceiro, 2),
        "ferias_proporcionais": round(ferias_prop, 2),
        "terco_ferias": round(terco_ferias, 2),
        "ferias_vencidas": round(ferias_vencidas, 2),
        "terco_vencidas": round(terco_vencidas, 2),
        "fgts_saldo": round(fgts_saldo, 2),
        "multa_fgts": round(multa_fgts, 2),
        "adicionais": round(adicionais, 2),
        "descontos": round(descontos, 2),
        "total_bruto": round(total_bruto, 2),
        "total_liquido": round(total_liquido, 2),
        "meses_trabalhados": round(meses_trabalhados, 1),
        "anos_servico": anos,
        "tipo_rescisao": tipo_rescisao,
        "tipo_rescisao_label": tipos_label.get(tipo_rescisao, tipo_rescisao),
        "data_admissao": data_admissao.strftime("%d/%m/%Y"),
        "data_demissao": data_demissao.strftime("%d/%m/%Y"),
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
async def tools_index(request: Request, db: Session = Depends(get_db)):
    """Tools landing page — list of available tools."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/index.html", get_context(request, db))


@router.get("/rescisao", response_class=HTMLResponse)
async def rescisao_form(request: Request, db: Session = Depends(get_db)):
    """Rescisão calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/rescisao.html", get_context(request, db))


@router.post("/rescisao/calcular", response_class=HTMLResponse)
async def rescisao_calcular(
    request: Request,
    db: Session = Depends(get_db),
    nome: str = Form(""),
    data_admissao: str = Form(...),
    data_demissao: str = Form(...),
    salario: float = Form(...),
    tipo_rescisao: str = Form(...),
    horas_extras: float = Form(0),
    adicional_noturno: float = Form(0),
    comissoes: float = Form(0),
    insalubridade: float = Form(0),
    periculosidade: float = Form(0),
    vale_transporte: float = Form(0),
    vale_refeicao: float = Form(0),
    ferias_vencidas_dias: int = Form(0),
    aviso_previo_trabalhado: bool = Form(False),
):
    """Calculate rescisão and return results page."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_admissao = datetime.strptime(data_admissao, "%Y-%m-%d").date()
        dt_demissao = datetime.strptime(data_demissao, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse("app/tools/rescisao.html", {
            **get_context(request, db),
            "error": "Datas inválidas. Use o formato correto.",
        })

    if dt_demissao <= dt_admissao:
        return templates.TemplateResponse("app/tools/rescisao.html", {
            **get_context(request, db),
            "error": "A data de demissão deve ser posterior à data de admissão.",
        })

    resultado = calcular_rescisao(
        salario=salario,
        data_admissao=dt_admissao,
        data_demissao=dt_demissao,
        tipo_rescisao=tipo_rescisao,
        horas_extras=horas_extras,
        adicional_noturno=adicional_noturno,
        comissoes=comissoes,
        insalubridade=insalubridade,
        periculosidade=periculosidade,
        vale_transporte=vale_transporte,
        vale_refeicao=vale_refeicao,
        ferias_vencidas_dias=ferias_vencidas_dias,
        aviso_previo_trabalhado=aviso_previo_trabalhado,
    )

    return templates.TemplateResponse("app/tools/rescisao.html", {
        **get_context(request, db),
        "resultado": resultado,
        "nome": nome,
        # Re-populate form fields
        "form": {
            "nome": nome,
            "data_admissao": data_admissao,
            "data_demissao": data_demissao,
            "salario": salario,
            "tipo_rescisao": tipo_rescisao,
            "horas_extras": horas_extras,
            "adicional_noturno": adicional_noturno,
            "comissoes": comissoes,
            "insalubridade": insalubridade,
            "periculosidade": periculosidade,
            "vale_transporte": vale_transporte,
            "vale_refeicao": vale_refeicao,
            "ferias_vencidas_dias": ferias_vencidas_dias,
            "aviso_previo_trabalhado": aviso_previo_trabalhado,
        },
    })


# ---------------------------------------------------------------------------
# Rescisão PDF Export
# ---------------------------------------------------------------------------

@router.post("/rescisao/pdf", response_class=HTMLResponse)
async def rescisao_pdf(
    request: Request,
    db: Session = Depends(get_db),
    nome: str = Form(""),
    data_admissao: str = Form(...),
    data_demissao: str = Form(...),
    salario: float = Form(...),
    tipo_rescisao: str = Form(...),
    horas_extras: float = Form(0),
    adicional_noturno: float = Form(0),
    comissoes: float = Form(0),
    insalubridade: float = Form(0),
    periculosidade: float = Form(0),
    vale_transporte: float = Form(0),
    vale_refeicao: float = Form(0),
    ferias_vencidas_dias: int = Form(0),
    aviso_previo_trabalhado: bool = Form(False),
):
    """Generate printable HTML for rescisão calculation (browser Print → PDF)."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_admissao = datetime.strptime(data_admissao, "%Y-%m-%d").date()
        dt_demissao = datetime.strptime(data_demissao, "%Y-%m-%d").date()
    except ValueError:
        return HTMLResponse("<h1>Erro: Datas inválidas</h1>", status_code=400)

    if dt_demissao <= dt_admissao:
        return HTMLResponse("<h1>Erro: Data de demissão deve ser posterior à admissão</h1>", status_code=400)

    r = calcular_rescisao(
        salario=salario,
        data_admissao=dt_admissao,
        data_demissao=dt_demissao,
        tipo_rescisao=tipo_rescisao,
        horas_extras=horas_extras,
        adicional_noturno=adicional_noturno,
        comissoes=comissoes,
        insalubridade=insalubridade,
        periculosidade=periculosidade,
        vale_transporte=vale_transporte,
        vale_refeicao=vale_refeicao,
        ferias_vencidas_dias=ferias_vencidas_dias,
        aviso_previo_trabalhado=aviso_previo_trabalhado,
    )

    # Build breakdown rows
    rows = f'<tr><td>Saldo de Salário ({dt_demissao.day} dias)</td><td class="val">R$ {r["saldo_salario"]:,.2f}</td></tr>'
    if r["aviso_previo"] > 0:
        rows += f'<tr><td>Aviso Prévio ({r["dias_aviso"]} dias)</td><td class="val">R$ {r["aviso_previo"]:,.2f}</td></tr>'
    if r["decimo_terceiro"] > 0:
        rows += f'<tr><td>13º Proporcional</td><td class="val">R$ {r["decimo_terceiro"]:,.2f}</td></tr>'
    if r["ferias_proporcionais"] > 0:
        rows += f'<tr><td>Férias Proporcionais</td><td class="val">R$ {r["ferias_proporcionais"]:,.2f}</td></tr>'
        rows += f'<tr><td>1/3 Constitucional (férias)</td><td class="val">R$ {r["terco_ferias"]:,.2f}</td></tr>'
    if r["ferias_vencidas"] > 0:
        rows += f'<tr><td>Férias Vencidas</td><td class="val">R$ {r["ferias_vencidas"]:,.2f}</td></tr>'
        rows += f'<tr><td>1/3 Constitucional (vencidas)</td><td class="val">R$ {r["terco_vencidas"]:,.2f}</td></tr>'
    rows += f'<tr><td>FGTS (saldo estimado)</td><td class="val">R$ {r["fgts_saldo"]:,.2f}</td></tr>'
    if r["multa_fgts"] > 0:
        pct = "20%" if tipo_rescisao == "acordo" else "40%"
        rows += f'<tr><td>Multa FGTS ({pct})</td><td class="val">R$ {r["multa_fgts"]:,.2f}</td></tr>'
    if r["adicionais"] > 0:
        rows += f'<tr><td>Adicionais</td><td class="val">R$ {r["adicionais"]:,.2f}</td></tr>'
    if r["descontos"] > 0:
        rows += f'<tr class="desc"><td>Descontos</td><td class="val">- R$ {r["descontos"]:,.2f}</td></tr>'
    rows += f'<tr class="bruto"><td><strong>TOTAL BRUTO</strong></td><td class="val"><strong>R$ {r["total_bruto"]:,.2f}</strong></td></tr>'
    rows += f'<tr class="liq"><td><strong>TOTAL LÍQUIDO</strong></td><td class="val"><strong>R$ {r["total_liquido"]:,.2f}</strong></td></tr>'

    nome_display = nome or "Não informado"
    today_str = date.today().strftime("%d/%m/%Y")

    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>Rescisão Trabalhista - {nome_display}</title>
<style>
@media print {{ @page {{ margin: 20mm; }} body {{ margin: 0; }} .no-print {{ display: none; }} }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; color: #333; line-height: 1.6; }}
h1 {{ color: #1C2447; font-size: 1.6em; margin-bottom: 4px; }}
h2 {{ color: #1C2447; font-size: 1.1em; margin-top: 24px; border-bottom: 2px solid #1C2447; padding-bottom: 4px; }}
.meta {{ color: #555; font-size: 0.95em; margin-bottom: 20px; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
td {{ padding: 8px 12px; border-bottom: 1px solid #e0e0e0; }}
td.val {{ text-align: right; font-family: 'Courier New', monospace; white-space: nowrap; }}
tr.desc td {{ color: #dc3545; }}
tr.bruto {{ background: #f8f9fa; }}
tr.liq {{ background: #d4edda; font-size: 1.15em; }}
.legal {{ margin-top: 24px; padding: 12px; background: #f8f9fa; border-radius: 6px; font-size: 0.85em; color: #666; }}
.footer {{ margin-top: 30px; font-size: 0.8em; color: #999; border-top: 1px solid #ddd; padding-top: 10px; }}
.print-btn {{ display: inline-block; margin-bottom: 20px; padding: 10px 24px; background: #1C2447; color: #fff; border: none; border-radius: 6px; font-size: 1em; cursor: pointer; }}
.print-btn:hover {{ background: #1a6069; }}
</style></head><body>
<button class="print-btn no-print" onclick="window.print()">Imprimir / Salvar PDF</button>
<h1>Cálculo de Rescisão Trabalhista</h1>
<div class="meta">
<strong>Funcionário:</strong> {nome_display}<br>
<strong>Admissão:</strong> {r['data_admissao']} &nbsp;|&nbsp; <strong>Demissão:</strong> {r['data_demissao']}<br>
<strong>Tempo de Serviço:</strong> {r['meses_trabalhados']} meses ({r['anos_servico']} ano{'s' if r['anos_servico'] != 1 else ''})<br>
<strong>Salário:</strong> R$ {salario:,.2f}<br>
<strong>Tipo:</strong> {r['tipo_rescisao_label']}
</div>
<h2>Discriminação das Verbas</h2>
<table>{rows}</table>
<div class="legal">
<strong>Fundamentação Legal:</strong> CLT Art. 477, CLT Art. 487, Lei 12.506/2011 (aviso prévio proporcional),
Lei 8.036/90 (FGTS), CF Art. 7º XVII (1/3 férias).
{'CLT Art. 484-A (rescisão por acordo).' if tipo_rescisao == 'acordo' else ''}
<br><em>Este cálculo é uma estimativa para fins de orientação. Valores finais podem variar conforme convenção coletiva e outros descontos legais.</em>
</div>
<div class="footer">
Gerado pelo CaseHub Lite em {today_str}
</div>
</body></html>"""

    return HTMLResponse(content=html)


# ---------------------------------------------------------------------------
# 1. Horas Extras + DSR + Reflexos
# ---------------------------------------------------------------------------
# Test cases:
# Input: salário=3000, jornada=44, horas_extras=20, percentual=50, dias_uteis=22, domingos_feriados=8
#   valor_hora = 3000 / (44 * 4.33) = 15.74
#   hora_extra = 15.74 * 1.5 = 23.61
#   total_he = 23.61 * 20 = 472.20
#   dsr = (472.20 / 22) * 8 = 171.71
#   reflexo_13 = (472.20 + 171.71) / 12 = 53.66
#   reflexo_ferias = (472.20 + 171.71) / 12 * 1.3333 = 71.54
#   reflexo_fgts = (472.20 + 171.71) * 0.08 = 51.51

def calcular_horas_extras(
    salario: float,
    jornada_semanal: float,
    horas_extras: float,
    percentual: float,
    dias_uteis: int,
    domingos_feriados: int,
) -> dict:
    """
    Calcula horas extras, DSR e reflexos.
    Fundamentação: CLT Art. 59, 67; Súmula 172 TST (reflexos).
    """
    valor_hora = salario / (jornada_semanal * 4.33)
    valor_he = valor_hora * (1 + percentual / 100)
    total_he = valor_he * horas_extras
    dsr = (total_he / dias_uteis) * domingos_feriados if dias_uteis > 0 else 0
    total_he_dsr = total_he + dsr
    reflexo_13 = total_he_dsr / 12
    reflexo_ferias = (total_he_dsr / 12) * (4 / 3)
    reflexo_fgts = total_he_dsr * 0.08

    return {
        "valor_hora_normal": round(valor_hora, 2),
        "valor_hora_extra": round(valor_he, 2),
        "total_horas_extras": round(total_he, 2),
        "dsr": round(dsr, 2),
        "total_he_dsr": round(total_he_dsr, 2),
        "reflexo_13": round(reflexo_13, 2),
        "reflexo_ferias": round(reflexo_ferias, 2),
        "reflexo_fgts": round(reflexo_fgts, 2),
        "total_geral": round(total_he_dsr + reflexo_13 + reflexo_ferias + reflexo_fgts, 2),
        "salario": salario,
        "jornada_semanal": jornada_semanal,
        "horas_extras": horas_extras,
        "percentual": percentual,
        "dias_uteis": dias_uteis,
        "domingos_feriados": domingos_feriados,
    }


@router.get("/horas-extras", response_class=HTMLResponse)
async def horas_extras_form(request: Request, db: Session = Depends(get_db)):
    """Horas Extras calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/horas_extras.html", get_context(request, db))


@router.post("/horas-extras/calcular", response_class=HTMLResponse)
async def horas_extras_calcular(
    request: Request,
    db: Session = Depends(get_db),
    salario: float = Form(...),
    jornada_semanal: float = Form(44),
    horas_extras: float = Form(...),
    percentual: float = Form(50),
    dias_uteis: int = Form(22),
    domingos_feriados: int = Form(8),
):
    """Calculate horas extras + DSR + reflexos."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_horas_extras(
        salario=salario,
        jornada_semanal=jornada_semanal,
        horas_extras=horas_extras,
        percentual=percentual,
        dias_uteis=dias_uteis,
        domingos_feriados=domingos_feriados,
    )

    return templates.TemplateResponse("app/tools/horas_extras.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "salario": salario,
            "jornada_semanal": jornada_semanal,
            "horas_extras": horas_extras,
            "percentual": percentual,
            "dias_uteis": dias_uteis,
            "domingos_feriados": domingos_feriados,
        },
    })


# ---------------------------------------------------------------------------
# 2. Adicional Noturno
# ---------------------------------------------------------------------------
# Test cases:
# Input: salário=3000, horas_noturnas=40, percentual=20
#   valor_hora = 3000 / 220 = 13.64
#   adicional = 13.64 * 0.20 * 40 = 109.09
#   hora_reduzida: 40h * (60/52.5) = 45.71 horas fictas
#   dsr = (109.09 / 22) * 8 = 39.67
#   reflexo_13 = (109.09 + 39.67) / 12 = 12.40
#   reflexo_ferias = (109.09 + 39.67) / 12 * 1.333 = 16.53

def calcular_adicional_noturno(
    salario: float,
    horas_noturnas: float,
    percentual: float = 20.0,
    dias_uteis: int = 22,
    domingos_feriados: int = 8,
) -> dict:
    """
    Calcula adicional noturno conforme CLT Art. 73.
    Hora noturna reduzida: 52min30s (Art. 73, §1º CLT).
    """
    valor_hora = salario / 220  # 220h = jornada mensal padrão CLT
    horas_fictas = horas_noturnas * (60 / 52.5)  # hora reduzida
    adicional = valor_hora * (percentual / 100) * horas_noturnas
    dsr = (adicional / dias_uteis) * domingos_feriados if dias_uteis > 0 else 0
    total_com_dsr = adicional + dsr
    reflexo_13 = total_com_dsr / 12
    reflexo_ferias = (total_com_dsr / 12) * (4 / 3)
    reflexo_fgts = total_com_dsr * 0.08

    return {
        "valor_hora": round(valor_hora, 2),
        "horas_fictas": round(horas_fictas, 2),
        "adicional": round(adicional, 2),
        "dsr": round(dsr, 2),
        "total_com_dsr": round(total_com_dsr, 2),
        "reflexo_13": round(reflexo_13, 2),
        "reflexo_ferias": round(reflexo_ferias, 2),
        "reflexo_fgts": round(reflexo_fgts, 2),
        "total_geral": round(total_com_dsr + reflexo_13 + reflexo_ferias + reflexo_fgts, 2),
        "salario": salario,
        "horas_noturnas": horas_noturnas,
        "percentual": percentual,
    }


@router.get("/adicional-noturno", response_class=HTMLResponse)
async def adicional_noturno_form(request: Request, db: Session = Depends(get_db)):
    """Adicional Noturno calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/adicional_noturno.html", get_context(request, db))


@router.get("/adicional_noturno")
async def adicional_noturno_legacy_alias():
    """Compatibility alias for older tool links that used underscores."""
    return RedirectResponse(url=f"{PREFIX}/tools/adicional-noturno", status_code=302)


@router.post("/adicional-noturno/calcular", response_class=HTMLResponse)
async def adicional_noturno_calcular(
    request: Request,
    db: Session = Depends(get_db),
    salario: float = Form(...),
    horas_noturnas: float = Form(...),
    percentual: float = Form(20),
    dias_uteis: int = Form(22),
    domingos_feriados: int = Form(8),
):
    """Calculate adicional noturno + reflexos."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_adicional_noturno(
        salario=salario,
        horas_noturnas=horas_noturnas,
        percentual=percentual,
        dias_uteis=dias_uteis,
        domingos_feriados=domingos_feriados,
    )

    return templates.TemplateResponse("app/tools/adicional_noturno.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "salario": salario,
            "horas_noturnas": horas_noturnas,
            "percentual": percentual,
            "dias_uteis": dias_uteis,
            "domingos_feriados": domingos_feriados,
        },
    })


# ---------------------------------------------------------------------------
# 3. Insalubridade / Periculosidade
# ---------------------------------------------------------------------------
# Test cases:
# Insalubridade grau máximo: SM=1518, grau=40 → 1518 * 0.40 = 607.20
# Insalubridade grau médio: SM=1518, grau=20 → 1518 * 0.20 = 303.60
# Insalubridade grau mínimo: SM=1518, grau=10 → 1518 * 0.10 = 151.80
# Periculosidade: salário_base=5000, 30% → 5000 * 0.30 = 1500.00
# Não cumulativo (Art. 193, §2º CLT)

SALARIO_MINIMO_2026 = 1518.00


def calcular_insalubridade_periculosidade(
    tipo: str,
    salario_base: float = 0,
    salario_minimo: float = SALARIO_MINIMO_2026,
    grau_insalubridade: int = 20,
) -> dict:
    """
    Calcula insalubridade ou periculosidade.
    Insalubridade: Art. 192 CLT — base = salário mínimo.
    Periculosidade: Art. 193 CLT — 30% sobre salário base.
    Não cumulativos (Art. 193, §2º CLT).
    """
    if tipo == "insalubridade":
        percentual = grau_insalubridade / 100
        valor = salario_minimo * percentual
        base = salario_minimo
        grau_label = {10: "Mínimo (10%)", 20: "Médio (20%)", 40: "Máximo (40%)"}.get(
            grau_insalubridade, f"{grau_insalubridade}%"
        )
    else:
        percentual = 0.30
        valor = salario_base * percentual
        base = salario_base
        grau_label = "30% sobre salário base"

    reflexo_13 = valor / 12
    reflexo_ferias = (valor / 12) * (4 / 3)
    reflexo_fgts = valor * 0.08

    return {
        "tipo": tipo,
        "tipo_label": "Insalubridade" if tipo == "insalubridade" else "Periculosidade",
        "base_calculo": round(base, 2),
        "grau_label": grau_label,
        "percentual": round(percentual * 100, 1),
        "valor_mensal": round(valor, 2),
        "reflexo_13": round(reflexo_13, 2),
        "reflexo_ferias": round(reflexo_ferias, 2),
        "reflexo_fgts": round(reflexo_fgts, 2),
        "total_anual_estimado": round(valor * 12 + valor + (valor / 3), 2),
        "salario_base": salario_base,
        "salario_minimo": salario_minimo,
    }


@router.get("/insalubridade", response_class=HTMLResponse)
async def insalubridade_form(request: Request, db: Session = Depends(get_db)):
    """Insalubridade/Periculosidade calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/insalubridade.html", get_context(request, db))


@router.post("/insalubridade/calcular", response_class=HTMLResponse)
async def insalubridade_calcular(
    request: Request,
    db: Session = Depends(get_db),
    tipo: str = Form(...),
    salario_base: float = Form(0),
    salario_minimo: float = Form(SALARIO_MINIMO_2026),
    grau_insalubridade: int = Form(20),
):
    """Calculate insalubridade or periculosidade."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_insalubridade_periculosidade(
        tipo=tipo,
        salario_base=salario_base,
        salario_minimo=salario_minimo,
        grau_insalubridade=grau_insalubridade,
    )

    return templates.TemplateResponse("app/tools/insalubridade.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "tipo": tipo,
            "salario_base": salario_base,
            "salario_minimo": salario_minimo,
            "grau_insalubridade": grau_insalubridade,
        },
    })


# ---------------------------------------------------------------------------
# 4. Férias + 1/3
# ---------------------------------------------------------------------------
# Test cases:
# Input: salário=3000, meses=12, faltas=4, abono=False
#   proporcional = (3000 / 12) * 12 = 3000.00
#   terço = 3000 / 3 = 1000.00
#   total = 4000.00 (sem desconto, faltas <= 5 → 30 dias)
# Input: salário=3000, meses=6, faltas=0, abono=True
#   proporcional = (3000 / 12) * 6 = 1500.00
#   terço = 1500 / 3 = 500.00
#   abono_pecuniario = 1500 / 3 = 500.00 (vende 10 dias de 30)
#   terco_abono = 500 / 3 = 166.67
#   total = 1500 + 500 + 500 + 166.67 = 2666.67
# Input: salário=3000, meses=12, faltas=10
#   Art.130: 6-14 faltas → 24 dias de férias
#   proporcional = 3000 * (24/30) = 2400.00, terço = 800.00, total = 3200.00

def calcular_ferias(
    salario: float,
    meses_trabalhados: int,
    faltas: int = 0,
    abono_pecuniario: bool = False,
) -> dict:
    """
    Calcula férias + 1/3 constitucional.
    Art. 130 CLT — tabela de desconto por faltas.
    Art. 143 CLT — abono pecuniário (venda de 1/3).
    CF Art. 7º, XVII — terço constitucional.
    """
    # Art. 130 CLT: dias de férias conforme faltas
    if faltas <= 5:
        dias_ferias = 30
    elif faltas <= 14:
        dias_ferias = 24
    elif faltas <= 23:
        dias_ferias = 18
    elif faltas <= 32:
        dias_ferias = 12
    else:
        dias_ferias = 0  # perde direito a férias

    proporcional = (salario / 12) * meses_trabalhados
    # Ajusta pelo desconto de dias
    proporcional = proporcional * (dias_ferias / 30)
    terco = proporcional / 3

    abono_valor = 0.0
    terco_abono = 0.0
    dias_abono = 0
    if abono_pecuniario and dias_ferias > 0:
        dias_abono = dias_ferias // 3
        abono_valor = (proporcional / dias_ferias) * dias_abono
        terco_abono = abono_valor / 3

    total = proporcional + terco + abono_valor + terco_abono

    return {
        "dias_ferias": dias_ferias,
        "proporcional": round(proporcional, 2),
        "terco": round(terco, 2),
        "abono_pecuniario": round(abono_valor, 2),
        "terco_abono": round(terco_abono, 2),
        "dias_abono": dias_abono,
        "total": round(total, 2),
        "salario": salario,
        "meses_trabalhados": meses_trabalhados,
        "faltas": faltas,
        "usou_abono": abono_pecuniario,
    }


@router.get("/ferias", response_class=HTMLResponse)
async def ferias_form(request: Request, db: Session = Depends(get_db)):
    """Férias calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/ferias.html", get_context(request, db))


@router.post("/ferias/calcular", response_class=HTMLResponse)
async def ferias_calcular(
    request: Request,
    db: Session = Depends(get_db),
    salario: float = Form(...),
    meses_trabalhados: int = Form(...),
    faltas: int = Form(0),
    abono_pecuniario: bool = Form(False),
):
    """Calculate férias + 1/3."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_ferias(
        salario=salario,
        meses_trabalhados=meses_trabalhados,
        faltas=faltas,
        abono_pecuniario=abono_pecuniario,
    )

    return templates.TemplateResponse("app/tools/ferias.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "salario": salario,
            "meses_trabalhados": meses_trabalhados,
            "faltas": faltas,
            "abono_pecuniario": abono_pecuniario,
        },
    })


# ---------------------------------------------------------------------------
# 5. 13º Salário Proporcional
# ---------------------------------------------------------------------------
# Test cases:
# Input: salário=3000, meses=12, adicionais=500
#   valor = (3000 + 500) / 12 * 12 = 3500.00
#   1ª parcela = 3500 / 2 = 1750.00
#   2ª parcela = 3500 / 2 = 1750.00 (antes dos descontos)
# Input: salário=3000, meses=6, adicionais=0
#   valor = 3000 / 12 * 6 = 1500.00
#   1ª parcela = 750.00, 2ª parcela = 750.00

def calcular_decimo_terceiro(
    salario: float,
    meses_trabalhados: int,
    adicionais_habituais: float = 0,
) -> dict:
    """
    Calcula 13º salário proporcional.
    Art. 1º Lei 4.090/62 — gratificação natalina.
    Art. 2º Lei 4.749/65 — pagamento em duas parcelas.
    >= 15 dias no mês = mês cheio.
    """
    base = salario + adicionais_habituais
    valor = (base / 12) * meses_trabalhados
    primeira_parcela = valor / 2  # até 30/nov
    segunda_parcela = valor / 2   # até 20/dez (desconto INSS/IR na 2ª)

    return {
        "base_calculo": round(base, 2),
        "valor_total": round(valor, 2),
        "primeira_parcela": round(primeira_parcela, 2),
        "segunda_parcela": round(segunda_parcela, 2),
        "salario": salario,
        "meses_trabalhados": meses_trabalhados,
        "adicionais_habituais": adicionais_habituais,
    }


@router.get("/decimo-terceiro", response_class=HTMLResponse)
async def decimo_terceiro_form(request: Request, db: Session = Depends(get_db)):
    """13º Salário calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/decimo_terceiro.html", get_context(request, db))


@router.post("/decimo-terceiro/calcular", response_class=HTMLResponse)
async def decimo_terceiro_calcular(
    request: Request,
    db: Session = Depends(get_db),
    salario: float = Form(...),
    meses_trabalhados: int = Form(...),
    adicionais_habituais: float = Form(0),
):
    """Calculate 13º salário proporcional."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_decimo_terceiro(
        salario=salario,
        meses_trabalhados=meses_trabalhados,
        adicionais_habituais=adicionais_habituais,
    )

    return templates.TemplateResponse("app/tools/decimo_terceiro.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "salario": salario,
            "meses_trabalhados": meses_trabalhados,
            "adicionais_habituais": adicionais_habituais,
        },
    })


# ---------------------------------------------------------------------------
# 6. FGTS + Multa 40%
# ---------------------------------------------------------------------------
# Test cases:
# Input: salário=3000, meses=24, saldo_anterior=0, tipo=sem_justa_causa
#   depósito_mensal = 3000 * 0.08 = 240.00
#   saldo = 240 * 24 = 5760.00
#   multa = 5760 * 0.40 = 2304.00
#   total = 5760 + 2304 = 8064.00
# Input: salário=3000, meses=24, saldo_anterior=1000, tipo=acordo
#   saldo = 240 * 24 + 1000 = 6760.00
#   multa = 6760 * 0.20 = 1352.00 (Art. 484-A)
#   total = 6760 + 1352 = 8112.00

def calcular_fgts(
    salario: float,
    meses_trabalhados: int,
    saldo_anterior: float = 0,
    tipo_rescisao: str = "sem_justa_causa",
) -> dict:
    """
    Calcula FGTS + multa rescisória.
    Lei 8.036/90 — depósito mensal de 8%.
    Art. 18 — multa 40% (sem justa causa).
    Art. 484-A CLT — multa 20% (acordo).
    """
    deposito_mensal = salario * 0.08
    saldo_depositado = deposito_mensal * meses_trabalhados
    saldo_total = saldo_depositado + saldo_anterior

    if tipo_rescisao == "sem_justa_causa":
        percentual_multa = 40
        multa = saldo_total * 0.40
    elif tipo_rescisao == "acordo":
        percentual_multa = 20
        multa = saldo_total * 0.20
    else:
        percentual_multa = 0
        multa = 0

    total = saldo_total + multa

    tipo_labels = {
        "sem_justa_causa": "Sem Justa Causa",
        "acordo": "Acordo (Art. 484-A CLT)",
        "pedido_demissao": "Pedido de Demissão",
        "justa_causa": "Justa Causa",
    }

    return {
        "deposito_mensal": round(deposito_mensal, 2),
        "saldo_depositado": round(saldo_depositado, 2),
        "saldo_anterior": round(saldo_anterior, 2),
        "saldo_total": round(saldo_total, 2),
        "percentual_multa": percentual_multa,
        "multa": round(multa, 2),
        "total": round(total, 2),
        "tipo_rescisao": tipo_rescisao,
        "tipo_label": tipo_labels.get(tipo_rescisao, tipo_rescisao),
        "salario": salario,
        "meses_trabalhados": meses_trabalhados,
    }


@router.get("/fgts", response_class=HTMLResponse)
async def fgts_form(request: Request, db: Session = Depends(get_db)):
    """FGTS + Multa calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/fgts.html", get_context(request, db))


@router.post("/fgts/calcular", response_class=HTMLResponse)
async def fgts_calcular(
    request: Request,
    db: Session = Depends(get_db),
    salario: float = Form(...),
    meses_trabalhados: int = Form(...),
    saldo_anterior: float = Form(0),
    tipo_rescisao: str = Form("sem_justa_causa"),
):
    """Calculate FGTS + multa rescisória."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_fgts(
        salario=salario,
        meses_trabalhados=meses_trabalhados,
        saldo_anterior=saldo_anterior,
        tipo_rescisao=tipo_rescisao,
    )

    return templates.TemplateResponse("app/tools/fgts.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "salario": salario,
            "meses_trabalhados": meses_trabalhados,
            "saldo_anterior": saldo_anterior,
            "tipo_rescisao": tipo_rescisao,
        },
    })


# ---------------------------------------------------------------------------
# 7. Seguro-Desemprego
# ---------------------------------------------------------------------------
# Test cases (faixas 2026 — valores aproximados, atualizados anualmente):
# Input: salarios=[1800, 1900, 1700], meses=18, vezes_recebeu=0
#   média = (1800+1900+1700)/3 = 1800.00
#   Faixa 1 (até R$2.041,39): parcela = média * 0.8 = 1440.00
#   1ª vez + 18 meses → 5 parcelas
# Input: salarios=[3000, 3200, 2800], meses=10, vezes_recebeu=1
#   média = 3000.00
#   Faixa 2 (R$2.041,40 a R$3.402,65):
#     parcela = 1633.11 + (3000 - 2041.39) * 0.5 = 1633.11 + 479.31 = 2112.42
#   2ª vez + 10 meses → 4 parcelas

# Faixas do seguro-desemprego 2026 (valores base — atualizados via resolução CODEFAT)
FAIXAS_SD_2026 = [
    {"ate": 2041.39, "fator": 0.8, "adicional": 0},
    {"ate": 3402.65, "fator": 0.5, "base_fixa": 1633.11},
    {"ate": 999999, "fixo": 2313.74},
]
PISO_SD = 1518.00  # salário mínimo
TETO_SD = 2313.74


def calcular_seguro_desemprego(
    salario_1: float,
    salario_2: float,
    salario_3: float,
    meses_trabalhados: int,
    vezes_recebeu: int = 0,
) -> dict:
    """
    Calcula seguro-desemprego conforme Lei 7.998/90, Art. 5.
    Parcelas: 3 a 5 conforme tempo de trabalho.
    Requisitos mínimos: 12m (1ª vez), 9m (2ª vez), 6m (3ª+ vez).
    """
    media = (salario_1 + salario_2 + salario_3) / 3

    # Calcular valor da parcela pelas faixas
    if media <= FAIXAS_SD_2026[0]["ate"]:
        parcela = media * FAIXAS_SD_2026[0]["fator"]
    elif media <= FAIXAS_SD_2026[1]["ate"]:
        parcela = FAIXAS_SD_2026[1]["base_fixa"] + (media - FAIXAS_SD_2026[0]["ate"]) * FAIXAS_SD_2026[1]["fator"]
    else:
        parcela = FAIXAS_SD_2026[2]["fixo"]

    # Piso = salário mínimo
    parcela = max(parcela, PISO_SD)
    # Teto
    parcela = min(parcela, TETO_SD)

    # Requisitos mínimos e número de parcelas
    if vezes_recebeu == 0:
        meses_minimo = 12
        if meses_trabalhados >= 24:
            num_parcelas = 5
        elif meses_trabalhados >= 12:
            num_parcelas = 4
        else:
            num_parcelas = 0
    elif vezes_recebeu == 1:
        meses_minimo = 9
        if meses_trabalhados >= 24:
            num_parcelas = 5
        elif meses_trabalhados >= 12:
            num_parcelas = 4
        elif meses_trabalhados >= 9:
            num_parcelas = 3
        else:
            num_parcelas = 0
    else:
        meses_minimo = 6
        if meses_trabalhados >= 24:
            num_parcelas = 5
        elif meses_trabalhados >= 12:
            num_parcelas = 4
        elif meses_trabalhados >= 6:
            num_parcelas = 3
        else:
            num_parcelas = 0

    tem_direito = meses_trabalhados >= meses_minimo
    total = parcela * num_parcelas if tem_direito else 0

    return {
        "media_salarial": round(media, 2),
        "valor_parcela": round(parcela, 2),
        "num_parcelas": num_parcelas,
        "total": round(total, 2),
        "tem_direito": tem_direito,
        "meses_minimo": meses_minimo,
        "salario_1": salario_1,
        "salario_2": salario_2,
        "salario_3": salario_3,
        "meses_trabalhados": meses_trabalhados,
        "vezes_recebeu": vezes_recebeu,
    }


@router.get("/seguro-desemprego", response_class=HTMLResponse)
async def seguro_desemprego_form(request: Request, db: Session = Depends(get_db)):
    """Seguro-Desemprego calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/seguro_desemprego.html", get_context(request, db))


@router.post("/seguro-desemprego/calcular", response_class=HTMLResponse)
async def seguro_desemprego_calcular(
    request: Request,
    db: Session = Depends(get_db),
    salario_1: float = Form(...),
    salario_2: float = Form(...),
    salario_3: float = Form(...),
    meses_trabalhados: int = Form(...),
    vezes_recebeu: int = Form(0),
):
    """Calculate seguro-desemprego."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_seguro_desemprego(
        salario_1=salario_1,
        salario_2=salario_2,
        salario_3=salario_3,
        meses_trabalhados=meses_trabalhados,
        vezes_recebeu=vezes_recebeu,
    )

    return templates.TemplateResponse("app/tools/seguro_desemprego.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "salario_1": salario_1,
            "salario_2": salario_2,
            "salario_3": salario_3,
            "meses_trabalhados": meses_trabalhados,
            "vezes_recebeu": vezes_recebeu,
        },
    })


# ---------------------------------------------------------------------------
# Excel / CSV Export endpoints
# ---------------------------------------------------------------------------

def _try_openpyxl_export(headers: list, rows: list, sheet_name: str = "Dados") -> tuple:
    """
    Try to export using openpyxl. Returns (BytesIO, content_type, extension).
    Falls back to CSV if openpyxl is not available.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_data in rows:
                if col_idx - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col_idx - 1] or "")))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

    except ImportError:
        logger.info("openpyxl not available, falling back to CSV export")
        return None, None, None


def _csv_export(headers: list, rows: list) -> tuple:
    """Fallback CSV export."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    content = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    return io.BytesIO(content), "text/csv; charset=utf-8", "csv"


def _export_response(headers: list, rows: list, filename_base: str, sheet_name: str = "Dados"):
    """Build a StreamingResponse for export (xlsx or csv fallback)."""
    buf, content_type, ext = _try_openpyxl_export(headers, rows, sheet_name)
    if buf is None:
        buf, content_type, ext = _csv_export(headers, rows)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{filename_base}_{timestamp}.{ext}"

    return StreamingResponse(
        buf,
        media_type=content_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/prazos")
async def export_prazos_excel(request: Request, db: Session = Depends(get_db)):
    """Export prazos processuais to Excel/CSV."""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")

    # Try to import PrazoProcessual model
    try:
        from models.prazo import PrazoProcessual
        prazos = tenant_query(db, PrazoProcessual, request.state.org_id).order_by(
            PrazoProcessual.data_vencimento.asc()
        ).all()

        headers = ["Processo", "Cliente", "Tipo", "Intimação", "Vencimento", "Responsável", "Status"]
        rows = []
        for p in prazos:
            rows.append([
                getattr(p, "numero_processo", ""),
                getattr(p, "cliente_nome", ""),
                getattr(p, "tipo_prazo", ""),
                str(getattr(p, "data_intimacao", "")) if getattr(p, "data_intimacao", None) else "",
                str(getattr(p, "data_vencimento", "")) if getattr(p, "data_vencimento", None) else "",
                getattr(p, "responsavel", ""),
                getattr(p, "status", ""),
            ])
    except (ImportError, Exception) as e:
        logger.warning("PrazoProcessual model not available, exporting empty: %s", e)
        headers = ["Processo", "Cliente", "Tipo", "Intimação", "Vencimento", "Responsável", "Status"]
        rows = []

    return _export_response(headers, rows, "prazos_processuais", "Prazos")


@router.get("/export/processos")
async def export_processos_excel(request: Request, db: Session = Depends(get_db)):
    """Export cases/processos to Excel/CSV."""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")

    cases = tenant_query(db, Case, request.state.org_id).order_by(Case.created_at.desc()).all()

    headers = ["Nº Processo", "Cliente", "Tipo", "Status", "Data Abertura", "Última Atualização", "Responsável"]
    rows = []
    for c in cases:
        client_name = ""
        if c.client:
            client_name = getattr(c.client, "name", "") or getattr(c.client, "nome", "")
        rows.append([
            getattr(c, "case_number", "") or getattr(c, "numero_processo", ""),
            client_name,
            getattr(c, "visa_type", "") or getattr(c, "tipo", ""),
            c.status or "",
            c.created_at.strftime("%d/%m/%Y") if c.created_at else "",
            c.updated_at.strftime("%d/%m/%Y") if c.updated_at else "",
            getattr(c, "assigned_to", "") or getattr(c, "responsavel", ""),
        ])

    return _export_response(headers, rows, "processos", "Processos")


@router.get("/export/clientes")
async def export_clientes_excel(request: Request, db: Session = Depends(get_db)):
    """Export clients to Excel/CSV."""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")

    clients = tenant_query(db, Client, request.state.org_id).order_by(Client.name.asc()).all()

    headers = ["Nome", "E-mail", "Telefone", "CPF/CNPJ", "Data Cadastro", "Status"]
    rows = []
    for cl in clients:
        rows.append([
            cl.name or "",
            cl.email or "",
            getattr(cl, "phone", "") or "",
            getattr(cl, "cpf", "") or getattr(cl, "document_number", "") or "",
            cl.created_at.strftime("%d/%m/%Y") if cl.created_at else "",
            getattr(cl, "status", "ativo") or "ativo",
        ])

    return _export_response(headers, rows, "clientes", "Clientes")


# ===========================================================================
# CALCULADORAS CÍVEIS
# ===========================================================================

# ---------------------------------------------------------------------------
# 9. Correção Monetária (com índices reais BCB + fallback simplificado)
# ---------------------------------------------------------------------------
# Usa API do Banco Central do Brasil para índices reais (IPCA, INPC, IGP-M, Selic, TR).
# Quando a API está indisponível, usa fórmula simplificada com taxas anuais aproximadas.
#
# Test cases (fallback):
# Input: valor=10000, data_inicial=2024-01-01, data_final=2026-03-01, indice=IPCA (~6%/ano), juros_moratorios=False
#   meses = 26
#   valor_corrigido = 10000 * (1.06)^(26/12) = 10000 * 1.1348 = 11348.17
#   sem juros moratórios → total = 11348.17

from services.indices_economicos import get_indice, calcular_correcao

TAXAS_ANUAIS = {
    "ipca": 0.06,     # ~6% ao ano (aproximação / fallback)
    "inpc": 0.055,    # ~5.5% ao ano
    "igpm": 0.07,     # ~7% ao ano
    "selic": 0.1375,  # taxa Selic vigente (aproximação)
    "tr": 0.02,       # ~2% ao ano
}

INDICE_LABELS = {
    "ipca": "IPCA",
    "inpc": "INPC",
    "igpm": "IGP-M",
    "selic": "Selic",
    "tr": "TR",
}


def _calcular_correcao_fallback(
    valor: float,
    data_inicial: date,
    data_final: date,
    indice: str,
    juros_moratorios: bool = False,
    taxa_juros_mensal: float = 1.0,
) -> dict:
    """
    Correção monetária simplificada (fallback quando BCB API indisponível).
    Fórmula: valor_corrigido = valor * (1 + taxa_anual)^(meses/12)
    """
    dias = (data_final - data_inicial).days
    meses = dias / 30.0
    taxa_anual = TAXAS_ANUAIS.get(indice, 0.06)

    fator_correcao = (1 + taxa_anual) ** (meses / 12)
    valor_corrigido = valor * fator_correcao
    diferenca_correcao = valor_corrigido - valor

    juros_valor = 0.0
    if juros_moratorios:
        juros_valor = valor_corrigido * (taxa_juros_mensal / 100) * meses

    total = valor_corrigido + juros_valor

    return {
        "valor_original": round(valor, 2),
        "data_inicial": data_inicial.strftime("%d/%m/%Y"),
        "data_final": data_final.strftime("%d/%m/%Y"),
        "meses": round(meses, 1),
        "indice": indice,
        "indice_label": f"{INDICE_LABELS.get(indice, indice.upper())} (estimativa)",
        "taxa_anual": round(taxa_anual * 100, 2),
        "fator_correcao": round(fator_correcao, 6),
        "valor_corrigido": round(valor_corrigido, 2),
        "diferenca_correcao": round(diferenca_correcao, 2),
        "juros_moratorios": juros_moratorios,
        "taxa_juros_mensal": taxa_juros_mensal,
        "juros_valor": round(juros_valor, 2),
        "total": round(total, 2),
        "fonte": "estimativa",
    }


async def calcular_correcao_monetaria(
    valor: float,
    data_inicial: date,
    data_final: date,
    indice: str,
    juros_moratorios: bool = False,
    taxa_juros_mensal: float = 1.0,
) -> dict:
    """
    Correção monetária com índices reais do BCB.
    Fallback para fórmula simplificada se a API estiver indisponível.
    Fundamentação: CPC Art. 524, Lei 14.905/2024 (nova regra Selic).
    """
    # Try BCB real indices first
    dt_inicio_str = data_inicial.strftime("%d/%m/%Y")
    dt_fim_str = data_final.strftime("%d/%m/%Y")

    try:
        indices = await get_indice(indice, dt_inicio_str, dt_fim_str)
    except Exception:
        indices = []

    if indices:
        # Use real BCB data
        resultado_bcb = calcular_correcao(valor, indices)
        dias = (data_final - data_inicial).days
        meses = dias / 30.0

        juros_valor = 0.0
        if juros_moratorios:
            juros_valor = resultado_bcb["valor_corrigido"] * (taxa_juros_mensal / 100) * meses

        total = resultado_bcb["valor_corrigido"] + juros_valor

        return {
            "valor_original": round(valor, 2),
            "data_inicial": data_inicial.strftime("%d/%m/%Y"),
            "data_final": data_final.strftime("%d/%m/%Y"),
            "meses": round(meses, 1),
            "indice": indice,
            "indice_label": f"{INDICE_LABELS.get(indice, indice.upper())} (BCB real)",
            "taxa_anual": None,
            "fator_correcao": resultado_bcb["fator_acumulado"],
            "valor_corrigido": resultado_bcb["valor_corrigido"],
            "diferenca_correcao": round(resultado_bcb["valor_corrigido"] - valor, 2),
            "variacao_percentual": resultado_bcb["variacao_percentual"],
            "juros_moratorios": juros_moratorios,
            "taxa_juros_mensal": taxa_juros_mensal,
            "juros_valor": round(juros_valor, 2),
            "total": round(total, 2),
            "fonte": "bcb",
            "detalhes_mensais": resultado_bcb["detalhes"],
        }

    # Fallback to simplified formula
    logger.warning("BCB API unavailable for %s, using fallback formula", indice)
    return _calcular_correcao_fallback(
        valor, data_inicial, data_final, indice,
        juros_moratorios, taxa_juros_mensal,
    )


@router.get("/correcao-monetaria", response_class=HTMLResponse)
async def correcao_monetaria_form(request: Request, db: Session = Depends(get_db)):
    """Correção Monetária calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/correcao_monetaria.html", get_context(request, db))


@router.post("/correcao-monetaria/calcular", response_class=HTMLResponse)
async def correcao_monetaria_calcular(
    request: Request,
    db: Session = Depends(get_db),
    valor: float = Form(...),
    data_inicial: str = Form(...),
    data_final: str = Form(...),
    indice: str = Form("ipca"),
    juros_moratorios: bool = Form(False),
    taxa_juros_mensal: float = Form(1.0),
):
    """Calculate correção monetária."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_inicial = datetime.strptime(data_inicial, "%Y-%m-%d").date()
        dt_final = datetime.strptime(data_final, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse("app/tools/correcao_monetaria.html", {
            **get_context(request, db),
            "error": "Datas inválidas. Use o formato correto.",
        })

    if dt_final <= dt_inicial:
        return templates.TemplateResponse("app/tools/correcao_monetaria.html", {
            **get_context(request, db),
            "error": "A data final deve ser posterior à data inicial.",
        })

    resultado = await calcular_correcao_monetaria(
        valor=valor,
        data_inicial=dt_inicial,
        data_final=dt_final,
        indice=indice,
        juros_moratorios=juros_moratorios,
        taxa_juros_mensal=taxa_juros_mensal,
    )

    return templates.TemplateResponse("app/tools/correcao_monetaria.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "valor": valor,
            "data_inicial": data_inicial,
            "data_final": data_final,
            "indice": indice,
            "juros_moratorios": juros_moratorios,
            "taxa_juros_mensal": taxa_juros_mensal,
        },
    })


# ---------------------------------------------------------------------------
# 10. Juros de Mora
# ---------------------------------------------------------------------------
# Test cases:
# Input: principal=5000, data_vencimento=2026-01-01, data_pagamento=2026-04-01, tipo=cc_406 (1%/mês)
#   dias_atraso = 90
#   juros = 5000 * 0.01 * (90 / 30) = 150.00
#   total = 5000 + 150 = 5150.00

def calcular_juros_mora(
    principal: float,
    data_vencimento: date,
    data_pagamento: date,
    tipo: str = "cc_406",
    incluir_multa_523: bool = False,
) -> dict:
    """
    Calcula juros de mora.
    Fundamentação: CC Art. 406, CPC Art. 523 §1º (multa 10% + honorários 10%).
    Tipo cc_406: 1% ao mês (Art. 406 CC).
    Tipo selic: taxa Selic (Art. 406 + Lei 14.905/2024).
    """
    dias_atraso = max((data_pagamento - data_vencimento).days, 0)
    meses_atraso = dias_atraso / 30.0

    if tipo == "selic":
        taxa_mensal = 0.1375 / 12  # Selic anual / 12
        taxa_label = "Selic (~1,15%/mês)"
    else:
        taxa_mensal = 0.01  # 1% ao mês
        taxa_label = "1% ao mês (Art. 406 CC)"

    juros = principal * taxa_mensal * meses_atraso
    subtotal = principal + juros

    multa_523 = 0.0
    honorarios_523 = 0.0
    if incluir_multa_523:
        multa_523 = subtotal * 0.10
        honorarios_523 = subtotal * 0.10

    total = subtotal + multa_523 + honorarios_523

    return {
        "principal": round(principal, 2),
        "data_vencimento": data_vencimento.strftime("%d/%m/%Y"),
        "data_pagamento": data_pagamento.strftime("%d/%m/%Y"),
        "dias_atraso": dias_atraso,
        "meses_atraso": round(meses_atraso, 1),
        "tipo": tipo,
        "taxa_label": taxa_label,
        "taxa_mensal_pct": round(taxa_mensal * 100, 2),
        "juros": round(juros, 2),
        "subtotal": round(subtotal, 2),
        "incluir_multa_523": incluir_multa_523,
        "multa_523": round(multa_523, 2),
        "honorarios_523": round(honorarios_523, 2),
        "total": round(total, 2),
    }


@router.get("/juros-mora", response_class=HTMLResponse)
async def juros_mora_form(request: Request, db: Session = Depends(get_db)):
    """Juros de Mora calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/juros_mora.html", get_context(request, db))


@router.post("/juros-mora/calcular", response_class=HTMLResponse)
async def juros_mora_calcular(
    request: Request,
    db: Session = Depends(get_db),
    principal: float = Form(...),
    data_vencimento: str = Form(...),
    data_pagamento: str = Form(...),
    tipo: str = Form("cc_406"),
    incluir_multa_523: bool = Form(False),
):
    """Calculate juros de mora."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_venc = datetime.strptime(data_vencimento, "%Y-%m-%d").date()
        dt_pgto = datetime.strptime(data_pagamento, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse("app/tools/juros_mora.html", {
            **get_context(request, db),
            "error": "Datas inválidas. Use o formato correto.",
        })

    if dt_pgto <= dt_venc:
        return templates.TemplateResponse("app/tools/juros_mora.html", {
            **get_context(request, db),
            "error": "A data de pagamento deve ser posterior à data de vencimento.",
        })

    resultado = calcular_juros_mora(
        principal=principal,
        data_vencimento=dt_venc,
        data_pagamento=dt_pgto,
        tipo=tipo,
        incluir_multa_523=incluir_multa_523,
    )

    return templates.TemplateResponse("app/tools/juros_mora.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "principal": principal,
            "data_vencimento": data_vencimento,
            "data_pagamento": data_pagamento,
            "tipo": tipo,
            "incluir_multa_523": incluir_multa_523,
        },
    })


# ---------------------------------------------------------------------------
# 11. Honorários Sucumbenciais
# ---------------------------------------------------------------------------
# Test cases:
# Input: valor_causa=100000, percentual=15, fase=conhecimento
#   honorarios = 100000 * 0.15 = 15000.00
# Fazenda (§3º): até 200 SM → mín 10%, máx 20%

def calcular_honorarios_sucumbenciais(
    valor_causa: float,
    percentual: float = 15.0,
    fase: str = "conhecimento",
    contra_fazenda: bool = False,
    sucumbencia_reciproca: bool = False,
    percentual_autor: float = 50.0,
) -> dict:
    """
    Calcula honorários sucumbenciais.
    Fundamentação: CPC Art. 85 §§2-11.
    Fases: conhecimento, recurso (§11 honorários recursais), execução.
    """
    honorarios = valor_causa * (percentual / 100)

    # Faixas do §3º para Fazenda Pública
    fazenda_faixa = ""
    if contra_fazenda:
        sm = SALARIO_MINIMO_2026
        if valor_causa <= 200 * sm:
            fazenda_faixa = f"Até 200 SM (R$ {200*sm:,.2f}): mín 10%, máx 20%"
        elif valor_causa <= 2000 * sm:
            fazenda_faixa = f"200-2000 SM: mín 8%, máx 10%"
        elif valor_causa <= 20000 * sm:
            fazenda_faixa = f"2000-20000 SM: mín 5%, máx 8%"
        elif valor_causa <= 100000 * sm:
            fazenda_faixa = f"20000-100000 SM: mín 3%, máx 5%"
        else:
            fazenda_faixa = f"Acima de 100000 SM: mín 1%, máx 3%"

    # Honorários recursais (§11)
    honorarios_recursais = 0.0
    if fase == "recurso":
        # Majoração de até 20% sobre o valor fixado (mín = 1% a mais)
        honorarios_recursais = honorarios * 0.10  # estimativa: 10% a mais

    # Sucumbência recíproca
    honorarios_autor = 0.0
    honorarios_reu = 0.0
    if sucumbencia_reciproca:
        honorarios_autor = honorarios * (percentual_autor / 100)
        honorarios_reu = honorarios * ((100 - percentual_autor) / 100)

    total = honorarios + honorarios_recursais

    fase_labels = {
        "conhecimento": "Conhecimento",
        "recurso": "Recurso (§11)",
        "execucao": "Execução",
    }

    return {
        "valor_causa": round(valor_causa, 2),
        "percentual": percentual,
        "fase": fase,
        "fase_label": fase_labels.get(fase, fase),
        "honorarios": round(honorarios, 2),
        "honorarios_recursais": round(honorarios_recursais, 2),
        "contra_fazenda": contra_fazenda,
        "fazenda_faixa": fazenda_faixa,
        "sucumbencia_reciproca": sucumbencia_reciproca,
        "honorarios_autor": round(honorarios_autor, 2),
        "honorarios_reu": round(honorarios_reu, 2),
        "percentual_autor": percentual_autor,
        "total": round(total, 2),
    }


@router.get("/honorarios", response_class=HTMLResponse)
async def honorarios_form(request: Request, db: Session = Depends(get_db)):
    """Honorários Sucumbenciais calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/honorarios.html", get_context(request, db))


@router.post("/honorarios/calcular", response_class=HTMLResponse)
async def honorarios_calcular(
    request: Request,
    db: Session = Depends(get_db),
    valor_causa: float = Form(...),
    percentual: float = Form(15.0),
    fase: str = Form("conhecimento"),
    contra_fazenda: bool = Form(False),
    sucumbencia_reciproca: bool = Form(False),
    percentual_autor: float = Form(50.0),
):
    """Calculate honorários sucumbenciais."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    if percentual < 10 or percentual > 20:
        return templates.TemplateResponse("app/tools/honorarios.html", {
            **get_context(request, db),
            "error": "O percentual deve estar entre 10% e 20% (Art. 85, §2º CPC).",
        })

    resultado = calcular_honorarios_sucumbenciais(
        valor_causa=valor_causa,
        percentual=percentual,
        fase=fase,
        contra_fazenda=contra_fazenda,
        sucumbencia_reciproca=sucumbencia_reciproca,
        percentual_autor=percentual_autor,
    )

    return templates.TemplateResponse("app/tools/honorarios.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "valor_causa": valor_causa,
            "percentual": percentual,
            "fase": fase,
            "contra_fazenda": contra_fazenda,
            "sucumbencia_reciproca": sucumbencia_reciproca,
            "percentual_autor": percentual_autor,
        },
    })


# ---------------------------------------------------------------------------
# 12. Pensão Alimentícia
# ---------------------------------------------------------------------------
# Test cases:
# Input: renda=5000, num_filhos=2, percentual=33, despesas_especiais=0
#   pensao_total = 5000 * 0.33 = 1650.00
#   pensao_por_filho = 1650.00 / 2 = 825.00
#   em SM: 1650 / 1518 = 1.09 SM

def calcular_pensao_alimenticia(
    renda: float,
    num_filhos: int = 1,
    percentual: float = 30.0,
    despesas_especiais: float = 0.0,
    salario_minimo: float = SALARIO_MINIMO_2026,
) -> dict:
    """
    Calcula pensão alimentícia.
    Fundamentação: CC Art. 1694-1710.
    Padrão: ~30% para 1 filho, ~33% para 2+.
    """
    pensao_total = renda * (percentual / 100)
    pensao_por_filho = pensao_total / num_filhos if num_filhos > 0 else pensao_total
    pensao_com_despesas = pensao_total + despesas_especiais
    pensao_por_filho_com_desp = pensao_com_despesas / num_filhos if num_filhos > 0 else pensao_com_despesas

    pensao_em_sm = pensao_total / salario_minimo
    pensao_por_filho_sm = pensao_por_filho / salario_minimo

    # Teto razoável (jurisprudência): em geral não ultrapassa 33% para 1 filho
    comprometimento = (pensao_com_despesas / renda) * 100 if renda > 0 else 0

    return {
        "renda": round(renda, 2),
        "num_filhos": num_filhos,
        "percentual": percentual,
        "despesas_especiais": round(despesas_especiais, 2),
        "pensao_total": round(pensao_total, 2),
        "pensao_por_filho": round(pensao_por_filho, 2),
        "pensao_com_despesas": round(pensao_com_despesas, 2),
        "pensao_por_filho_com_desp": round(pensao_por_filho_com_desp, 2),
        "pensao_em_sm": round(pensao_em_sm, 2),
        "pensao_por_filho_sm": round(pensao_por_filho_sm, 2),
        "salario_minimo": salario_minimo,
        "comprometimento": round(comprometimento, 1),
        "pensao_anual": round(pensao_com_despesas * 12, 2),
        "pensao_13": round(pensao_com_despesas, 2),
    }


@router.get("/pensao-alimenticia", response_class=HTMLResponse)
async def pensao_alimenticia_form(request: Request, db: Session = Depends(get_db)):
    """Pensão Alimentícia calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/pensao_alimenticia.html", get_context(request, db))


@router.post("/pensao-alimenticia/calcular", response_class=HTMLResponse)
async def pensao_alimenticia_calcular(
    request: Request,
    db: Session = Depends(get_db),
    renda: float = Form(...),
    num_filhos: int = Form(1),
    percentual: float = Form(30.0),
    despesas_especiais: float = Form(0.0),
):
    """Calculate pensão alimentícia."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    if num_filhos < 1:
        return templates.TemplateResponse("app/tools/pensao_alimenticia.html", {
            **get_context(request, db),
            "error": "Número de filhos deve ser pelo menos 1.",
        })

    resultado = calcular_pensao_alimenticia(
        renda=renda,
        num_filhos=num_filhos,
        percentual=percentual,
        despesas_especiais=despesas_especiais,
    )

    return templates.TemplateResponse("app/tools/pensao_alimenticia.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "renda": renda,
            "num_filhos": num_filhos,
            "percentual": percentual,
            "despesas_especiais": despesas_especiais,
        },
    })


# ---------------------------------------------------------------------------
# 13. Atualização de Débito Judicial (Art. 523-524 CPC)
# ---------------------------------------------------------------------------
# Test cases:
# Input: valor_condenacao=20000, data_sentenca=2025-06-01, data_atual=2026-03-29,
#        indice=ipca, tem_multa_523=True
#   meses = 9.93
#   valor_corrigido = 20000 * (1.06)^(9.93/12) = 20000 * 1.0496 = 20991.68
#   juros_mora = 20991.68 * 0.01 * 9.93 = 2084.47
#   subtotal = 20991.68 + 2084.47 = 23076.15
#   multa_10 = 23076.15 * 0.10 = 2307.62
#   honorarios_10 = 23076.15 * 0.10 = 2307.62
#   total = 23076.15 + 2307.62 + 2307.62 = 27691.38

def calcular_debito_judicial(
    valor_condenacao: float,
    data_sentenca: date,
    data_atual: date,
    indice: str = "ipca",
    tem_multa_523: bool = True,
    taxa_juros_mensal: float = 1.0,
) -> dict:
    """
    Atualização de débito judicial conforme Art. 523-524 CPC.
    Se não pagou em 15 dias: multa 10% + honorários 10% automáticos.
    """
    dias = (data_atual - data_sentenca).days
    meses = dias / 30.0
    taxa_anual = TAXAS_ANUAIS.get(indice, 0.06)

    fator_correcao = (1 + taxa_anual) ** (meses / 12)
    valor_corrigido = valor_condenacao * fator_correcao
    correcao = valor_corrigido - valor_condenacao

    juros_mora = valor_corrigido * (taxa_juros_mensal / 100) * meses
    subtotal = valor_corrigido + juros_mora

    multa_10 = 0.0
    honorarios_10 = 0.0
    if tem_multa_523:
        multa_10 = subtotal * 0.10
        honorarios_10 = subtotal * 0.10

    total = subtotal + multa_10 + honorarios_10

    indice_labels = {
        "ipca": "IPCA (~6%/ano)",
        "inpc": "INPC (~5,5%/ano)",
        "igpm": "IGP-M (~7%/ano)",
        "selic": "Selic (~13,75%/ano)",
        "tr": "TR (~2%/ano)",
    }

    return {
        "valor_condenacao": round(valor_condenacao, 2),
        "data_sentenca": data_sentenca.strftime("%d/%m/%Y"),
        "data_atual": data_atual.strftime("%d/%m/%Y"),
        "meses": round(meses, 1),
        "indice": indice,
        "indice_label": indice_labels.get(indice, indice.upper()),
        "taxa_anual": round(taxa_anual * 100, 2),
        "fator_correcao": round(fator_correcao, 6),
        "valor_corrigido": round(valor_corrigido, 2),
        "correcao": round(correcao, 2),
        "taxa_juros_mensal": taxa_juros_mensal,
        "juros_mora": round(juros_mora, 2),
        "subtotal": round(subtotal, 2),
        "tem_multa_523": tem_multa_523,
        "multa_10": round(multa_10, 2),
        "honorarios_10": round(honorarios_10, 2),
        "total": round(total, 2),
    }


@router.get("/debito-judicial", response_class=HTMLResponse)
async def debito_judicial_form(request: Request, db: Session = Depends(get_db)):
    """Atualização de Débito Judicial calculator form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/debito_judicial.html", get_context(request, db))


@router.post("/debito-judicial/calcular", response_class=HTMLResponse)
async def debito_judicial_calcular(
    request: Request,
    db: Session = Depends(get_db),
    valor_condenacao: float = Form(...),
    data_sentenca: str = Form(...),
    data_atual: str = Form(...),
    indice: str = Form("ipca"),
    tem_multa_523: bool = Form(False),
    taxa_juros_mensal: float = Form(1.0),
):
    """Calculate atualização de débito judicial."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_sentenca = datetime.strptime(data_sentenca, "%Y-%m-%d").date()
        dt_atual = datetime.strptime(data_atual, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse("app/tools/debito_judicial.html", {
            **get_context(request, db),
            "error": "Datas inválidas. Use o formato correto.",
        })

    if dt_atual <= dt_sentenca:
        return templates.TemplateResponse("app/tools/debito_judicial.html", {
            **get_context(request, db),
            "error": "A data atual deve ser posterior à data da sentença.",
        })

    resultado = calcular_debito_judicial(
        valor_condenacao=valor_condenacao,
        data_sentenca=dt_sentenca,
        data_atual=dt_atual,
        indice=indice,
        tem_multa_523=tem_multa_523,
        taxa_juros_mensal=taxa_juros_mensal,
    )

    return templates.TemplateResponse("app/tools/debito_judicial.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "valor_condenacao": valor_condenacao,
            "data_sentenca": data_sentenca,
            "data_atual": data_atual,
            "indice": indice,
            "tem_multa_523": tem_multa_523,
            "taxa_juros_mensal": taxa_juros_mensal,
        },
    })


# ===========================================================================
# CALCULADORAS PREVIDENCIARIAS (EC 103/2019)
# ===========================================================================

# ---------------------------------------------------------------------------
# 1. Simulação de Tempo de Contribuição Faltante
# ---------------------------------------------------------------------------
# Test cases:
# Mulher, nascida 29/03/1971 (55 anos), 12 anos contribuição, início contrib 01/01/2007
#   Regra geral: 62 idade + 15 anos contrib → falta 3 anos contrib + 7 anos idade
#   Data estimada: 29/03/2033 (quando completa 62)
# Homem, nascido 01/01/1965 (61 anos), 18 anos contribuição, início contrib 01/01/2008
#   Regra geral: 65 idade + 20 anos contrib → falta 2 anos contrib + 4 anos idade
#   Data estimada: 01/01/2030 (quando completa 65)

SALARIO_MINIMO_2026 = 1518.00
DATA_REFORMA = date(2019, 11, 13)


def calcular_tempo_contribuicao(
    data_nascimento: date,
    sexo: str,
    anos_contribuicao: int,
    meses_contribuicao: int,
    dias_contribuicao: int,
    data_inicio_contribuicao: date,
) -> dict:
    """
    Simula tempo de contribuição faltante para aposentadoria.
    Fundamentação: EC 103/2019, Art. 15-20.
    """
    hoje = date.today()

    # Idade atual
    idade_atual = relativedelta(hoje, data_nascimento)
    idade_anos = idade_atual.years

    # Requisitos por sexo
    if sexo == "M":
        idade_minima = 65
        contrib_minima_anos = 20
    else:
        idade_minima = 62
        contrib_minima_anos = 15

    # Tempo de contribuição total em dias
    total_dias_contrib = anos_contribuicao * 365 + meses_contribuicao * 30 + dias_contribuicao
    total_anos_contrib = total_dias_contrib / 365.25

    # Tempo faltante de contribuição
    falta_contrib_anos = max(0, contrib_minima_anos - total_anos_contrib)
    falta_contrib_anos_int = int(falta_contrib_anos)
    falta_contrib_meses = int((falta_contrib_anos - falta_contrib_anos_int) * 12)
    falta_contrib_dias = int(((falta_contrib_anos - falta_contrib_anos_int) * 12 - falta_contrib_meses) * 30)

    # Tempo faltante de idade
    falta_idade_anos = max(0, idade_minima - idade_anos)

    # Data estimada por idade
    data_idade_minima = data_nascimento + relativedelta(years=idade_minima)

    # Data estimada por contribuição
    data_contrib_completa = hoje + relativedelta(years=falta_contrib_anos_int, months=falta_contrib_meses, days=falta_contrib_dias)

    # Data estimada de aposentadoria = a mais distante
    data_estimada = max(data_idade_minima, data_contrib_completa)

    # Regra de transição - pedágio 100%
    # Calcula tempo que faltava na data da reforma
    tempo_ate_reforma = relativedelta(DATA_REFORMA, data_inicio_contribuicao)
    anos_ate_reforma = tempo_ate_reforma.years + tempo_ate_reforma.months / 12
    faltava_reforma = max(0, contrib_minima_anos - anos_ate_reforma)
    pedagio_100 = faltava_reforma * 2  # precisa cumprir o dobro

    # Já contribuiu após reforma
    contrib_pos_reforma = relativedelta(hoje, DATA_REFORMA)
    anos_pos_reforma = contrib_pos_reforma.years + contrib_pos_reforma.months / 12
    pedagio_restante = max(0, pedagio_100 - anos_pos_reforma)
    pedagio_restante_anos = int(pedagio_restante)
    pedagio_restante_meses = int((pedagio_restante - pedagio_restante_anos) * 12)

    data_pedagio = hoje + relativedelta(years=pedagio_restante_anos, months=pedagio_restante_meses)
    data_pedagio_final = max(data_pedagio, data_idade_minima)

    return {
        "sexo": sexo,
        "sexo_label": "Masculino" if sexo == "M" else "Feminino",
        "idade_anos": idade_anos,
        "idade_minima": idade_minima,
        "contrib_minima_anos": contrib_minima_anos,
        "anos_contribuicao": anos_contribuicao,
        "meses_contribuicao": meses_contribuicao,
        "dias_contribuicao": dias_contribuicao,
        "total_anos_contrib": round(total_anos_contrib, 1),
        "falta_contrib_anos": falta_contrib_anos_int,
        "falta_contrib_meses": falta_contrib_meses,
        "falta_contrib_dias": falta_contrib_dias,
        "falta_idade_anos": falta_idade_anos,
        "data_estimada": data_estimada.strftime("%d/%m/%Y"),
        "data_idade_minima": data_idade_minima.strftime("%d/%m/%Y"),
        "data_contrib_completa": data_contrib_completa.strftime("%d/%m/%Y"),
        # Transição pedágio 100%
        "faltava_reforma": round(faltava_reforma, 1),
        "pedagio_100": round(pedagio_100, 1),
        "pedagio_restante_anos": pedagio_restante_anos,
        "pedagio_restante_meses": pedagio_restante_meses,
        "data_pedagio_final": data_pedagio_final.strftime("%d/%m/%Y"),
    }


@router.get("/tempo-contribuicao", response_class=HTMLResponse)
async def tempo_contribuicao_form(request: Request, db: Session = Depends(get_db)):
    """Tempo de contribuição faltante - form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/tempo_contribuicao.html", get_context(request, db))


@router.post("/tempo-contribuicao/calcular", response_class=HTMLResponse)
async def tempo_contribuicao_calcular(
    request: Request,
    db: Session = Depends(get_db),
    data_nascimento: str = Form(...),
    sexo: str = Form(...),
    anos_contribuicao: int = Form(0),
    meses_contribuicao: int = Form(0),
    dias_contribuicao: int = Form(0),
    data_inicio_contribuicao: str = Form(...),
):
    """Calcular tempo de contribuição faltante."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_nasc = datetime.strptime(data_nascimento, "%Y-%m-%d").date()
        dt_inicio = datetime.strptime(data_inicio_contribuicao, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse("app/tools/tempo_contribuicao.html", {
            **get_context(request, db),
            "error": "Data invalida. Use o formato correto.",
        })

    resultado = calcular_tempo_contribuicao(
        data_nascimento=dt_nasc,
        sexo=sexo,
        anos_contribuicao=anos_contribuicao,
        meses_contribuicao=meses_contribuicao,
        dias_contribuicao=dias_contribuicao,
        data_inicio_contribuicao=dt_inicio,
    )

    return templates.TemplateResponse("app/tools/tempo_contribuicao.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "data_nascimento": data_nascimento,
            "sexo": sexo,
            "anos_contribuicao": anos_contribuicao,
            "meses_contribuicao": meses_contribuicao,
            "dias_contribuicao": dias_contribuicao,
            "data_inicio_contribuicao": data_inicio_contribuicao,
        },
    })


# ---------------------------------------------------------------------------
# 2. BPC/LOAS Análise
# ---------------------------------------------------------------------------
# Test cases:
# Família 4 pessoas, renda R$1.200 → per capita R$300 → ELEGÍVEL (R$300 <= R$379,50)
# Família 3 pessoas, renda R$1.500 → per capita R$500 → NÃO ELEGÍVEL (R$500 > R$379,50)
# Idoso 65 anos, sem deficiência, família 2, renda R$700 → per capita R$350 → ELEGÍVEL
# Pessoa 40 anos, com deficiência, família 1, renda R$300 → per capita R$300 → ELEGÍVEL

def calcular_bpc_loas(
    renda_familiar: float,
    num_membros: int,
    idade: int,
    possui_deficiencia: bool,
) -> dict:
    """
    Analisa elegibilidade ao BPC/LOAS.
    Fundamentação: Lei 8.742/93 (LOAS), Art. 20-21.
    Valor: 1 salário mínimo (R$1.518,00 em 2026).
    Critério de renda: per capita <= 1/4 do SM.
    """
    sm = SALARIO_MINIMO_2026
    limite_renda = sm / 4  # R$379,50

    renda_per_capita = renda_familiar / max(num_membros, 1)

    criterio_renda = renda_per_capita <= limite_renda
    criterio_idade = idade >= 65
    criterio_deficiencia = possui_deficiencia
    criterio_pessoal = criterio_idade or criterio_deficiencia

    elegivel = criterio_renda and criterio_pessoal

    # Motivos de inelegibilidade
    motivos = []
    if not criterio_renda:
        motivos.append(f"Renda per capita (R$ {renda_per_capita:.2f}) excede o limite de 1/4 do SM (R$ {limite_renda:.2f})")
    if not criterio_pessoal:
        motivos.append("Requerente nao possui 65 anos ou mais e nao possui deficiencia")

    return {
        "renda_familiar": round(renda_familiar, 2),
        "num_membros": num_membros,
        "idade": idade,
        "possui_deficiencia": possui_deficiencia,
        "renda_per_capita": round(renda_per_capita, 2),
        "limite_renda": round(limite_renda, 2),
        "salario_minimo": sm,
        "criterio_renda": criterio_renda,
        "criterio_idade": criterio_idade,
        "criterio_deficiencia": criterio_deficiencia,
        "criterio_pessoal": criterio_pessoal,
        "elegivel": elegivel,
        "valor_beneficio": sm if elegivel else 0,
        "motivos_inelegibilidade": motivos,
    }


@router.get("/bpc-loas", response_class=HTMLResponse)
async def bpc_loas_form(request: Request, db: Session = Depends(get_db)):
    """BPC/LOAS analysis form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/bpc_loas.html", get_context(request, db))


@router.post("/bpc-loas/calcular", response_class=HTMLResponse)
async def bpc_loas_calcular(
    request: Request,
    db: Session = Depends(get_db),
    renda_familiar: float = Form(...),
    num_membros: int = Form(...),
    idade: int = Form(...),
    possui_deficiencia: str = Form("nao"),
):
    """Calcular elegibilidade BPC/LOAS."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_bpc_loas(
        renda_familiar=renda_familiar,
        num_membros=num_membros,
        idade=idade,
        possui_deficiencia=(possui_deficiencia == "sim"),
    )

    return templates.TemplateResponse("app/tools/bpc_loas.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "renda_familiar": renda_familiar,
            "num_membros": num_membros,
            "idade": idade,
            "possui_deficiencia": possui_deficiencia,
        },
    })


# ---------------------------------------------------------------------------
# 3. Aposentadoria por Idade (Regra Geral + Transição)
# ---------------------------------------------------------------------------
# Test cases:
# Mulher 62 anos, 20 anos contribuição, média salarial R$3.000, 240 meses carência
#   Coeficiente: 60% + 2% * (20-15) = 70% → R$3.000 * 70% = R$2.100
# Homem 65 anos, 25 anos contribuição, média R$4.000, 300 meses carência
#   Coeficiente: 60% + 2% * (25-20) = 70% → R$4.000 * 70% = R$2.800
# Mulher 60 anos, 15 anos contribuição → NÃO ELEGÍVEL (idade < 62)

# Idade mínima progressiva para mulheres (transição): 60 em 2020, +6m/ano
TRANSICAO_IDADE_MULHER = {
    2020: 60.0, 2021: 60.5, 2022: 61.0, 2023: 61.5, 2024: 62.0,
}
# Para homens: sempre 65 (sem transição de idade)


def calcular_aposentadoria_idade(
    data_nascimento: date,
    sexo: str,
    anos_contribuicao: int,
    meses_contribuicao: int,
    carencia_meses: int,
    media_salarial: float,
) -> dict:
    """
    Calcula aposentadoria por idade conforme EC 103/2019, Art. 15, 26.
    Regra geral + transição + cálculo do benefício.
    """
    hoje = date.today()
    idade = relativedelta(hoje, data_nascimento).years
    total_contrib_anos = anos_contribuicao + meses_contribuicao / 12

    # Requisitos por sexo
    if sexo == "M":
        idade_minima = 65
        contrib_minima = 20
        carencia_minima = 180
    else:
        idade_minima = 62
        contrib_minima = 15
        carencia_minima = 180

    # Verificar elegibilidade
    cumpre_idade = idade >= idade_minima
    cumpre_contribuicao = total_contrib_anos >= contrib_minima
    cumpre_carencia = carencia_meses >= carencia_minima
    elegivel = cumpre_idade and cumpre_contribuicao and cumpre_carencia

    # Cálculo do coeficiente (Art. 26 EC 103/2019)
    # 60% + 2% por ano acima do mínimo
    anos_acima_minimo = max(0, total_contrib_anos - contrib_minima)
    coeficiente = 60 + (2 * int(anos_acima_minimo))
    coeficiente = min(coeficiente, 100)  # teto 100%

    # Valor do benefício
    valor_beneficio = media_salarial * (coeficiente / 100)
    valor_beneficio = max(valor_beneficio, SALARIO_MINIMO_2026)  # piso = SM

    # Regra de transição (idade progressiva - já encerrada para mulheres em 2024)
    ano_atual = hoje.year
    if sexo == "F" and ano_atual <= 2024:
        idade_transicao = TRANSICAO_IDADE_MULHER.get(ano_atual, 62.0)
    else:
        idade_transicao = idade_minima

    # Data estimada se não elegível
    data_estimada_idade = data_nascimento + relativedelta(years=idade_minima)
    falta_contrib = max(0, contrib_minima - total_contrib_anos)
    falta_contrib_anos = int(falta_contrib)
    falta_contrib_meses = int((falta_contrib - falta_contrib_anos) * 12)
    data_estimada_contrib = hoje + relativedelta(years=falta_contrib_anos, months=falta_contrib_meses)
    data_estimada = max(data_estimada_idade, data_estimada_contrib) if not elegivel else hoje

    # Motivos de inelegibilidade
    motivos = []
    if not cumpre_idade:
        motivos.append(f"Idade atual ({idade} anos) inferior a {idade_minima} anos")
    if not cumpre_contribuicao:
        motivos.append(f"Tempo de contribuicao ({total_contrib_anos:.1f} anos) inferior a {contrib_minima} anos")
    if not cumpre_carencia:
        motivos.append(f"Carencia ({carencia_meses} meses) inferior a {carencia_minima} meses")

    return {
        "sexo": sexo,
        "sexo_label": "Masculino" if sexo == "M" else "Feminino",
        "idade": idade,
        "idade_minima": idade_minima,
        "total_contrib_anos": round(total_contrib_anos, 1),
        "contrib_minima": contrib_minima,
        "carencia_meses": carencia_meses,
        "carencia_minima": carencia_minima,
        "media_salarial": round(media_salarial, 2),
        "cumpre_idade": cumpre_idade,
        "cumpre_contribuicao": cumpre_contribuicao,
        "cumpre_carencia": cumpre_carencia,
        "elegivel": elegivel,
        "coeficiente": coeficiente,
        "valor_beneficio": round(valor_beneficio, 2),
        "motivos_inelegibilidade": motivos,
        "data_estimada": data_estimada.strftime("%d/%m/%Y"),
        "falta_contrib_anos": falta_contrib_anos,
        "falta_contrib_meses": falta_contrib_meses,
        "idade_transicao": idade_transicao,
    }


@router.get("/aposentadoria-idade", response_class=HTMLResponse)
async def aposentadoria_idade_form(request: Request, db: Session = Depends(get_db)):
    """Aposentadoria por idade - form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/aposentadoria_idade.html", get_context(request, db))


@router.post("/aposentadoria-idade/calcular", response_class=HTMLResponse)
async def aposentadoria_idade_calcular(
    request: Request,
    db: Session = Depends(get_db),
    data_nascimento: str = Form(...),
    sexo: str = Form(...),
    anos_contribuicao: int = Form(0),
    meses_contribuicao: int = Form(0),
    carencia_meses: int = Form(0),
    media_salarial: float = Form(...),
):
    """Calcular aposentadoria por idade."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    try:
        dt_nasc = datetime.strptime(data_nascimento, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse("app/tools/aposentadoria_idade.html", {
            **get_context(request, db),
            "error": "Data de nascimento invalida.",
        })

    resultado = calcular_aposentadoria_idade(
        data_nascimento=dt_nasc,
        sexo=sexo,
        anos_contribuicao=anos_contribuicao,
        meses_contribuicao=meses_contribuicao,
        carencia_meses=carencia_meses,
        media_salarial=media_salarial,
    )

    return templates.TemplateResponse("app/tools/aposentadoria_idade.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "data_nascimento": data_nascimento,
            "sexo": sexo,
            "anos_contribuicao": anos_contribuicao,
            "meses_contribuicao": meses_contribuicao,
            "carencia_meses": carencia_meses,
            "media_salarial": media_salarial,
        },
    })


# ---------------------------------------------------------------------------
# 4. Pensão por Morte
# ---------------------------------------------------------------------------
# Test cases:
# Benefício R$4.000, cônjuge + 1 filho menor → 2 dependentes → cota 70% → R$2.800
# Benefício R$3.000, cônjuge sozinho → 1 dependente → cota 60% → R$1.800
# Benefício R$5.000, 3 dependentes → cota 80% → R$4.000
# Benefício R$2.000, 5 dependentes → cota 100% → R$2.000

# Tabela de duração da pensão para cônjuge (Art. 77, §5º, Lei 8.213/91)
DURACAO_PENSAO_CONJUGE = [
    (21, 3),    # menos de 21 anos: 3 anos
    (26, 6),    # 21-26: 6 anos
    (29, 10),   # 27-29: 10 anos
    (40, 15),   # 30-40: 15 anos
    (43, 20),   # 41-43: 20 anos
    (999, 0),   # 44+: vitalícia (0 = sem limite)
]


def calcular_pensao_morte(
    valor_beneficio: float,
    num_dependentes: int,
    idade_conjuge: int = 0,
    tem_menor_21: bool = False,
) -> dict:
    """
    Calcula pensão por morte conforme EC 103/2019, Art. 23.
    Cota familiar: 50% + 10% por dependente (max 100%).
    """
    sm = SALARIO_MINIMO_2026

    # Cota familiar
    cota_percentual = 50 + (10 * num_dependentes)
    cota_percentual = min(cota_percentual, 100)

    valor_pensao = valor_beneficio * (cota_percentual / 100)
    valor_pensao = max(valor_pensao, sm)  # piso = SM

    # Duração da pensão para cônjuge
    duracao_anos = None
    duracao_label = ""
    if idade_conjuge > 0:
        for limite_idade, anos in DURACAO_PENSAO_CONJUGE:
            if idade_conjuge <= limite_idade:
                duracao_anos = anos
                break
        if duracao_anos == 0:
            duracao_label = "Vitalicia (conjuge com 44 anos ou mais)"
        elif duracao_anos:
            duracao_label = f"{duracao_anos} anos (conjuge com {idade_conjuge} anos)"

    # Cota por dependente
    cota_individual = valor_pensao / max(num_dependentes, 1)

    return {
        "valor_beneficio_original": round(valor_beneficio, 2),
        "num_dependentes": num_dependentes,
        "cota_percentual": cota_percentual,
        "valor_pensao": round(valor_pensao, 2),
        "cota_individual": round(cota_individual, 2),
        "idade_conjuge": idade_conjuge,
        "duracao_anos": duracao_anos,
        "duracao_label": duracao_label,
        "tem_menor_21": tem_menor_21,
        "salario_minimo": sm,
        "piso_aplicado": valor_pensao == sm and (valor_beneficio * cota_percentual / 100) < sm,
    }


@router.get("/pensao-morte", response_class=HTMLResponse)
async def pensao_morte_form(request: Request, db: Session = Depends(get_db)):
    """Pensão por morte - form."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)
    return templates.TemplateResponse("app/tools/pensao_morte.html", get_context(request, db))


@router.post("/pensao-morte/calcular", response_class=HTMLResponse)
async def pensao_morte_calcular(
    request: Request,
    db: Session = Depends(get_db),
    valor_beneficio: float = Form(...),
    num_dependentes: int = Form(...),
    idade_conjuge: int = Form(0),
    tem_menor_21: str = Form("nao"),
):
    """Calcular pensão por morte."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    resultado = calcular_pensao_morte(
        valor_beneficio=valor_beneficio,
        num_dependentes=num_dependentes,
        idade_conjuge=idade_conjuge,
        tem_menor_21=(tem_menor_21 == "sim"),
    )

    return templates.TemplateResponse("app/tools/pensao_morte.html", {
        **get_context(request, db),
        "resultado": resultado,
        "form": {
            "valor_beneficio": valor_beneficio,
            "num_dependentes": num_dependentes,
            "idade_conjuge": idade_conjuge,
            "tem_menor_21": tem_menor_21,
        },
    })
