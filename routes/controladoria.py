"""
CaseHub Lite — Controladoria Juridica
Modulo de controle de prazos processuais.
Substitui e melhora o app "Prazo Certo" (Lovable) da instancia legada.

Routes:
    GET  /controladoria              — Dashboard principal
    POST /controladoria/novo-prazo   — Criar novo prazo
    POST /controladoria/buscar-intimacoes — Buscar intimacoes via DataJud
    POST /controladoria/{id}/update   — Inline edit field (Asana-style)
    POST /controladoria/{id}/concluir — Marcar prazo como concluido
    POST /controladoria/{id}/excluir  — Excluir prazo
    GET  /controladoria/export/excel  — Exportar prazos para Excel
    GET  /controladoria/api/stats     — Estatisticas JSON para os cards
"""
from fastapi import APIRouter, Depends, Request, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import bindparam, text
from datetime import date, datetime, timedelta
from typing import Any, List, Optional
import hashlib
import hmac
import logging
import io
import re
import json
import unicodedata

logger = logging.getLogger(__name__)

from core.template_config import templates, PREFIX, inject_org_context
from auth import get_current_user
from config import settings
from models import get_db, Case
from models.tenant import tenant_query
from middleware.permissions import has_permission
from services.prazos_cpc import (
    calcular_prazo,
    calcular_prazo_corrido,
    calcular_prazo_detalhado,
    eh_dia_util,
    prazos_comuns,
    proximo_dia_util,
)
from services.calendario_judicial import inferir_tribunal, normalizar_tribunal

router = APIRouter(prefix="/controladoria", tags=["controladoria"])

CONTROLADORIA_RENDER_LIMIT = 300
CONTROLADORIA_CASE_OPTION_LIMIT = 250
BRAZIL_UFS = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS",
    "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC",
    "SP", "SE", "TO",
}

TRIBUNAL_PATTERNS = {
    "TRT3": "%.5.03.%",
    "TRT1": "%.5.01.%",
    "TRT2": "%.5.02.%",
    "TRF1": "%.4.01.%",
    "TRF2": "%.4.02.%",
    "TRF3": "%.4.03.%",
    "TRF4": "%.4.04.%",
    "TRF5": "%.4.05.%",
    "TRF6": "%.4.06.%",
    "TJMG": "%.8.13.%",
    "TJSP": "%.8.26.%",
    "TJRJ": "%.8.19.%",
}

PRAZOS_CALCULATION_ENGINE_VERSION = "prazos_cpc:v1"
OFFICIAL_INTIMATION_SOURCE_MARKERS = (
    "comunicaapi",
    "pdpj",
    "domicilio judicial",
    "domicilio judicial eletronico",
)

TIPOS_PRODUTIVIDADE = (
    "Pet. Simples",
    "Impugnação/decote RPV",
    "Manif. Laudo",
    "Informa perícia/audiência",
    "Pet. Complexa",
    "Defesas",
    "Réplicas",
    "Recursos",
    "C.razões",
    "Outros",
)


def _tipo_produtividade_key(value: Optional[str]) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_text = "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower()
    ascii_text = ascii_text.replace("&", " e ")
    return re.sub(r"[^a-z0-9]+", " ", ascii_text).strip()


_TIPO_PRODUTIVIDADE_ALIASES: dict[str, str] = {}


def _register_tipo_produtividade(canonical: str, *aliases: str) -> None:
    for label in (canonical, *aliases):
        key = _tipo_produtividade_key(label)
        if key:
            _TIPO_PRODUTIVIDADE_ALIASES[key] = canonical


_register_tipo_produtividade(
    "Pet. Simples",
    "Pet Simples",
    "Peticao Simples",
    "Petição Simples",
    "Manifestacao generica",
    "Manifestação genérica",
    "manifestacao",
)
_register_tipo_produtividade(
    "Impugnação/decote RPV",
    "Impugnacao ou decote RPV",
    "Impugnação ou decote RPV",
    "Impugnacao decote RPV",
    "Impugnação decote RPV",
    "Impugnacao ao Cumprimento de Sentenca",
    "Impugnação ao Cumprimento de Sentença",
    "impugnacao_cumprimento",
    "cumprimento_sentenca",
)
_register_tipo_produtividade(
    "Manif. Laudo",
    "Manif Laudo",
    "Manif. Laudo/Quesitos",
    "Manifestacao Laudo",
    "Manifestação Laudo",
    "Manifestacao de Laudo",
    "Manifestação de Laudo",
)
_register_tipo_produtividade(
    "Informa perícia/audiência",
    "Informa pericia ou audiencia",
    "Informa perícia ou audiência",
    "Informa pericia audiencia",
    "Informa perícia audiência",
)
_register_tipo_produtividade(
    "Pet. Complexa",
    "Pet Complexa",
    "Peticao Complexa",
    "Petição Complexa",
    "Aditamento da inicial tutela de urgencia antecedente",
    "Aditamento da inicial tutela de urgência antecedente",
    "tutela_urgencia",
)
_register_tipo_produtividade(
    "Defesas",
    "Defesa",
    "Contestacao procedimento comum",
    "Contestação procedimento comum",
    "contestacao",
    "Embargos a Execucao",
    "Embargos à Execução",
    "embargos_execucao",
    "Reconvencao",
    "Reconvenção",
    "reconvencao",
)
_register_tipo_produtividade(
    "Réplicas",
    "Replicas",
    "Replica",
    "Réplica",
    "Replica resposta a contestacao",
    "Réplica resposta a contestação",
    "replica",
)
_register_tipo_produtividade(
    "Recursos",
    "Recurso",
    "Recurso de Apelacao",
    "Recurso de Apelação",
    "Recurso Especial STJ",
    "Recurso Extraordinario STF",
    "Recurso Extraordinário STF",
    "Agravo de Instrumento",
    "Agravo Interno",
    "Embargos de Declaracao",
    "Embargos de Declaração",
    "Recurso Ordinario Trabalhista",
    "Recurso Ordinário Trabalhista",
    "recurso_apelacao",
    "recurso_especial",
    "recurso_extraordinario",
    "agravo_instrumento",
    "agravo_interno",
    "embargos_declaracao",
    "recurso_ordinario",
)
_register_tipo_produtividade(
    "C.razões",
    "C razoes",
    "Contrarrazoes",
    "Contrarrazões",
    "Contrarrazoes de Apelacao",
    "Contrarrazões de Apelação",
    "contrarrazoes",
)
_register_tipo_produtividade(
    "Outros",
    "Outro",
    "Sem tipo",
    "Prazo manual",
    "Manual",
)


def normalizar_tipo_produtividade(tipo_peticao: Optional[str] = None, tipo: Optional[str] = None) -> str:
    """Converte labels antigos/CPC para a taxonomia gerencial da planilha do UsuarioDemo."""
    for candidate in (tipo_peticao, tipo):
        key = _tipo_produtividade_key(candidate)
        if key:
            if key == "sem tipo":
                continue
            return _TIPO_PRODUTIVIDADE_ALIASES.get(key, "Outros")
    return "Outros"


def _aggregate_tipos_produtividade(rows: list) -> tuple[dict[str, int], list[str], int]:
    totals = {label: 0 for label in TIPOS_PRODUTIVIDADE}
    total = 0
    for row in rows:
        qtd = int(getattr(row, "qtd", 0) or 0)
        label = normalizar_tipo_produtividade(
            getattr(row, "tipo_peticao", None),
            getattr(row, "tipo", None),
        )
        totals[label] += qtd
        total += qtd
    dist = {label: totals[label] for label in TIPOS_PRODUTIVIDADE if totals[label] > 0}
    return dist, list(dist.keys()), total


def _initials(name: Optional[str]) -> str:
    parts = [p for p in re.split(r"\s+", (name or "").strip()) if p]
    if not parts:
        return "?"
    return "".join(part[0].upper() for part in parts[:2])


def _user_key(name: Optional[str]) -> str:
    raw = (name or "").strip()
    raw = re.sub(r"^(dr\.?|dra\.?|doutor|doutora)\s+", "", raw, flags=re.IGNORECASE)
    normalized = unicodedata.normalize("NFKD", raw)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", normalized).strip().lower()


def _user_payload(row) -> dict:
    return {
        "id": row.id,
        "name": row.name or "",
        "initials": _initials(row.name),
        "photo_url": row.photo_url or "",
        "color": row.color or "#1C2447",
    }


def _parse_oab_number(value: Optional[str]) -> tuple[str, str]:
    """Return (number, uf) from common OAB formats without logging the raw value."""
    raw = (value or "").strip().upper()
    if not raw:
        return "", "MG"
    number = "".join(re.findall(r"\d+", raw))
    uf = ""
    for match in re.findall(r"[A-Z]{2}", raw):
        if match in BRAZIL_UFS:
            uf = match
            break
    return number, uf or "MG"


def _get_user_directory(db: Session, org_id: int) -> dict:
    rows = db.execute(
        text("""
            SELECT id, name, email, color, photo_url
            FROM users
            WHERE org_id = :org_id AND enabled = TRUE
            ORDER BY name
        """),
        {"org_id": org_id},
    ).fetchall()
    by_id = {}
    by_name = {}
    users = []
    for row in rows:
        payload = _user_payload(row)
        by_id[payload["id"]] = payload
        key = _user_key(payload["name"])
        if key:
            by_name[key] = payload
        users.append(payload)
    return {"users": users, "by_id": by_id, "by_name": by_name}


def _resolve_responsavel(directory: dict, user_id: Optional[int], name: Optional[str]) -> Optional[dict]:
    if user_id and user_id in directory.get("by_id", {}):
        return directory["by_id"][user_id]
    key = _user_key(name)
    if key:
        return directory.get("by_name", {}).get(key)
    return None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _business_days_between(start: date, end: date, estado: str = "MG", tribunal: Optional[str] = None) -> int:
    """Count business days (Mon-Fri) between start and end (signed).
    Positive = end is in the future, negative = end is in the past."""
    if start == end:
        return 0
    sign = 1 if end >= start else -1
    a, b = (start, end) if sign == 1 else (end, start)
    count = 0
    current = a
    while current < b:
        current += timedelta(days=1)
        if eh_dia_util(current, estado, tribunal=tribunal):
            count += 1
    return count * sign


def _get_org_id(request: Request) -> int:
    """Extract org_id from request state (set by TenantMiddleware)."""
    return getattr(request.state, "org_id", 1)


# UsuarioDemo 03/06: "horário específico para vencimento de prazos processuais".
# Regex p/ validar "HH:MM" (00:00–23:59). Mesmo formato do due_time do Kanban.
_HORA_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")


def _ensure_controladoria_schema(db):
    """Schema aditivo idempotente p/ Controladoria — mantém DBs antigas do alpha
    compatíveis sem migração manual (mesmo padrão de _ensure_kanban_schema).
    Roda nos caminhos lazy que leem/escrevem prazos."""
    bind = db.get_bind()
    dialect = bind.dialect.name if bind is not None else "sqlite"
    additions = [
        ("prazos_processuais", "processo_override", "VARCHAR"),
        ("prazos_processuais", "cliente_override", "VARCHAR"),
        # hora_vencimento: horário do prazo ("HH:MM", nullable). UsuarioDemo 03/06.
        ("prazos_processuais", "hora_vencimento", "VARCHAR(5)"),
        # dias_corridos: prazo administrativo contado em dias corridos (não úteis).
        # Reunião PessoaDemo/UsuarioDemo 10/06 — processos administrativos (INSS etc).
        ("prazos_processuais", "dias_corridos", "BOOLEAN DEFAULT FALSE"),
        # parte_contraria_override: permite editar parte contrária em prazos avulsos
        # (sem case_id). Análogo a processo_override/cliente_override. 11/06.
        ("prazos_processuais", "parte_contraria_override", "VARCHAR"),
        ("prazos_processuais", "source_provider", "VARCHAR(120)"),
        ("prazos_processuais", "source_status", "VARCHAR(50) DEFAULT 'manual'"),
        ("prazos_processuais", "source_reference", "VARCHAR(255)"),
        ("prazos_processuais", "source_url", "TEXT"),
        ("prazos_processuais", "source_payload_hash", "VARCHAR(64)"),
        ("prazos_processuais", "source_fetched_at", "TIMESTAMP"),
        ("prazos_processuais", "source_version", "VARCHAR(80)"),
        ("prazos_processuais", "official_source", "BOOLEAN DEFAULT FALSE"),
        ("prazos_processuais", "calculation_engine_version", "VARCHAR(80)"),
        ("prazos_processuais", "calculation_notes", "TEXT"),
    ]
    for table, column, definition in additions:
        try:
            if dialect == "postgresql":
                exists = db.execute(
                    text("""
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = :table AND column_name = :column
                    """),
                    {"table": table, "column": column},
                ).first()
            else:
                exists = any(row[1] == column for row in db.execute(text(f"PRAGMA table_info({table})")).fetchall())
            if not exists:
                db.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {definition}"))
                db.commit()
        except Exception:
            db.rollback()


def _get_org_id_strict(request: Request):
    """org_id SEM fallback — para escrita sensivel (meta). Ruling 2026-06-03 (Sentinela):
    proibido gravar em org 1 por inercia; retorna None se ausente -> handler responde 400."""
    return getattr(request.state, "org_id", None)


# Meta mensal de produtividade: por org em organizations.settings JSONB (existente —
# ruling 2026-06-03/Janitor: reusar o store JSONB, NAO criar tabela org_settings que
# nao existe no alpha). Sugestao/original (UsuarioDemo 02/06) = 100.
META_KEY = "controladoria_meta_mensal"
META_SUGERIDA = 100


def _org_settings(db: Session, org_id) -> dict:
    """Le organizations.settings JSONB como dict (org-scoped)."""
    raw = db.execute(
        text("SELECT settings FROM organizations WHERE id = :oid"),
        {"oid": org_id},
    ).scalar()
    if raw is None:
        return {}
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            return {}
    return dict(raw) if raw else {}


def _get_meta(db: Session, org_id: int) -> int:
    """Le a meta mensal da org de organizations.settings JSONB (fallback = META_SUGERIDA)."""
    try:
        v = int(str(_org_settings(db, org_id).get(META_KEY, "")).strip() or 0)
        if 1 <= v <= 100000:
            return v
    except Exception:
        pass
    return META_SUGERIDA


DRAG_KEY = "controladoria_drag_enabled"


def _get_drag_enabled(db: Session, org_id) -> bool:
    """Drag-and-drop de prazos é nativo e ligado por padrão; a org pode
    desligar pelas configurações (Equipe CaseHub 2026-06-15). Persistido em
    organizations.settings JSONB (sem migração)."""
    try:
        val = _org_settings(db, org_id).get(DRAG_KEY, True)
    except Exception:
        return True
    if isinstance(val, str):
        return val.strip().lower() not in ("false", "0", "no", "off", "")
    return bool(val)


def _utc_now_iso() -> str:
    """Return a compact UTC timestamp for operator-facing diagnostics."""
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _safe_api_error(value: Any) -> str:
    """Keep upstream diagnostics useful without risking token/secret leakage."""
    if not value:
        return ""
    text_value = str(value)
    replacements = [
        (r"(?i)(Bearer)\s+[^\s,;]+", r"\1 <redacted>"),
        (r"(?i)(access_token|refresh_token|client_secret)(\s*['\"]?\s*[:=]\s*['\"]?)([^\s'\",}&]+)", r"\1\2<redacted>"),
        (r"eyJ[A-Za-z0-9_\-]{20,}\.eyJ[A-Za-z0-9_\-]{20,}\.[A-Za-z0-9_\-]+", "<redacted-jwt>"),
    ]
    for pattern, replacement in replacements:
        text_value = re.sub(pattern, replacement, text_value, flags=re.IGNORECASE)
    return text_value[:160]


def _normalize_ascii(value: Any) -> str:
    text_value = unicodedata.normalize("NFKD", str(value or ""))
    return text_value.encode("ascii", "ignore").decode("ascii").lower()


def _is_official_intimation_source(value: Any) -> bool:
    source = value
    if isinstance(value, dict):
        source = value.get("source") or value.get("provider") or value.get("fonte") or ""
    normalized = _normalize_ascii(source)
    if not normalized or "demo" in normalized or "mock" in normalized:
        return False
    return any(marker in normalized for marker in OFFICIAL_INTIMATION_SOURCE_MARKERS)


def _deadline_payload_hash(item: dict) -> str:
    safe_item = {
        key: value
        for key, value in dict(item or {}).items()
        if _normalize_ascii(key) not in {
            "access_token",
            "refresh_token",
            "client_secret",
            "authorization",
            "source_signature",
        }
    }
    payload = json.dumps(safe_item, sort_keys=True, default=str, ensure_ascii=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _deadline_source_signature(org_id: int, item: dict) -> str:
    payload_hash = _deadline_payload_hash(item)
    message = f"{int(org_id)}:{payload_hash}".encode("utf-8")
    key = settings.SECRET_KEY.encode("utf-8")
    return hmac.new(key, message, hashlib.sha256).hexdigest()


def _attach_deadline_source_signatures(items: list, org_id: int) -> list:
    signed = []
    for raw in items or []:
        item = dict(raw or {})
        if _is_official_intimation_source(item) and item.get("importable", True) is not False:
            item["source_signature"] = _deadline_source_signature(org_id, item)
        signed.append(item)
    return signed


def _valid_deadline_source_signature(item: dict, org_id: int) -> bool:
    provided = str(item.get("source_signature") or "")
    if not provided:
        return False
    expected = _deadline_source_signature(org_id, item)
    return hmac.compare_digest(provided, expected)


def _deadline_source_metadata(item: dict) -> dict:
    source = item.get("source") or item.get("provider") or item.get("fonte") or ""
    importable = item.get("importable", True) is not False
    official = importable and _is_official_intimation_source(source)
    return {
        "source_provider": (source or "desconhecida")[:120],
        "source_status": "official" if official else "manual_review_required",
        "official_source": official,
        "source_reference": str(
            item.get("id")
            or item.get("numero_comunicacao")
            or item.get("numeroComunicacao")
            or item.get("communication_id")
            or item.get("numero_processo")
            or ""
        )[:255],
        "source_payload_hash": _deadline_payload_hash(item),
        "source_version": str(item.get("source_version") or item.get("version") or "")[:80] or None,
    }


def _item_first(item: dict, *keys: str) -> Any:
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return value
    return ""


def _provider_attempt(
    provider: str,
    status: str,
    reason: str,
    *,
    error: Any = "",
    count: int = 0,
) -> dict:
    return {
        "provider": provider,
        "status": status,
        "reason": reason,
        "error": _safe_api_error(error),
        "count": count,
        "attempted_at": _utc_now_iso(),
    }


def _default_tribunal_for_uf(uf: str) -> str:
    return {
        "AC": "TJAC",
        "AL": "TJAL",
        "AM": "TJAM",
        "AP": "TJAP",
        "BA": "TJBA",
        "CE": "TJCE",
        "DF": "TJDFT",
        "ES": "TJES",
        "GO": "TJGO",
        "MA": "TJMA",
        "MG": "TJMG",
        "MS": "TJMS",
        "MT": "TJMT",
        "PA": "TJPA",
        "PB": "TJPB",
        "PE": "TJPE",
        "PI": "TJPI",
        "PR": "TJPR",
        "RJ": "TJRJ",
        "RN": "TJRN",
        "RO": "TJRO",
        "RR": "TJRR",
        "RS": "TJRS",
        "SC": "TJSC",
        "SE": "TJSE",
        "SP": "TJSP",
        "TO": "TJTO",
    }.get(uf.upper(), "TJMG")


def _extract_nested_name(value: Any) -> str:
    if isinstance(value, dict):
        return value.get("nome") or value.get("name") or ""
    return str(value or "")


def _date_only(value: Any) -> str:
    if not value:
        return ""
    return str(value)[:10]


def _normalize_process_fallback_item(item: dict, provider: str, tribunal: str) -> dict:
    numero = item.get("numeroProcesso") or item.get("numero_cnj") or item.get("numero_processo") or ""
    orgao = _extract_nested_name(item.get("orgaoJulgador")) or item.get("orgao") or ""
    data_ref = item.get("dataHoraUltimaAtualizacao") or item.get("dataAjuizamento") or item.get("ultima_atualizacao")
    return {
        "id": f"{provider.lower().replace(' ', '-')}-{numero or hash(str(item))}",
        "numero_processo": numero,
        "tribunal": item.get("tribunal") or item.get("siglaTribunal") or tribunal,
        "orgao": orgao,
        "tipo_comunicacao": "Processo localizado",
        "texto": (
            "Resultado subsidiario por processo/OAB. PDPJ/ComunicaAPI nao retornou "
            "intimacoes autenticadas; confira a fonte antes de importar prazo."
        ),
        "data_disponibilizacao": _date_only(data_ref),
        "source": provider,
        "importable": False,
    }


def _normalize_publication_item(item: dict, provider: str) -> dict:
    texto = item.get("texto") or item.get("conteudo") or item.get("descricao") or ""
    numero = item.get("numero_processo") or item.get("numeroProcesso") or item.get("processo") or ""
    data_ref = item.get("data_disponibilizacao") or item.get("data") or item.get("data_publicacao")
    return {
        "id": f"{provider.lower().replace(' ', '-')}-{item.get('id') or hash(str(item))}",
        "numero_processo": numero,
        "tribunal": item.get("tribunal") or item.get("diario") or "",
        "orgao": item.get("orgao") or item.get("caderno") or "",
        "tipo_comunicacao": item.get("tipo_comunicacao") or item.get("tipo") or "Publicacao",
        "texto": texto,
        "data_disponibilizacao": _date_only(data_ref),
        "source": provider,
        "importable": False,
    }


def _comunicaapi_reason(error_code: str, source: str, message: str = "") -> str:
    if message:
        return message
    if error_code == "missing_credentials":
        return "PDPJ_CLIENT_ID/PDPJ_CLIENT_SECRET ausentes ou nao lidos pelo runtime."
    if error_code == "invalid_client":
        return "PDPJ rejeitou client_id/client_secret; conferir credencial CNJ e client profile."
    if error_code == "unauthorized_client":
        return "O client PDPJ nao esta autorizado para o grant type configurado."
    if error_code == "invalid_grant":
        return "Refresh token PDPJ expirou ou foi revogado; reconexao necessaria."
    if error_code == "no_access_token":
        return "PDPJ nao emitiu access_token; conferir client profile e credenciais no Keycloak."
    if error_code == "auth_failure":
        return "Falha de autenticacao PDPJ; consultar logs sanitizados do VPS."
    if error_code == "network":
        return "Erro de rede ao contactar PDPJ/ComunicaAPI."
    if error_code and error_code.startswith("http_"):
        return f"ComunicaAPI retornou {error_code.replace('_', ' ').upper()}."
    if "demo" in source.lower():
        return "DEMO_MODE ativo; dados mockados."
    return "ComunicaAPI respondeu sem itens para a OAB/periodo informado."


def _controladoria_api_card(
    name: str,
    status: str,
    reason: str,
    detail: str = "",
    badge_text: str = "",
) -> dict:
    return {
        "name": name,
        "status": status,
        "reason": reason,
        "detail": detail,
        "badge_text": badge_text,
    }


def _controladoria_api_status_cards(org_id: Optional[int] = None) -> List[dict]:
    """Operator-facing integration truth for the Controladoria header."""
    cards: List[dict] = []

    try:
        from services.comunicaapi import pdpj_auth

        auth = pdpj_auth.public_status(org_id)
        if auth.get("token_cached"):
            cards.append(_controladoria_api_card(
                "PDPJ/CNJ",
                "ok",
                "Autenticacao PDPJ encontrada neste servidor.",
                "Fonte preferencial para intimacoes oficiais.",
                "Oficial PDPJ",
            ))
        elif auth.get("configured"):
            code = auth.get("last_error_code") or "auth_pending"
            if code == "auth_pending":
                reason = "Credenciais PDPJ existem, mas o token ainda nao foi validado neste servidor."
            else:
                reason = _comunicaapi_reason(code, "ComunicaAPI PJE/CNJ", auth.get("last_error_message") or "")
            cards.append(_controladoria_api_card(
                "PDPJ/CNJ",
                "down",
                reason,
                "Enquanto isso, use Novo Prazo e a busca subsidiaria.",
                "Oficial PDPJ",
            ))
        else:
            cards.append(_controladoria_api_card(
                "PDPJ/CNJ",
                "down",
                "Credenciais PDPJ/CNJ ausentes ou nao lidas pelo runtime.",
                "Sem autorizacao valida, o CNJ nao libera intimacoes por OAB.",
                "Oficial PDPJ",
            ))
    except Exception as exc:
        cards.append(_controladoria_api_card(
            "PDPJ/CNJ",
            "down",
            "Falha ao ler o status local da integracao PDPJ.",
            _safe_api_error(exc),
            "Oficial PDPJ",
        ))

    cards.append(_controladoria_api_card(
        "DataJud",
        "warn",
        "Fallback publico CNJ disponivel para consulta subsidiaria de processos.",
        "Nao substitui a importacao oficial de prazos/intimacoes do PDPJ.",
        "Fallback publico",
    ))

    cards.append(_controladoria_api_card(
        "Escavador",
        "ok" if settings.ESCAVADOR_API_KEY else "down",
        "Chave configurada para busca subsidiaria de publicacoes." if settings.ESCAVADOR_API_KEY else "ESCAVADOR_API_KEY ausente neste servidor.",
        "Usado como apoio quando PDPJ nao retorna intimacoes oficiais.",
        "Provider ativo" if settings.ESCAVADOR_API_KEY else "Chave ausente",
    ))

    cards.append(_controladoria_api_card(
        "JusBrasil",
        "ok" if settings.JUSBRASIL_API_KEY else "down",
        "Chave configurada para diarios e publicacoes." if settings.JUSBRASIL_API_KEY else "JUSBRASIL_API_KEY ausente neste servidor.",
        "Usado como apoio quando fontes oficiais nao retornam resultado.",
        "Provider ativo" if settings.JUSBRASIL_API_KEY else "Chave ausente",
    ))

    return cards


def _is_integration_failure(attempt: dict) -> bool:
    return attempt.get("status") in {"failed", "unavailable"} and bool(attempt.get("error"))


def _failed_status_code(code: str) -> int:
    if code == "missing_credentials":
        return 503
    if code == "unexpected":
        return 500
    return 502


async def _try_comunicaapi_provider(
    numero_oab: str,
    uf_oab: str,
    data_inicio: str,
    data_fim: str,
    org_id: Optional[int] = None,
) -> dict:
    provider = "ComunicaAPI PJE/CNJ"
    try:
        from services.comunicaapi import comunicaapi_client, pdpj_auth
    except ImportError as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "unavailable", "Servico ComunicaAPI nao disponivel.", error=e),
            "auth_status": "unavailable",
            "grant_attempted": "none",
        }

    try:
        resultado = await comunicaapi_client.buscar_por_oab(
            numero_oab,
            uf_oab,
            data_inicio=data_inicio or None,
            data_fim=data_fim or None,
            org_id=org_id,
        )
    except Exception as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "failed", "Erro inesperado na ComunicaAPI.", error=e),
            "auth_status": "error",
            "grant_attempted": pdpj_auth._state(org_id).last_grant_type or "client_credentials",
        }

    items = resultado.get("items", [])
    error_code = resultado.get("error", "")
    source = resultado.get("source", provider)
    status = "ok" if items else ("failed" if error_code else "empty")
    auth_snapshot = pdpj_auth.public_status(org_id)
    auth_status = "configured" if auth_snapshot.get("configured") else "missing_credentials"
    grant_attempted = pdpj_auth._state(org_id).last_grant_type
    if not grant_attempted:
        grant_attempted = "none" if auth_status == "missing_credentials" else "client_credentials"
    for item in items:
        item.setdefault("importable", _is_official_intimation_source(source))
    return {
        "items": items,
        "attempt": _provider_attempt(
            provider,
            status,
            _comunicaapi_reason(error_code, source, resultado.get("message", "")),
            error=error_code,
            count=len(items),
        ),
        "auth_status": auth_status,
        "grant_attempted": grant_attempted,
        "source": source,
    }


async def _try_datajud_provider(numero_oab: str, uf_oab: str) -> dict:
    provider = "DataJud (CNJ)"
    tribunal = _default_tribunal_for_uf(uf_oab)
    try:
        from services.datajud import datajud_client
        resultados = await datajud_client.buscar_por_advogado(numero_oab, tribunal=tribunal)
    except Exception as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "failed", "DataJud falhou na busca por OAB.", error=e),
        }

    items = [_normalize_process_fallback_item(item, provider, tribunal) for item in resultados]
    return {
        "items": items,
        "attempt": _provider_attempt(
            provider,
            "ok" if items else "empty",
            (
                "Busca subsidiaria por processos da OAB; resultados nao substituem "
                "intimacoes autenticadas do PDPJ."
                if items else
                "DataJud nao retornou processos para a OAB informada."
            ),
            count=len(items),
        ),
    }


async def _try_escavador_provider(numero_oab: str, uf_oab: str, data_inicio: str) -> dict:
    provider = "Escavador"
    try:
        from services.escavador import escavador_client
    except ImportError as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "unavailable", "Servico Escavador nao disponivel.", error=e),
        }
    if not getattr(escavador_client, "is_configured", False):
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "skipped", "ESCAVADOR_API_KEY nao configurada."),
        }
    try:
        publicacoes = await escavador_client.buscar_publicacoes(
            oab=f"{uf_oab.upper()}{numero_oab}",
            data_inicio=data_inicio or None,
        )
    except Exception as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "failed", "Escavador falhou na busca de publicacoes.", error=e),
        }
    items = [_normalize_publication_item(item, provider) for item in publicacoes if not item.get("_mock")]
    return {
        "items": items,
        "attempt": _provider_attempt(
            provider,
            "ok" if items else "empty",
            "Busca subsidiaria em publicacoes do Escavador." if items else "Escavador nao retornou publicacoes reais.",
            count=len(items),
        ),
    }


async def _try_jusbrasil_provider(numero_oab: str, uf_oab: str, data_inicio: str) -> dict:
    provider = "JusBrasil"
    try:
        from services.jusbrasil import jusbrasil_client
    except ImportError as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "unavailable", "Servico JusBrasil nao disponivel.", error=e),
        }
    if not getattr(jusbrasil_client, "is_configured", False):
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "skipped", "JUSBRASIL_API_KEY nao configurada."),
        }
    tribunal = _default_tribunal_for_uf(uf_oab)
    try:
        diarios = await jusbrasil_client.buscar_diarios(
            f"OAB {uf_oab.upper()} {numero_oab}",
            data=data_inicio or None,
            tribunal=tribunal,
        )
    except Exception as e:
        return {
            "items": [],
            "attempt": _provider_attempt(provider, "failed", "JusBrasil falhou na busca de diarios.", error=e),
        }
    items = [_normalize_publication_item(item, provider) for item in diarios if not item.get("_mock")]
    return {
        "items": items,
        "attempt": _provider_attempt(
            provider,
            "ok" if items else "empty",
            "Busca subsidiaria em diarios do JusBrasil." if items else "JusBrasil nao retornou publicacoes reais.",
            count=len(items),
        ),
    }


async def _search_intimacoes_oab_chain(
    numero_oab: str,
    uf_oab: str,
    data_inicio: str,
    data_fim: str,
    org_id: Optional[int] = None,
) -> dict:
    """Search intimation sources in a transparent, operator-visible order."""
    attempts = []
    auth_status = "unknown"
    grant_attempted = "none"

    comunica = await _try_comunicaapi_provider(numero_oab, uf_oab, data_inicio, data_fim, org_id=org_id)
    attempts.append(comunica["attempt"])
    auth_status = comunica.get("auth_status", auth_status)
    grant_attempted = comunica.get("grant_attempted", grant_attempted)
    if comunica["items"]:
        signed_items = _attach_deadline_source_signatures(
            comunica["items"],
            int(org_id) if org_id is not None else 1,
        )
        return {
            "items": signed_items,
            "provider": "ComunicaAPI PJE/CNJ",
            "provider_status": "ok",
            "reason": comunica["attempt"]["reason"],
            "last_error": comunica["attempt"]["error"],
            "fallback_active": False,
            "fallback_chain": attempts,
            "auth_status": auth_status,
            "grant_attempted": grant_attempted,
            "source": comunica.get("source", "ComunicaAPI PJE/CNJ"),
            "last_attempt_at": comunica["attempt"]["attempted_at"],
            "code": comunica["attempt"]["error"] or None,
        }

    primary_failed = _is_integration_failure(comunica["attempt"])
    fallback_process_results = None

    if not primary_failed:
        return {
            "items": [],
            "provider": "ComunicaAPI PJE/CNJ",
            "provider_status": "empty",
            "reason": "Consulta executada sem intimacoes no periodo selecionado.",
            "last_error": "",
            "fallback_active": False,
            "fallback_chain": attempts,
            "auth_status": auth_status,
            "grant_attempted": grant_attempted,
            "source": comunica.get("source", "ComunicaAPI PJE/CNJ"),
            "last_attempt_at": comunica["attempt"]["attempted_at"],
            "code": None,
        }

    datajud = await _try_datajud_provider(numero_oab, uf_oab)
    attempts.append(datajud["attempt"])
    if datajud["items"]:
        fallback_process_results = datajud

    escavador = await _try_escavador_provider(numero_oab, uf_oab, data_inicio)
    attempts.append(escavador["attempt"])
    if escavador["items"]:
        return {
            "items": escavador["items"],
            "provider": "Escavador",
            "provider_status": "fallback",
            "reason": "PDPJ/ComunicaAPI falhou; usando publicacoes do Escavador.",
            "last_error": comunica["attempt"]["error"],
            "fallback_active": True,
            "fallback_chain": attempts,
            "auth_status": auth_status,
            "grant_attempted": grant_attempted,
            "source": "Escavador",
            "last_attempt_at": escavador["attempt"]["attempted_at"],
            "code": comunica["attempt"]["error"] or "upstream_failed",
        }

    jusbrasil = await _try_jusbrasil_provider(numero_oab, uf_oab, data_inicio)
    attempts.append(jusbrasil["attempt"])
    if jusbrasil["items"]:
        return {
            "items": jusbrasil["items"],
            "provider": "JusBrasil",
            "provider_status": "fallback",
            "reason": "PDPJ/ComunicaAPI falhou; usando diarios do JusBrasil.",
            "last_error": comunica["attempt"]["error"],
            "fallback_active": True,
            "fallback_chain": attempts,
            "auth_status": auth_status,
            "grant_attempted": grant_attempted,
            "source": "JusBrasil",
            "last_attempt_at": jusbrasil["attempt"]["attempted_at"],
            "code": comunica["attempt"]["error"] or "upstream_failed",
        }

    if fallback_process_results:
        return {
            "items": fallback_process_results["items"],
            "provider": "DataJud (CNJ)",
            "provider_status": "fallback_limited",
            "reason": (
                "PDPJ/ComunicaAPI falhou; DataJud retornou processos da OAB, "
                "mas nao intimações prontas para importacao."
            ),
            "last_error": comunica["attempt"]["error"],
            "fallback_active": True,
            "fallback_chain": attempts,
            "auth_status": auth_status,
            "grant_attempted": grant_attempted,
            "source": "DataJud (CNJ)",
            "last_attempt_at": fallback_process_results["attempt"]["attempted_at"],
            "code": comunica["attempt"]["error"] or "upstream_failed",
        }

    primary_failure_reason = comunica["attempt"].get("reason") or "PDPJ/ComunicaAPI falhou."

    return {
        "items": [],
        "provider": "Nenhuma API",
        "provider_status": "failed",
        "reason": (
            f"{primary_failure_reason} "
            "Fallbacks nao retornaram intimacoes ou resultados para a OAB/periodo informado."
        ),
        "last_error": comunica["attempt"]["error"],
        "fallback_active": True,
        "fallback_chain": attempts,
        "auth_status": auth_status,
        "grant_attempted": grant_attempted,
        "source": None,
        "last_attempt_at": attempts[-1]["attempted_at"] if attempts else _utc_now_iso(),
        "code": comunica["attempt"]["error"] or "integration_failed",
    }


def _get_stats(db: Session, org_id: int) -> dict:
    """Compute stat card values."""
    base = "SELECT COUNT(*) FROM prazos_processuais WHERE org_id = :org_id"
    total = db.execute(text(base), {"org_id": org_id}).scalar() or 0
    concluidos = db.execute(
        text(base + " AND status = 'concluido'"),
        {"org_id": org_id},
    ).scalar() or 0

    ativos = _get_prazos(db, org_id)
    fatais_hoje = sum(1 for p in ativos if p["urgencia"] == "fatal")
    vencidos = sum(1 for p in ativos if p["urgencia"] == "vencido" and p["status"] != "perdido")
    proximos = sum(1 for p in ativos if p["urgencia"] == "amarelo")

    return {
        "total": total,
        # pendentes: prazos ainda não concluídos (hero KPI — FB1 alpha UsuarioDemo).
        # 'total' permanece intacto: eficiencia_geral usa total como denominador.
        "pendentes": max(int(total) - int(concluidos), 0),
        "fatais_hoje": fatais_hoje,
        "vencidos": vencidos,
        "proximos": proximos,
        "concluidos": concluidos,
    }


def _append_prazos_filters(
    query: str,
    params: dict,
    *,
    search: str = "",
    status_filter: str = "",
    mes: str = "",
    tribunal: str = "",
) -> str:
    """Apply dashboard filters shared by list and derived counts."""
    if search:
        query += """
            AND (
                LOWER(c.case_number) LIKE LOWER(:search)
                OR LOWER(c.numero_processo) LIKE LOWER(:search)
                OR LOWER(COALESCE(p.processo_override, '')) LIKE LOWER(:search)
                OR LOWER(COALESCE(p.cliente_override, '')) LIKE LOWER(:search)
                OR LOWER(COALESCE(cl.first_name, '') || ' ' || COALESCE(cl.last_name, '')) LIKE LOWER(:search)
                OR LOWER(COALESCE(c.case_name, '')) LIKE LOWER(:search)
                OR LOWER(COALESCE(p.responsavel, '')) LIKE LOWER(:search)
                OR LOWER(COALESCE(p.tipo, '')) LIKE LOWER(:search)
            )
        """
        params["search"] = f"%{search}%"

    if status_filter and status_filter != "todos":
        query += " AND p.status = :status_filter"
        params["status_filter"] = status_filter
    elif not status_filter:
        query += " AND p.status NOT IN ('concluido')"

    if mes:
        try:
            ano, m = mes.split("-")
            query += " AND EXTRACT(YEAR FROM p.data_vencimento) = :ano AND EXTRACT(MONTH FROM p.data_vencimento) = :mes"
            params["ano"] = int(ano)
            params["mes"] = int(m)
        except (ValueError, AttributeError):
            pass

    # C1 fix: Filtro por tribunal — mapear codigo para pattern do no processo
    if tribunal:
        if tribunal in TRIBUNAL_PATTERNS:
            query += (
                " AND (UPPER(COALESCE(c.tribunal, '')) = :tribunal_code "
                "OR COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE :trib_pat)"
            )
            params["tribunal_code"] = tribunal.upper()
            params["trib_pat"] = TRIBUNAL_PATTERNS[tribunal]
        elif tribunal == "Outro":
            known_pats = " AND ".join(
                f"COALESCE(p.processo_override, c.numero_processo, c.case_number, '') NOT LIKE '{v}'"
                for v in TRIBUNAL_PATTERNS.values()
            )
            query += f" AND (COALESCE(c.tribunal, '') = '' AND {known_pats})"

    return query


def _get_prazos(
    db: Session,
    org_id: int,
    search: str = "",
    status_filter: str = "",
    mes: str = "",
    tribunal: str = "",
    limit: Optional[int] = None,
    sort: str = "",
    direction: str = "asc",
) -> list:
    """Fetch prazos with optional filters, joined with cases."""
    _ensure_controladoria_schema(db)
    query = """
        SELECT
            p.id, p.case_id, p.tipo, p.data_intimacao, p.data_inicio,
            p.data_vencimento, p.hora_vencimento, p.dias_prazo, p.responsavel, p.status,
            p.descricao, p.uf, p.dobro, p.created_at, p.tipo_peticao,
            COALESCE(p.dias_corridos, FALSE) AS dias_corridos,
            p.responsavel_user_id,
            COALESCE(p.source_provider, 'manual') AS source_provider,
            COALESCE(p.source_status, 'manual') AS source_status,
            COALESCE(p.official_source, FALSE) AS official_source,
            p.source_reference,
            p.source_payload_hash,
            p.calculation_engine_version,
            COALESCE(p.ordem, 0) AS ordem,
            COALESCE(p.processo_override, c.numero_processo, c.case_number) AS processo,
            COALESCE(p.cliente_override, TRIM(COALESCE(cl.first_name, '') || ' ' || COALESCE(cl.last_name, '')), c.case_name) AS cliente,
            COALESCE(CASE WHEN p.case_id IS NULL THEN NULLIF(p.parte_contraria_override, '') END, NULLIF(c.polo_passivo, ''), '') AS parte_contraria,
            COALESCE(NULLIF(c.tribunal, ''), '') AS tribunal_cadastrado
        FROM prazos_processuais p
        LEFT JOIN cases c ON p.case_id = c.id AND c.org_id = :org_id
        LEFT JOIN clients cl ON c.client_id = cl.id AND cl.org_id = :org_id
        WHERE p.org_id = :org_id
    """
    params = {"org_id": org_id}

    query = _append_prazos_filters(
        query,
        params,
        search=search,
        status_filter=status_filter,
        mes=mes,
        tribunal=tribunal,
    )

    sort = (sort or "").lower()
    direction_sql = "DESC" if (direction or "").lower() == "desc" else "ASC"
    tribunal_expr = (
        "COALESCE(NULLIF(c.tribunal, ''), "
        "CASE "
        "WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.5.03.%' THEN 'TRT3' "
        "WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.4.06.%' THEN 'TRF6' "
        "WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.4.02.%' THEN 'TRF2' "
        "WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.8.13.%' THEN 'TJMG' "
        "WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.8.26.%' THEN 'TJSP' "
        "ELSE 'Outro' END)"
    )
    sort_exprs = {
        "urgencia": "p.data_vencimento",
        "tribunal": tribunal_expr,
        "processo": "COALESCE(p.processo_override, c.numero_processo, c.case_number, '')",
        "cliente": "COALESCE(p.cliente_override, TRIM(COALESCE(cl.first_name, '') || ' ' || COALESCE(cl.last_name, '')), c.case_name, '')",
        "tipo": "COALESCE(p.tipo, '')",
        "inicio": "p.data_intimacao",
        "dias": "p.dias_prazo",
        "parte_contraria": "COALESCE(c.polo_passivo, '')",
        "observacao": "COALESCE(p.descricao, '')",
        "vencimento": "p.data_vencimento",
        # alias usado pela aba "Concluídos" (sort="data_vencimento", desc).
        "data_vencimento": "p.data_vencimento",
        "responsavel": "COALESCE(p.responsavel, '')",
        "status": "COALESCE(p.status, '')",
    }
    is_pg = (db.get_bind().dialect.name == "postgresql") if db.get_bind() is not None else False
    # Ordenações ancoradas na data de vencimento desempatam pelo horário do
    # prazo (hora_vencimento) — mesma data, hora mais cedo primeiro — e por id.
    # NULLs sempre ao final, independente da direção (UsuarioDemo 2026-06-15).
    venc_keys = {"vencimento", "urgencia", "data_vencimento"}

    def _venc_tiebreak(dir_sql):
        """Colunas de desempate por vencimento: data, depois hora, depois id.
        hora_vencimento NULL **ou vazia** ('') vai sempre ao fim ("sem hora =
        fim do dia") — em PG via NULLIF + NULLS LAST, em SQLite emulado por CASE,
        para que o comportamento seja idêntico nos dois bancos."""
        if is_pg:
            return (
                f" p.data_vencimento {dir_sql} NULLS LAST,"
                f" NULLIF(p.hora_vencimento, '') {dir_sql} NULLS LAST,"
                " p.id ASC"
            )
        return (
            " CASE WHEN p.data_vencimento IS NULL THEN 1 ELSE 0 END ASC,"
            f" p.data_vencimento {dir_sql},"
            " CASE WHEN p.hora_vencimento IS NULL OR p.hora_vencimento = '' THEN 1 ELSE 0 END ASC,"
            f" p.hora_vencimento {dir_sql},"
            " p.id ASC"
        )

    if sort == "manual":
        # Ordem manual (drag-and-drop). Entra-se nela arrastando uma linha; a
        # URL passa a ?sort=manual e sobrevive ao F5. Linhas com a mesma ordem
        # (ou nunca arrastadas) desempatam pelo mesmo critério de vencimento.
        query += " ORDER BY COALESCE(p.ordem, 999999)," + _venc_tiebreak("ASC")
    elif sort in venc_keys:
        query += " ORDER BY" + _venc_tiebreak(direction_sql)
    elif sort in sort_exprs:
        sort_expr = sort_exprs[sort]
        if is_pg:
            query += f" ORDER BY {sort_expr} {direction_sql} NULLS LAST, p.id ASC"
        else:
            query += f" ORDER BY CASE WHEN {sort_expr} IS NULL THEN 1 ELSE 0 END, {sort_expr} {direction_sql}, p.id ASC"
    else:
        # Default (sem sort / sort desconhecido): vencimento mais próximo no
        # topo. NÃO usa p.ordem — o drag precisa ser escolhido explicitamente
        # para a página abrir ordenada por prazo legal (feedback alpha UsuarioDemo).
        query += " ORDER BY" + _venc_tiebreak("ASC")
    if limit is not None:
        query += " LIMIT :limit"
        params["limit"] = limit

    rows = db.execute(text(query), params).fetchall()
    if not rows:
        return []

    hoje = date.today()
    alerta_7dias = hoje + timedelta(days=7)
    prazos = []
    user_dir = _get_user_directory(db, org_id)
    for row in rows:
        venc = row.data_vencimento
        uf = row.uf or "MG"
        processo = row.processo or ""
        tribunal_codigo = normalizar_tribunal(row.tribunal_cadastrado or inferir_tribunal(processo))
        if tribunal_codigo == "CNJ" and not row.tribunal_cadastrado:
            tribunal_codigo = "Outro"
        # Prazo administrativo (dias corridos): sem deslocamento p/ dia útil e
        # dias restantes em dias de calendário. Reunião 10/06.
        eh_corrido = bool(row.dias_corridos)
        venc_efetivo = venc
        prazo_suspenso = False
        if isinstance(venc, date) and not eh_corrido:
            venc_efetivo = proximo_dia_util(venc, uf, tribunal=tribunal_codigo)
            prazo_suspenso = venc_efetivo != venc
        dias_restantes = None
        if isinstance(venc_efetivo, date):
            if eh_corrido:
                dias_restantes = (venc_efetivo - hoje).days
            else:
                dias_restantes = _business_days_between(hoje, venc_efetivo, uf, tribunal_codigo)

        # Urgencia: TODA linha ativa recebe uma cor
        if row.status == "concluido":
            urgencia = "concluido"
        elif row.status == "perdido":
            urgencia = "vencido"  # perdido = vermelho
        elif not isinstance(venc_efetivo, date):
            urgencia = "verde"  # sem data = assume ok
        elif venc_efetivo < hoje:
            urgencia = "vencido"
        elif venc_efetivo == hoje:
            urgencia = "fatal"
        elif venc_efetivo <= alerta_7dias:
            urgencia = "amarelo"  # <=7 dias
        else:
            urgencia = "verde"  # >7 dias

        responsavel_user = _resolve_responsavel(user_dir, row.responsavel_user_id, row.responsavel)

        prazos.append({
            "id": row.id,
            "case_id": row.case_id,
            "processo": processo or "—",
            "cliente": row.cliente or "—",
            "parte_contraria": row.parte_contraria or "—",
            "tipo": row.tipo,
            "data_intimacao": row.data_intimacao,
            "data_inicio": row.data_inicio,
            "data_vencimento": venc,
            "hora_vencimento": row.hora_vencimento or "",
            "data_vencimento_efetiva": venc_efetivo,
            "prazo_suspenso": prazo_suspenso,
            "dias_prazo": row.dias_prazo,
            "dias_restantes": dias_restantes,
            "responsavel": row.responsavel or "—",
            "responsavel_user": responsavel_user,
            "responsavel_user_id": responsavel_user["id"] if responsavel_user else None,
            "status": row.status,
            "descricao": row.descricao or "",
            "uf": uf,
            "tribunal": tribunal_codigo,
            "dobro": row.dobro,
            "dias_corridos": eh_corrido,
            "urgencia": urgencia,
            "tipo_peticao": normalizar_tipo_produtividade(row.tipo_peticao, row.tipo)
            if row.tipo_peticao or row.status == "concluido"
            else "",
            "source_provider": row.source_provider or "manual",
            "source_status": row.source_status or "manual",
            "official_source": bool(row.official_source),
            "source_reference": row.source_reference or "",
            "source_payload_hash": row.source_payload_hash or "",
            "calculation_engine_version": row.calculation_engine_version or "",
            "ordem": row.ordem or 0,
        })

    return prazos


def _get_prazos_vencidos_total(
    db: Session,
    org_id: int,
    search: str = "",
    status_filter: str = "",
    mes: str = "",
    tribunal: str = "",
) -> int:
    """Count overdue deadlines using the same filters and predicate as the table."""
    prazos = _get_prazos(
        db,
        org_id,
        search=search,
        status_filter=status_filter,
        mes=mes,
        tribunal=tribunal,
    )
    return sum(1 for prazo in prazos if prazo.get("urgencia") == "vencido")


def _get_responsaveis(db: Session, org_id: int) -> list:
    """List distinct responsaveis for filter dropdown."""
    rows = db.execute(
        text(
            "SELECT DISTINCT responsavel FROM prazos_processuais "
            "WHERE org_id = :org_id AND responsavel IS NOT NULL "
            "ORDER BY responsavel"
        ),
        {"org_id": org_id},
    ).fetchall()
    return [r[0] for r in rows if r[0]]


def _get_produtividade_setores(db: Session, org_id: int) -> List[dict]:
    rows = db.execute(
        text("""
            SELECT COALESCE(NULLIF(u.name, ''), NULLIF(p.responsavel, ''), 'Sem responsável') AS setor,
                   u.id AS user_id,
                   u.photo_url,
                   u.color,
                   COUNT(*) AS total,
                   SUM(CASE WHEN COALESCE(p.status, 'pendente') NOT IN ('concluido', 'perdido') THEN 1 ELSE 0 END) AS pendentes,
                   SUM(CASE WHEN p.status = 'concluido' THEN 1 ELSE 0 END) AS concluidos,
                   SUM(CASE WHEN p.data_vencimento < :today AND COALESCE(p.status, 'pendente') NOT IN ('concluido', 'perdido') THEN 1 ELSE 0 END) AS vencidos,
                   SUM(CASE WHEN p.data_vencimento >= :today
                             AND p.data_vencimento <= :today_plus_7
                             AND COALESCE(p.status, 'pendente') NOT IN ('concluido', 'perdido') THEN 1 ELSE 0 END) AS proximos
            FROM prazos_processuais p
            LEFT JOIN users u
              ON u.org_id = p.org_id
             AND u.enabled = TRUE
             AND (
                u.id = p.responsavel_user_id
                OR (p.responsavel_user_id IS NULL AND LOWER(TRIM(u.name)) = LOWER(TRIM(COALESCE(p.responsavel, ''))))
             )
            WHERE p.org_id = :org_id
            GROUP BY COALESCE(NULLIF(u.name, ''), NULLIF(p.responsavel, ''), 'Sem responsável'), u.id, u.photo_url, u.color
            ORDER BY vencidos DESC, pendentes DESC, total DESC, setor ASC
            LIMIT 8
        """),
        {"org_id": org_id, "today": date.today(), "today_plus_7": date.today() + timedelta(days=7)},
    ).fetchall()
    data = []
    for row in rows:
        total = int(row.total or 0)
        pendentes = int(row.pendentes or 0)
        concluidos = int(row.concluidos or 0)
        vencidos = int(row.vencidos or 0)
        proximos = int(row.proximos or 0)
        em_dia = max(total - vencidos, 0)
        data.append({
            "setor": row.setor,
            "user_id": row.user_id,
            "photo_url": row.photo_url or "",
            "color": row.color or "#1C2447",
            "initials": _initials(row.setor),
            "total": total,
            "pendentes": pendentes,
            "concluidos": concluidos,
            "vencidos": vencidos,
            "proximos": proximos,
            "eficiencia": round((concluidos / total) * 100) if total else 0,
            "produtividade": em_dia,
            "em_dia": em_dia,
            "status_label": "Atenção" if vencidos else ("Semana crítica" if proximos else "Em dia"),
        })
    return data


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse)
async def controladoria_dashboard(
    request: Request,
    search: str = "",
    status: str = "",
    mes: str = "",
    tribunal: str = "",
    sort: str = "vencimento",
    direction: str = "asc",
    db: Session = Depends(get_db),
):
    """Dashboard principal da Controladoria."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    org_id = _get_org_id(request)
    _ensure_controladoria_schema(db)  # lazy: garante hora_vencimento em DBs antigas
    drag_enabled = _get_drag_enabled(db, org_id)
    stats = _get_stats(db, org_id)
    prazos_render = _get_prazos(
        db,
        org_id,
        search=search,
        status_filter=status,
        mes=mes,
        tribunal=tribunal,
        limit=CONTROLADORIA_RENDER_LIMIT + 1,
        sort=sort,
        direction=direction,
    )
    prazos_truncated = len(prazos_render) > CONTROLADORIA_RENDER_LIMIT
    prazos = prazos_render[:CONTROLADORIA_RENDER_LIMIT]

    # P9: Extrair tribunais disponíveis para o dropdown
    tribunal_rows = db.execute(
        text("""
            SELECT DISTINCT
                COALESCE(NULLIF(c.tribunal, ''), CASE
                    WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.5.03.%' THEN 'TRT3'
                    WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.5.01.%' THEN 'TRT1'
                    WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.4.06.%' THEN 'TRF6'
                    WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.4.02.%' THEN 'TRF2'
                    WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.8.13.%' THEN 'TJMG'
                    WHEN COALESCE(p.processo_override, c.numero_processo, c.case_number, '') LIKE '%.8.26.%' THEN 'TJSP'
                    ELSE 'Outro'
                END) AS tribunal
            FROM prazos_processuais p
            LEFT JOIN cases c ON p.case_id = c.id
            WHERE p.org_id = :org_id AND p.status NOT IN ('concluido')
            ORDER BY 1
        """),
        {"org_id": org_id},
    ).fetchall()
    tribunais_disponiveis = [r.tribunal for r in tribunal_rows if r.tribunal and r.tribunal != 'Outro']
    if any(r.tribunal == 'Outro' for r in tribunal_rows):
        tribunais_disponiveis.append('Outro')

    # Buscar casos para o modal de novo prazo
    cases_query = (
        tenant_query(db, Case, org_id)
        .order_by(Case.case_number.asc())
        .limit(CONTROLADORIA_CASE_OPTION_LIMIT + 1)
        .all()
    )
    cases_truncated = len(cases_query) > CONTROLADORIA_CASE_OPTION_LIMIT
    cases_query = cases_query[:CONTROLADORIA_CASE_OPTION_LIMIT]

    # Prazos vencidos para popup de alerta
    prazos_vencidos = [p for p in prazos if p["urgencia"] == "vencido"]
    if prazos_truncated:
        prazos_vencidos_total = max(
            _get_prazos_vencidos_total(
                db,
                org_id,
                search=search,
                status_filter=status,
                mes=mes,
                tribunal=tribunal,
            ),
            len(prazos_vencidos),
        )
    else:
        prazos_vencidos_total = len(prazos_vencidos)
    prazos_vencidos_remaining = max(prazos_vencidos_total - min(len(prazos_vencidos), 15), 0)

    # Users da org para dropdown de responsável
    org_users = db.execute(
        text("SELECT id, name, email, user_type, color, photo_url, oab_number FROM users WHERE org_id = :org_id AND enabled = TRUE ORDER BY name"),
        {"org_id": org_id},
    ).fetchall()
    users_list = []
    oab_options = []
    for u in org_users:
        oab_numero, oab_uf = _parse_oab_number(getattr(u, "oab_number", "") or "")
        users_list.append({
            "id": u.id,
            "name": u.name,
            "initials": _initials(u.name),
            "photo_url": u.photo_url or "",
            "color": u.color or "#1C2447",
            "oab_number": oab_numero,
            "oab_uf": oab_uf if oab_numero else "",
        })
        if oab_numero:
            oab_options.append({
                "user_id": u.id,
                "name": u.name or "Advogado",
                "numero": oab_numero,
                "uf": oab_uf,
            })

    # Opções pros dropdowns de inline-edit Cliente/Processo (29/05 Equipe CaseHub): clicar
    # a célula abre um <select> com os existentes (tenant-isolado) + "cadastrar novo".
    # cliente_override/processo_override guardam o texto escolhido.
    clientes_opc = [
        {"id": r.id, "name": (r.name or "").strip()}
        for r in db.execute(
            text(
                "SELECT id, TRIM(COALESCE(first_name,'') || ' ' || COALESCE(last_name,'')) AS name "
                "FROM clients WHERE org_id = :org_id ORDER BY first_name, last_name LIMIT 500"
            ),
            {"org_id": org_id},
        ).fetchall()
        if (r.name or "").strip()
    ]
    processos_opc = [
        {"id": c.id, "label": (c.case_number or c.case_name or f"Processo {c.id}")}
        for c in cases_query
    ]

    prazos_concluidos = _get_prazos(
        db,
        org_id,
        status_filter="concluido",
        search=search,
        tribunal=tribunal,
        limit=500,
        sort="data_vencimento",
        direction="desc",
    )

    comuns = prazos_comuns()
    org_ctx = inject_org_context(request)

    return templates.TemplateResponse(
        "app/controladoria/dashboard.html",
        {
            "request": request,
            "user": user,
            "PREFIX": PREFIX,
            "stats": stats,
            "prazos": prazos,
            "prazos_truncated": prazos_truncated,
            "prazos_concluidos": prazos_concluidos,
            "controladoria_render_limit": CONTROLADORIA_RENDER_LIMIT,
            "prazos_vencidos": prazos_vencidos,
            "prazos_vencidos_total": prazos_vencidos_total,
            "prazos_vencidos_remaining": prazos_vencidos_remaining,
            "cases": cases_query,
            "cases_truncated": cases_truncated,
            "controladoria_case_option_limit": CONTROLADORIA_CASE_OPTION_LIMIT,
            "prazos_comuns": comuns,
            "org_users": users_list,
            "oab_options": oab_options,
            "clientes_opc": clientes_opc,
            "processos_opc": processos_opc,
            "tipos_produtividade": TIPOS_PRODUTIVIDADE,
            "produtividade_setores": _get_produtividade_setores(db, org_id),
            "api_status_cards": _controladoria_api_status_cards(org_id),
            "search": search,
            "status_filter": status,
            "mes_filter": mes,
            "tribunal_filter": tribunal,
            "sort_key": sort,
            "sort_direction": "desc" if direction.lower() == "desc" else "asc",
            "drag_enabled": drag_enabled,
            "tribunais_disponiveis": tribunais_disponiveis,
            **org_ctx,
        },
    )


@router.get("/cases/search")
async def buscar_cases_controladoria(
    request: Request,
    q: str = Query("", max_length=80),
    db: Session = Depends(get_db),
):
    """Search tenant cases for the Novo Prazo selector without rendering all options."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)

    term = (q or "").strip()
    if len(term) < 2 and not term.isdigit():
        return {"items": []}

    org_id = _get_org_id(request)
    rows = db.execute(
        text("""
            SELECT
                c.id,
                c.case_number,
                c.numero_processo,
                c.case_name,
                COALESCE(cl.first_name, '') || ' ' || COALESCE(cl.last_name, '') AS cliente
            FROM cases c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.org_id = :org_id
              AND (
                CAST(c.id AS TEXT) = :id_term
                OR LOWER(COALESCE(c.case_number, '')) LIKE LOWER(:term)
                OR LOWER(COALESCE(c.numero_processo, '')) LIKE LOWER(:term)
                OR LOWER(COALESCE(c.case_name, '')) LIKE LOWER(:term)
                OR LOWER(COALESCE(cl.first_name, '') || ' ' || COALESCE(cl.last_name, '')) LIKE LOWER(:term)
              )
            ORDER BY
                CASE
                    WHEN COALESCE(c.case_number, c.numero_processo, '') = '' THEN 1
                    ELSE 0
                END ASC,
                COALESCE(c.case_number, c.numero_processo, '') ASC,
                c.id ASC
            LIMIT 20
        """),
        {"org_id": org_id, "id_term": term if term.isdigit() else "", "term": f"%{term}%"},
    ).fetchall()

    items = []
    for row in rows:
        numero = row.case_number or row.numero_processo or "Sem numero"
        nome = row.case_name or "Sem nome"
        cliente = (row.cliente or "").strip()
        label = f"{numero} - {nome}"
        if cliente:
            label = f"{label} ({cliente})"
        items.append({"id": row.id, "label": label})

    return {"items": items}


def _notify_prazo_assignee_dm(
    db: Session,
    org_id: int,
    actor_id: int,
    assignee_id: int,
    tipo: str,
    data_vencimento,
    hora_vencimento=None,
) -> None:
    if not assignee_id or assignee_id == actor_id:
        return
    try:
        from routes.team_messages import post_system_dm_to_user
        title = str(tipo or "prazo").strip()[:160]
        due = str(data_vencimento or "").strip()
        hour = str(hora_vencimento or "").strip()[:5]
        if hour:
            due = f"{due} {hour}".strip()
        due_label = f" Vence em {due}." if due else ""
        post_system_dm_to_user(
            db,
            int(org_id),
            int(assignee_id),
            f"Controladoria: prazo \"{title}\" foi designado(a) para você.{due_label}",
        )
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass


@router.post("/novo-prazo")
async def criar_prazo(request: Request, db: Session = Depends(get_db)):
    """Criar novo prazo processual."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    # Ruling 2026-06-03 (CWE-862): exige cargo com escrita (gestor/advogado/estagiario),
    # nao basta estar logado. has_permission via cookie auth (require_permission e bearer).
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada para criar prazo"}, status_code=403)

    org_id = _get_org_id(request)
    _ensure_controladoria_schema(db)

    try:
        data = await request.json()
    except Exception:
        # Fallback: form data
        form = await request.form()
        data = dict(form)

    case_id = data.get("case_id")
    processo_manual = data.get("processo_manual", "").strip()
    cliente_manual = data.get("cliente_manual", "").strip()
    parte_contraria_manual = data.get("parte_contraria_manual", "").strip()

    # Prazo avulso: o advogado pode informar apenas o número do processo,
    # sem obrigar cadastro completo em cases.
    processo_override = processo_manual or None
    if case_id == "__manual__":
        case_id = None

    tipo = data.get("tipo", "").strip()
    data_intimacao_str = data.get("data_intimacao", "")
    dias_prazo = data.get("dias_prazo")
    responsavel = data.get("responsavel", "").strip()
    responsavel_user_id = data.get("responsavel_user_id") or data.get("responsavel_id")
    uf = data.get("uf", "MG").upper()
    descricao = data.get("descricao", "").strip()
    dobro = data.get("dobro", False)

    if isinstance(dobro, str):
        dobro = dobro.lower() in ("true", "1", "on", "sim")

    # Reunião 10/06 (PessoaDemo/UsuarioDemo) — prazo administrativo em dias corridos
    # (INSS etc): conta dias de calendário, sem dia útil/feriado/tribunal.
    dias_corridos = data.get("dias_corridos", False)
    if isinstance(dias_corridos, str):
        dias_corridos = dias_corridos.lower() in ("true", "1", "on", "sim")

    # Tribunal explícito (reunião 10/06): UF sozinha é ambígua (MG -> TJMG/TRT3/TRF6).
    # Quando informado, tem prioridade sobre a inferência por número/UF.
    tribunal_param = (data.get("tribunal") or "").strip()

    # UsuarioDemo 02/06 (C11) — "enviar pro chat" opcional + DESATIVAVEL (default = nao avisa).
    # Checkbox de form chega como "on"; JSON pode mandar true/1. Ausente = false.
    notificar_chat = data.get("notificar_chat", False)
    if isinstance(notificar_chat, str):
        notificar_chat = notificar_chat.lower() in ("true", "1", "on", "sim")

    try:
        responsavel_user_id = int(responsavel_user_id) if responsavel_user_id else None
    except (TypeError, ValueError):
        responsavel_user_id = None
    if responsavel_user_id:
        user_row = db.execute(
            text("SELECT id, name FROM users WHERE id = :id AND org_id = :org_id AND enabled = TRUE"),
            {"id": responsavel_user_id, "org_id": org_id},
        ).fetchone()
        if user_row:
            responsavel = user_row.name
        else:
            responsavel_user_id = None

    if not tipo:
        return JSONResponse({"error": "Tipo de prazo e obrigatorio"}, status_code=400)
    if not data_intimacao_str:
        return JSONResponse({"error": "Data de intimacao e obrigatoria"}, status_code=400)

    # UsuarioDemo 03/06: horário opcional do vencimento ("HH:MM"). Vazio = sem hora.
    hora_vencimento = (data.get("hora_vencimento") or "").strip()
    if hora_vencimento and not _HORA_RE.match(hora_vencimento):
        return JSONResponse({"error": "Horario invalido (use HH:MM)"}, status_code=400)
    hora_vencimento = hora_vencimento or None

    try:
        data_intimacao = date.fromisoformat(data_intimacao_str)
    except ValueError:
        return JSONResponse({"error": "Data de intimacao invalida (use AAAA-MM-DD)"}, status_code=400)

    # Determinar dias do prazo
    if not dias_prazo:
        comuns = prazos_comuns()
        if tipo in comuns:
            dias_prazo = comuns[tipo]["dias"]
        else:
            dias_prazo = 15  # padrao
    else:
        try:
            dias_prazo = int(dias_prazo)
        except (TypeError, ValueError):
            return JSONResponse({"error": "Dias do prazo deve ser um numero"}, status_code=400)

    case_tribunal = ""
    case_processo = ""
    if case_id:
        case_row = db.execute(
            text("""
                SELECT tribunal, numero_processo, case_number
                FROM cases
                WHERE id = :case_id AND org_id = :org_id
            """),
            {"case_id": int(case_id), "org_id": org_id},
        ).fetchone()
        if case_row:
            case_tribunal = case_row.tribunal or ""
            case_processo = case_row.numero_processo or case_row.case_number or ""
    processo_ref = processo_override or case_processo
    tribunal_inferido = inferir_tribunal(processo_ref)
    # Tribunal explícito do formulário > tribunal do case > inferência por número.
    tribunal_codigo = normalizar_tribunal(tribunal_param or case_tribunal or tribunal_inferido)
    if tribunal_codigo == "CNJ" and not (tribunal_param or case_tribunal) and tribunal_inferido == "CNJ":
        tribunal_codigo = "Outro"

    # Calcular datas
    if dias_corridos:
        # Administrativo: dias corridos (calendário); começa no dia seguinte.
        data_inicio = data_intimacao + timedelta(days=1)
        data_vencimento = calcular_prazo_corrido(data_intimacao, dias_prazo)
    else:
        data_inicio = proximo_dia_util(data_intimacao + timedelta(days=1), uf, tribunal=tribunal_codigo)
        data_vencimento = calcular_prazo(data_intimacao, dias_prazo, uf, dobro, tribunal=tribunal_codigo)

    # Inserir no banco
    try:
        db.execute(
            text("""
                INSERT INTO prazos_processuais
                    (case_id, org_id, tipo, data_intimacao, data_inicio, data_vencimento, hora_vencimento,
                     dias_prazo, responsavel, responsavel_user_id, status, descricao, uf, dobro, dias_corridos,
                     processo_override, cliente_override, parte_contraria_override,
                     source_provider, source_status, official_source,
                     calculation_engine_version, calculation_notes)
                VALUES
                    (:case_id, :org_id, :tipo, :data_intimacao, :data_inicio, :data_vencimento, :hora_vencimento,
                     :dias_prazo, :responsavel, :responsavel_user_id, 'pendente', :descricao, :uf, :dobro, :dias_corridos,
                     :processo_override, :cliente_override, :parte_contraria_override,
                     :source_provider, :source_status, :official_source,
                     :calculation_engine_version, :calculation_notes)
            """),
            {
                "case_id": int(case_id) if case_id else None,
                "org_id": org_id,
                "tipo": tipo,
                "data_intimacao": data_intimacao,
                "data_inicio": data_inicio,
                "data_vencimento": data_vencimento,
                "hora_vencimento": hora_vencimento,
                "dias_prazo": dias_prazo,
                "responsavel": responsavel or None,
                "responsavel_user_id": responsavel_user_id,
                "descricao": descricao or None,
                "uf": uf,
                "dobro": dobro,
                "dias_corridos": dias_corridos,
                "processo_override": processo_override,
                "cliente_override": cliente_manual or None,
                "parte_contraria_override": parte_contraria_manual or None,
                "source_provider": "manual",
                "source_status": "manual",
                "official_source": False,
                "calculation_engine_version": PRAZOS_CALCULATION_ENGINE_VERSION,
                "calculation_notes": "Prazo criado manualmente pelo operador.",
            },
        )
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error("Erro ao criar prazo: %s", e)
        return JSONResponse({"error": f"Erro ao salvar: {str(e)}"}, status_code=500)

    logger.info(
        "Prazo criado: tipo=%s, intimacao=%s, vencimento=%s, responsavel=%s",
        tipo, data_intimacao, data_vencimento, responsavel,
    )
    _notify_prazo_assignee_dm(
        db,
        org_id,
        user.id,
        responsavel_user_id,
        tipo,
        data_vencimento,
        hora_vencimento,
    )

    # UsuarioDemo 02/06 (C11) — opcao "enviar pro chat": ao criar o prazo, avisa a equipe.
    # DESATIVAVEL: so' roda quando notificar_chat=true. Best-effort (nunca derruba o
    # fluxo). Reusa o sininho (Notification + poller que ja' toca som) e o chat real
    # (#equipe via team_messages), org-scoped por construcao.
    if notificar_chat:
        proc_label = (processo_override or "").strip()
        resp_label = (responsavel or "").strip()
        titulo = "Novo prazo criado"
        partes = [f"Prazo '{tipo}'"]
        if proc_label:
            partes.append(f"no processo {proc_label}")
        if resp_label:
            partes.append(f"para {resp_label}")
        partes.append(f"— vence em {data_vencimento.isoformat()}")
        msg = " ".join(partes)

        # (1) Sininho + som: notificacao in-app para toda a equipe da org.
        try:
            from services.notifications.in_app import create_notification_for_all_staff
            create_notification_for_all_staff(
                db=db,
                title=titulo,
                notification_type="deadline_approaching",
                message=msg,
                severity="info",
                org_id=org_id,
                action_url=f"{PREFIX}/controladoria",
            )
            db.commit()
        except Exception as e:
            db.rollback()
            logger.error("Falha ao notificar equipe sobre novo prazo: %s", e)

        # (2) Chat de equipe: posta no #equipe (sem duplicar o sistema de chat).
        try:
            from routes.team_messages import post_system_message_to_equipe
            post_system_message_to_equipe(db, org_id, user.id, f"{titulo}: {msg}")
        except Exception as e:
            logger.error("Falha ao postar novo prazo no chat de equipe: %s", e)

    # Se veio de JSON, retorna JSON. Se form, redireciona.
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        return JSONResponse({
            "success": True,
            "prazo": {
                "tipo": tipo,
                "data_intimacao": data_intimacao.isoformat(),
                "data_inicio": data_inicio.isoformat(),
                "data_vencimento": data_vencimento.isoformat(),
                "dias_prazo": dias_prazo,
                "responsavel": responsavel,
                "responsavel_user_id": responsavel_user_id,
                "processo": processo_override,
                "cliente": cliente_manual,
                "parte_contraria": parte_contraria_manual,
                "tribunal": tribunal_codigo,
            },
        })
    return RedirectResponse(url=f"{PREFIX}/controladoria", status_code=303)


@router.post("/buscar-intimacoes")
async def buscar_intimacoes(request: Request, db: Session = Depends(get_db)):
    """Buscar intimacoes via DataJud para os processos monitorados."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)

    org_id = _get_org_id(request)

    try:
        data = await request.json()
    except Exception:
        data = {}

    data_inicio_str = data.get("data_inicio", "")
    data_fim_str = data.get("data_fim", "")

    # Buscar processos monitorados (cases com case_number preenchido)
    cases = tenant_query(db, Case, org_id).filter(
        Case.case_number.isnot(None),
        Case.case_number != "",
    ).all()

    if not cases:
        return JSONResponse({
            "success": True,
            "message": "Nenhum processo monitorado encontrado",
            "intimacoes": [],
        })

    # Tentar importar DataJud
    try:
        from services.datajud import datajud_client
    except ImportError:
        return JSONResponse({
            "error": "Servico DataJud nao disponivel",
        }, status_code=503)

    # Incident-class fix (2026-07-01 outage pattern — see
    # buscar_comunicaapi/api_datajud_busca below): extract the plain values
    # this loop needs BEFORE closing the session, then release it before the
    # (up to N-cases * slow) DataJud round-trips. Otherwise the request sits
    # idle-in-transaction holding a lock on `users`/`cases` for the entire
    # loop duration.
    case_rows = [
        {"case_id": case.id, "case_name": case.case_name or "", "case_number": case.case_number}
        for case in cases
    ]
    db.close()

    intimacoes = []
    erros = []

    for row in case_rows:
        numero = row["case_number"]
        if not numero or len(numero.replace("-", "").replace(".", "")) < 15:
            continue  # Pular numeros que nao parecem ser CNJ

        try:
            movimentacoes = await datajud_client.get_movimentacoes(numero)
            for mov in movimentacoes:
                nome_mov = mov.get("nome", "").lower()
                # Filtrar por intimacoes e citacoes
                if any(kw in nome_mov for kw in ["intimacao", "intimação", "citacao", "citação", "notificacao", "notificação"]):
                    data_hora = mov.get("dataHora", "")
                    # Filtrar por periodo se informado
                    if data_inicio_str and data_hora < data_inicio_str:
                        continue
                    if data_fim_str and data_hora > data_fim_str + "T23:59:59":
                        continue

                    intimacoes.append({
                        "processo": numero,
                        "case_id": row["case_id"],
                        "case_name": row["case_name"],
                        "tipo": mov.get("nome", ""),
                        "data": data_hora[:10] if data_hora else "",
                        "complemento": ", ".join(
                            c.get("descricao", "")
                            for c in mov.get("complementosTabelados", [])
                        ),
                    })
        except Exception as e:
            erros.append({"processo": numero, "erro": str(e)})
            logger.warning("Erro ao buscar intimacoes do processo %s: %s", numero, e)

    intimacoes.sort(key=lambda x: x.get("data", ""), reverse=True)

    return JSONResponse({
        "success": True,
        "total": len(intimacoes),
        "intimacoes": intimacoes,
        "erros": erros,
        "processos_consultados": len(case_rows),
    })


@router.post("/buscar-comunicaapi")
async def buscar_comunicaapi(request: Request, db: Session = Depends(get_db)):
    """Buscar intimacoes via ComunicaAPI PJE/CNJ por numero OAB.
    Mais eficiente que buscar processo a processo: uma unica chamada retorna tudo."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)

    try:
        data = await request.json()
    except Exception:
        data = {}

    numero_oab = data.get("numero_oab", "").strip()
    uf_oab = data.get("uf_oab", "MG").strip().upper()
    data_inicio = data.get("data_inicio", "")
    data_fim = data.get("data_fim", "")
    org_id = _get_org_id(request)

    if not numero_oab:
        return JSONResponse({"error": "Numero OAB obrigatorio"}, status_code=400)

    # Incident-class fix (2026-07-01 outage pattern): the ComunicaAPI/PDPJ
    # chain below can take up to ~2min across providers + fallbacks.
    # _search_intimacoes_oab_chain takes only plain args (no db), so release
    # the session before the await — it transparently reopens a connection
    # for the client-matching queries after the chain returns.
    db.close()

    try:
        resultado = await _search_intimacoes_oab_chain(numero_oab, uf_oab, data_inicio, data_fim, org_id=org_id)
        intimacoes = resultado.get("items", [])
        failed = resultado.get("provider_status") == "failed"
        payload = {
            "success": not failed,
            "code": resultado.get("code"),
            "message": resultado.get("reason"),
            "error": resultado.get("reason") if failed else None,
            "total": len(intimacoes),
            "intimacoes": intimacoes,
            "source": resultado.get("source") or resultado.get("provider"),
            "provider": resultado.get("provider"),
            "provider_status": resultado.get("provider_status"),
            "reason": resultado.get("reason"),
            "last_error": resultado.get("last_error"),
            "last_attempt_at": resultado.get("last_attempt_at"),
            "fallback_active": resultado.get("fallback_active", False),
            "fallback_chain": resultado.get("fallback_chain", []),
            "auth_status": resultado.get("auth_status"),
            "grant_attempted": resultado.get("grant_attempted"),
        }
        if not failed:
            payload.pop("error", None)
        status_code = _failed_status_code(str(resultado.get("code") or "")) if failed else 200
        return JSONResponse(payload, status_code=status_code)
    except Exception as e:
        logger.error("ComunicaAPI search error: %s", e, exc_info=True)
        return JSONResponse({
            "success": False,
            "code": "unexpected",
            "message": "Falha interna ao executar a cadeia de APIs.",
            "error": "Falha interna ao executar a cadeia de APIs.",
            "total": 0,
            "intimacoes": [],
            "provider": "Controladoria",
            "provider_status": "failed",
            "reason": "Falha interna ao executar a cadeia de APIs.",
            "last_error": _safe_api_error(e),
            "fallback_active": False,
            "fallback_chain": [],
        }, status_code=500)


@router.post("/importar-intimacoes")
async def importar_intimacoes(request: Request, db: Session = Depends(get_db)):
    """Importar intimacoes selecionadas do ComunicaAPI como novos prazos processuais."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada"}, status_code=403)

    org_id = _get_org_id(request)

    try:
        data = await request.json()
    except Exception:
        return JSONResponse({"error": "JSON invalido"}, status_code=400)

    intimacoes = data.get("intimacoes", [])
    if not intimacoes:
        return JSONResponse({"error": "Nenhuma intimacao enviada"}, status_code=400)

    imported = 0
    skipped = 0
    blocked = 0

    for item in intimacoes:
        numero_processo = str(_item_first(
            item,
            "numero_processo",
            "numeroProcesso",
            "numeroprocessocommascara",
            "numeroProcessoComMascara",
        )).strip()
        texto_raw = str(_item_first(item, "texto", "textoComunicacao", "conteudo", "descricao")).strip()
        data_disponibilizacao = _item_first(item, "data_disponibilizacao", "dataDisponibilizacao", "data", "data_publicacao")
        tribunal = str(_item_first(item, "tribunal", "siglaTribunal")).strip()
        orgao = str(_item_first(item, "orgao", "nomeOrgao", "orgaoJulgador")).strip()
        source_meta = _deadline_source_metadata(item)
        if not source_meta["official_source"] or not _valid_deadline_source_signature(item, org_id):
            blocked += 1
            skipped += 1
            continue

        # C4: Limpar HTML/códigos da descrição
        texto = re.sub(r'<[^>]+>', '', texto_raw)  # remove HTML tags
        texto = re.sub(r'&[a-zA-Z]+;', ' ', texto)  # remove &nbsp; etc
        texto = re.sub(r'\s+', ' ', texto).strip()  # normalizar espaços

        # C2+C3: Extrair prazo da publicação via regex
        dias_extraido = None
        prazo_patterns = [
            r'prazo\s+de\s+(\d+)\s*dias?',
            r'em\s+(\d+)\s*dias?\s*[uú]teis',
            r'no\s+prazo\s+de\s+(\d+)',
            r'(\d+)\s*dias?\s*para\s+(?:responder|manifestar|contestar|cumprir)',
            r'intimad[oa]\s+.*?(\d+)\s*dias?',
        ]
        for pat in prazo_patterns:
            m = re.search(pat, texto, re.IGNORECASE)
            if m:
                dias_extraido = int(m.group(1))
                if 1 <= dias_extraido <= 365:  # sanity check
                    break
                else:
                    dias_extraido = None
        if not dias_extraido:
            blocked += 1
            skipped += 1
            continue
        dias_prazo = dias_extraido

        # Parse date
        try:
            dt_intimacao = date.fromisoformat(data_disponibilizacao)
        except (ValueError, TypeError):
            blocked += 1
            skipped += 1
            continue

        # Check for duplicates by (processo + data_intimacao) OR (descricao + data_intimacao)
        dup_count = db.execute(
            text(
                "SELECT COUNT(*) FROM prazos_processuais p "
                "LEFT JOIN cases c ON p.case_id = c.id "
                "WHERE p.org_id = :org_id AND p.data_intimacao = :data "
                "AND (p.descricao = :texto OR COALESCE(p.processo_override, c.numero_processo, c.case_number) = :proc)"
            ),
            {"org_id": org_id, "texto": texto, "data": dt_intimacao, "proc": numero_processo},
        ).scalar() or 0

        if dup_count > 0:
            skipped += 1
            continue

        # Try to match case by numero_processo
        case_id = None
        if numero_processo:
            case_row = db.execute(
                text(
                    "SELECT id FROM cases "
                    "WHERE org_id = :org_id AND (case_number = :case_number OR numero_processo = :case_number) LIMIT 1"
                ),
                {"org_id": org_id, "case_number": numero_processo},
            ).fetchone()
            if case_row:
                case_id = case_row.id

        # Calculate vencimento using extracted or default days
        try:
            tribunal_codigo = normalizar_tribunal(tribunal or inferir_tribunal(numero_processo))
            dt_inicio = proximo_dia_util(dt_intimacao + timedelta(days=1), estado="MG", tribunal=tribunal_codigo)
            dt_vencimento = calcular_prazo(dt_intimacao, dias_prazo, estado="MG", tribunal=tribunal_codigo)
        except Exception:
            blocked += 1
            skipped += 1
            continue

        # Determine processo_override (only if no case match)
        processo_override = numero_processo if (numero_processo and not case_id) else None

        db.execute(
            text(
                "INSERT INTO prazos_processuais "
                "(org_id, case_id, tipo, data_intimacao, data_inicio, dias_prazo, "
                "data_vencimento, descricao, status, responsavel, uf, processo_override, "
                "source_provider, source_status, source_reference, source_url, source_payload_hash, "
                "source_fetched_at, source_version, official_source, calculation_engine_version, calculation_notes, created_at) "
                "VALUES (:org_id, :case_id, :tipo, :data_intimacao, :data_inicio, :dias_prazo, "
                ":data_vencimento, :descricao, :status, :responsavel, :uf, :processo_override, "
                ":source_provider, :source_status, :source_reference, :source_url, :source_payload_hash, "
                ":source_fetched_at, :source_version, :official_source, :calculation_engine_version, :calculation_notes, CURRENT_TIMESTAMP)"
            ),
            {
                "org_id": org_id,
                "case_id": case_id,
                "tipo": "Prazo Processual",
                "data_intimacao": dt_intimacao,
                "data_inicio": dt_inicio,
                "dias_prazo": dias_prazo,
                "data_vencimento": dt_vencimento,
                "descricao": texto,
                "status": "pendente",
                "responsavel": None,
                "uf": "MG",
                "processo_override": processo_override,
                "source_provider": source_meta["source_provider"],
                "source_status": source_meta["source_status"],
                "source_reference": source_meta["source_reference"],
                "source_url": item.get("url") or item.get("source_url") or None,
                "source_payload_hash": source_meta["source_payload_hash"],
                "source_fetched_at": datetime.utcnow(),
                "source_version": source_meta["source_version"],
                "official_source": source_meta["official_source"],
                "calculation_engine_version": PRAZOS_CALCULATION_ENGINE_VERSION,
                "calculation_notes": f"Dias do prazo extraidos da intimacao ({dias_prazo}); vencimento calculado por {PRAZOS_CALCULATION_ENGINE_VERSION}.",
            },
        )
        imported += 1

    db.commit()
    logger.info(
        "Importacao intimacoes: %d importados, %d ignorados, %d bloqueados (org_id=%d)",
        imported, skipped, blocked, org_id,
    )

    return JSONResponse({
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "blocked": blocked,
        "total": imported + skipped,
    })


@router.post("/{prazo_id}/concluir")
async def concluir_prazo(prazo_id: int, request: Request, db: Session = Depends(get_db)):
    """Marcar prazo como concluido, com tipo de peticao e data de conclusao."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada"}, status_code=403)

    org_id = _get_org_id(request)

    # Extract tipo_peticao and data_conclusao from request body
    tipo_peticao = None
    data_conclusao = None

    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        try:
            data = await request.json()
        except Exception:
            return JSONResponse({"error": "JSON invalido"}, status_code=400)
        tipo_peticao = data.get("tipo_peticao")
        data_conclusao_str = data.get("data_conclusao")
        if data_conclusao_str:
            try:
                data_conclusao = date.fromisoformat(data_conclusao_str)
            except ValueError:
                return JSONResponse({"error": "Data de conclusao invalida"}, status_code=400)
    else:
        try:
            form = await request.form()
        except Exception:
            return JSONResponse({"error": "Formulario invalido"}, status_code=400)
        tipo_peticao = form.get("tipo_peticao")
        data_conclusao_str = form.get("data_conclusao")
        if data_conclusao_str:
            try:
                data_conclusao = date.fromisoformat(data_conclusao_str)
            except ValueError:
                return JSONResponse({"error": "Data de conclusao invalida"}, status_code=400)

    if not data_conclusao:
        data_conclusao = date.today()

    prazo_row = db.execute(
        text("SELECT tipo FROM prazos_processuais WHERE id = :id AND org_id = :org_id"),
        {"id": prazo_id, "org_id": org_id},
    ).fetchone()
    if not prazo_row:
        return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)

    tipo_peticao = normalizar_tipo_produtividade(tipo_peticao, prazo_row.tipo)

    result = db.execute(
        text(
            "UPDATE prazos_processuais SET status = 'concluido', updated_at = CURRENT_TIMESTAMP, "
            "tipo_peticao = :tipo_peticao, data_conclusao = :data_conclusao "
            "WHERE id = :id AND org_id = :org_id"
        ),
        {
            "id": prazo_id,
            "org_id": org_id,
            "tipo_peticao": tipo_peticao,
            "data_conclusao": data_conclusao,
        },
    )
    db.commit()

    if result.rowcount == 0:
        return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)

    logger.info("Prazo %d marcado como concluido (tipo_peticao=%s)", prazo_id, tipo_peticao)

    if "application/json" in content_type:
        return JSONResponse({
            "success": True,
            "tipo_peticao": tipo_peticao,
            "data_conclusao": data_conclusao.isoformat(),
        })
    return RedirectResponse(url=f"{PREFIX}/controladoria", status_code=303)


@router.get("/api/datajud")
async def api_datajud_busca(
    request: Request,
    tipo: str = Query("numero", description="numero | oab | nome"),
    q: str = Query("", description="Termo de busca"),
    tribunal: str = Query("TJMG", description="Codigo do tribunal (ex.: TJMG, TJSP)"),
    db: Session = Depends(get_db),
):
    """Busca publica no DataJud (CNJ) embutida na controladoria — por numero CNJ, OAB do
    advogado ou nome da parte. API publica do CNJ, NAO precisa de credencial/secret."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"ok": False, "error": "Nao autenticado"}, status_code=401)

    termo = (q or "").strip()
    if not termo:
        return JSONResponse({"ok": False, "error": "Informe um numero, OAB ou nome para buscar."}, status_code=400)
    tribunal_code = (tribunal or "TJMG").strip().upper()

    # Incident-class fix (2026-07-01 outage pattern): `db` is not used again
    # in this handler after auth, so release it before the DataJud
    # round-trip instead of sitting idle-in-transaction on `users`.
    db.close()

    try:
        from services.datajud import datajud_client
        if tipo == "numero":
            raw = await datajud_client.consultar_processo(termo, tribunal=tribunal_code)
            records = [raw] if raw else []
        elif tipo == "oab":
            records = await datajud_client.buscar_por_advogado(termo, tribunal=tribunal_code)
        else:
            records = await datajud_client.buscar_por_parte(termo, tribunal=tribunal_code)
    except Exception as exc:
        logger.error("DataJud busca falhou (tipo=%s, tribunal=%s): %s", tipo, tribunal_code, exc)
        return JSONResponse(
            {"ok": False, "error": "Falha ao consultar o CNJ. Tente outro tribunal ou termo."},
            status_code=502,
        )

    def _fmt(rec):
        classe = rec.get("classe") or {}
        orgao = rec.get("orgaoJulgador") or {}
        assuntos = rec.get("assuntos") or []
        movs = rec.get("movimentos") or []
        ultimo = None
        if movs:
            try:
                ultimo = sorted(movs, key=lambda m: (m or {}).get("dataHora") or "", reverse=True)[0]
            except Exception:
                ultimo = movs[-1]
        return {
            "numero": rec.get("numeroProcesso") or "",
            "classe": (classe.get("nome") if isinstance(classe, dict) else "") or "",
            "tribunal": rec.get("tribunal") or tribunal_code,
            "orgao": (orgao.get("nome") if isinstance(orgao, dict) else "") or "",
            "assunto": (assuntos[0].get("nome") if assuntos and isinstance(assuntos[0], dict) else "") or "",
            "grau": rec.get("grau") or "",
            "data_ajuizamento": rec.get("dataAjuizamento") or "",
            "ultimo_movimento": (ultimo.get("nome") if isinstance(ultimo, dict) else "") or "",
            "ultimo_movimento_data": (ultimo.get("dataHora") if isinstance(ultimo, dict) else "") or "",
        }

    results = [_fmt(r) for r in records if r]
    return JSONResponse({"ok": True, "source": "DataJud (CNJ)", "count": len(results), "results": results})


@router.get("/api/produtividade")
async def api_produtividade(
    request: Request,
    mes: str = Query(None, description="Mes no formato YYYY-MM"),
    groupby: str = Query("urgencia", description="Eixo do donut: urgencia | status | responsavel | tipo"),
    db: Session = Depends(get_db),
):
    """Painel holístico da controladoria.

    Equipe CaseHub 03/06: o donut precisa refletir a CONTROLADORIA COMO UM TODO usando dados que
    EXISTEM. `groupby` controla o eixo do donut:
      - urgencia    (default): distribuição da carteira ATIVA por cor (fatal/vencido/amarelo/verde)
                    — mesma regra de cor das linhas (rota _get_prazos), sobre TODOS os prazos.
      - status      : distribuição de TODOS os prazos por status (pendente/concluido/perdido).
      - responsavel : carga da carteira ativa por responsável (NULL -> 'Sem responsável').
      - tipo        : petições concluídas no mês pela taxonomia gerencial da planilha.
    Sempre devolve `carteira` (a foto do todo) + `por_responsavel` (concluídos-no-mês) + meta.
    Org-scoped em toda query.
    """
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)

    org_id = _get_org_id(request)

    # Default to current month
    if not mes:
        mes = date.today().strftime("%Y-%m")

    try:
        ano, m = mes.split("-")
        ano = int(ano)
        m = int(m)
    except (ValueError, AttributeError):
        return JSONResponse({"error": "Formato de mes invalido. Use YYYY-MM"}, status_code=400)

    if m < 1 or m > 12:
        return JSONResponse({"error": "Formato de mes invalido. Use YYYY-MM"}, status_code=400)

    period_start = date(ano, m, 1)
    period_end = date(ano + 1, 1, 1) if m == 12 else date(ano, m + 1, 1)

    # C2 (UsuarioDemo 02/06): um prazo concluido conta na produtividade do mes quando a
    # DATA DE CONCLUSAO **ou** a DATA DE VENCIMENTO cai no mes. Antes so contava por
    # data_conclusao (com updated_at de fallback), entao prazos concluidos com
    # vencimento em maio/01-jun ficavam de fora da contagem mensal.
    rows = db.execute(
        text("""
            SELECT NULLIF(tipo_peticao,'') AS tipo_peticao,
                   NULLIF(tipo,'') AS tipo,
                   COUNT(*) as qtd
            FROM prazos_processuais
            WHERE org_id = :org_id
              AND status = 'concluido'
              AND (
                  (data_conclusao IS NOT NULL AND data_conclusao >= :period_start AND data_conclusao < :period_end)
                  OR
                  (data_vencimento IS NOT NULL AND data_vencimento >= :period_start AND data_vencimento < :period_end)
                  OR
                  (data_conclusao IS NULL AND data_vencimento IS NULL AND updated_at >= :period_start AND updated_at < :period_end)
              )
            GROUP BY NULLIF(tipo_peticao,''), NULLIF(tipo,'')
            ORDER BY qtd DESC
        """),
        {"org_id": org_id, "period_start": period_start, "period_end": period_end},
    ).fetchall()

    tipos, tipos_order, total = _aggregate_tipos_produtividade(rows)

    # Concluidos por responsavel no mes (MESMA janela de contagem dos tipos) — alimenta o
    # grafico de barras horizontais ao lado do donut (Equipe CaseHub 03/06: preencher o espaco
    # horizontal vazio com mais informacao = accountability por pessoa). Org-scoped.
    resp_rows = db.execute(
        text("""
            SELECT COALESCE(NULLIF(responsavel, ''), 'Sem responsável') AS nome, COUNT(*) AS qtd
            FROM prazos_processuais
            WHERE org_id = :org_id
              AND status = 'concluido'
              AND (
                  (data_conclusao IS NOT NULL AND data_conclusao >= :period_start AND data_conclusao < :period_end)
                  OR
                  (data_vencimento IS NOT NULL AND data_vencimento >= :period_start AND data_vencimento < :period_end)
                  OR
                  (data_conclusao IS NULL AND data_vencimento IS NULL AND updated_at >= :period_start AND updated_at < :period_end)
              )
            GROUP BY COALESCE(NULLIF(responsavel, ''), 'Sem responsável')
            ORDER BY qtd DESC
            LIMIT 8
        """),
        {"org_id": org_id, "period_start": period_start, "period_end": period_end},
    ).fetchall()
    por_responsavel = [{"nome": r.nome, "qtd": int(r.qtd or 0)} for r in resp_rows]

    # --- A FOTO DO TODO (carteira) — reusa _get_stats (org-scoped, all-time) + o total mensal.
    # Equipe CaseHub 03/06: KPIs holísticos no topo do painel. eficiencia_geral = concluídos/total da carteira.
    carteira_stats = _get_stats(db, org_id)
    c_total = int(carteira_stats.get("total") or 0)
    carteira = {
        "total": c_total,
        "pendentes": max(c_total - int(carteira_stats.get("concluidos") or 0), 0),
        "concluidos_total": int(carteira_stats.get("concluidos") or 0),
        "concluidos_mes": total,
        "vencidos": int(carteira_stats.get("vencidos") or 0),
        "fatais_hoje": int(carteira_stats.get("fatais_hoje") or 0),
        "proximos": int(carteira_stats.get("proximos") or 0),
        "eficiencia_geral": round((int(carteira_stats.get("concluidos") or 0) / c_total) * 100) if c_total else 0,
    }

    # --- DONUT: eixo escolhido. urgencia/status/responsavel = carteira REAL (dados preenchidos);
    # tipo = concluídos-no-mês pela taxonomia do escritório. dist_order preserva a ordem.
    gb = (groupby or "urgencia").strip().lower()
    if gb not in ("urgencia", "status", "responsavel", "tipo"):
        gb = "urgencia"

    dist: dict = {}
    dist_order: list = []
    dist_label = "Por urgência"
    if gb == "tipo":
        dist = dict(tipos)
        dist_order = list(tipos_order)
        dist_label = "Petições produzidas no mês"
    elif gb == "responsavel":
        # Carga da carteira ATIVA (não concluído/perdido) por responsável — mostra quem segura o quê.
        r2 = db.execute(
            text("""
                SELECT COALESCE(NULLIF(responsavel, ''), 'Sem responsável') AS k, COUNT(*) AS qtd
                FROM prazos_processuais
                WHERE org_id = :org_id
                  AND COALESCE(status, 'pendente') NOT IN ('concluido', 'perdido')
                GROUP BY COALESCE(NULLIF(responsavel, ''), 'Sem responsável')
                ORDER BY qtd DESC LIMIT 10
            """),
            {"org_id": org_id},
        ).fetchall()
        for row in r2:
            dist[row.k] = int(row.qtd or 0)
            dist_order.append(row.k)
        dist_label = "Carteira ativa por responsável"
    elif gb == "status":
        s2 = db.execute(
            text("""
                SELECT COALESCE(NULLIF(status, ''), 'pendente') AS k, COUNT(*) AS qtd
                FROM prazos_processuais
                WHERE org_id = :org_id
                GROUP BY COALESCE(NULLIF(status, ''), 'pendente')
            """),
            {"org_id": org_id},
        ).fetchall()
        _STATUS_LABEL = {"pendente": "Pendente", "concluido": "Concluído", "perdido": "Perdido", "em_andamento": "Em andamento"}
        raw = {row.k: int(row.qtd or 0) for row in s2}
        for k in sorted(raw, key=lambda x: -raw[x]):
            label = _STATUS_LABEL.get(k, k.replace("_", " ").capitalize())
            dist[label] = raw[k]
            dist_order.append(label)
        dist_label = "Por status"
    else:  # urgencia — MESMA regra de cor das linhas (_get_prazos): fatal/vencido/amarelo/verde.
        u = db.execute(
            text("""
                SELECT
                  SUM(CASE WHEN status = 'perdido'
                            OR (status <> 'concluido' AND data_vencimento IS NOT NULL AND data_vencimento < :today)
                           THEN 1 ELSE 0 END) AS vencido,
                  SUM(CASE WHEN status <> 'concluido' AND status <> 'perdido'
                            AND data_vencimento = :today THEN 1 ELSE 0 END) AS fatal,
                  SUM(CASE WHEN status <> 'concluido' AND status <> 'perdido'
                            AND data_vencimento > :today
                            AND data_vencimento <= :today_plus_7 THEN 1 ELSE 0 END) AS amarelo,
                  SUM(CASE WHEN status <> 'concluido' AND status <> 'perdido'
                            AND (data_vencimento IS NULL OR data_vencimento > :today_plus_7)
                           THEN 1 ELSE 0 END) AS verde,
                  SUM(CASE WHEN status = 'concluido' THEN 1 ELSE 0 END) AS concluido
                FROM prazos_processuais
                WHERE org_id = :org_id
            """),
            {"org_id": org_id, "today": date.today(), "today_plus_7": date.today() + timedelta(days=7)},
        ).fetchone()
        # Ordem semântica de gravidade (Lei de Miller: poucas faixas, fáceis de ler).
        for label, key in (("Fatal hoje", "fatal"), ("Vencidos", "vencido"),
                           ("Próx. 7 dias", "amarelo"), ("Em dia", "verde"), ("Concluídos", "concluido")):
            qtd = int(getattr(u, key) or 0) if u else 0
            if qtd:
                dist[label] = qtd
                dist_order.append(label)
        dist_label = "Carteira por urgência"

    META = _get_meta(db, org_id)

    return JSONResponse({
        "total": total,
        "mes": mes,
        "groupby": gb,
        "dist": dist,
        "dist_order": dist_order,
        "dist_label": dist_label,
        "carteira": carteira,
        "tipos": tipos,
        "tipos_produtividade": list(TIPOS_PRODUTIVIDADE),
        "por_responsavel": por_responsavel,
        "meta_batida": total >= META,
        "meta": META,
        "meta_sugerida": META_SUGERIDA,
        "pode_editar_meta": has_permission(user.user_type or "", "settings.edit"),
    })


@router.post("/meta")
async def atualizar_meta(request: Request, db: Session = Depends(get_db)):
    """Atualiza a meta mensal de produtividade da org (UsuarioDemo 02/06: gestor edita a meta).
    Ruling 2026-06-03: role-gated (settings.edit = gestor/super_admin), org-scoped SEM
    fallback->1, input validado, audit log. Storage em org_settings (existente)."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "settings.edit"):
        return JSONResponse(
            {"error": "Permissao negada: apenas gestor/administrador pode editar a meta"},
            status_code=403,
        )
    org_id = _get_org_id_strict(request)
    if not org_id:
        return JSONResponse({"error": "Organizacao nao identificada"}, status_code=400)
    try:
        data = await request.json()
        meta = int(data.get("meta"))
    except (TypeError, ValueError):
        return JSONResponse({"error": "Meta invalida (use um numero inteiro)"}, status_code=400)
    if meta < 1 or meta > 100000:
        return JSONResponse({"error": "Meta fora da faixa permitida (1 a 100000)"}, status_code=400)
    try:
        old = _get_meta(db, org_id)
        settings = _org_settings(db, org_id)
        settings[META_KEY] = meta
        db.execute(
            text("UPDATE organizations SET settings = :s WHERE id = :oid"),
            {"s": json.dumps(settings), "oid": org_id},
        )
        db.commit()
        try:
            from services.audit import log_action
            log_action(
                db, action="controladoria.meta_update", entity_type="org_settings",
                entity_id=org_id, user_id=user.id,
                description=f"Meta mensal {old} -> {meta} (org {org_id})",
                details={"de": old, "para": meta, "org_id": org_id}, request=request,
            )
        except Exception as _audit_err:
            logger.warning("meta_update sem audit log: %s", _audit_err)
        return JSONResponse({"success": True, "meta": meta, "meta_sugerida": META_SUGERIDA})
    except Exception as e:
        db.rollback()
        logger.error("Erro ao salvar meta: %s", e)
        return JSONResponse({"error": "Falha ao salvar a meta"}, status_code=500)


@router.post("/drag-toggle")
async def atualizar_drag_toggle(request: Request, db: Session = Depends(get_db)):
    """Liga/desliga o drag-and-drop de reordenação de prazos (org-scoped).
    Equipe CaseHub 2026-06-15: drag é nativo e ativo por padrão, com opção de desativar
    nas configurações. Role-gated (settings.edit), audit log, sem migração —
    storage em organizations.settings (mesmo padrão de /meta)."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "settings.edit"):
        return JSONResponse(
            {"error": "Permissao negada: apenas gestor/administrador pode alterar esta configuracao"},
            status_code=403,
        )
    org_id = _get_org_id_strict(request)
    if not org_id:
        return JSONResponse({"error": "Organizacao nao identificada"}, status_code=400)
    try:
        data = await request.json()
        raw = data.get("enabled")
        if isinstance(raw, str):
            enabled = raw.strip().lower() in ("true", "1", "yes", "on")
        else:
            enabled = bool(raw)
    except (TypeError, ValueError):
        return JSONResponse({"error": "Valor invalido (use enabled true/false)"}, status_code=400)
    try:
        old = _get_drag_enabled(db, org_id)
        settings = _org_settings(db, org_id)
        settings[DRAG_KEY] = enabled
        db.execute(
            text("UPDATE organizations SET settings = :s WHERE id = :oid"),
            {"s": json.dumps(settings), "oid": org_id},
        )
        db.commit()
        try:
            from services.audit import log_action
            log_action(
                db, action="controladoria.drag_toggle", entity_type="org_settings",
                entity_id=org_id, user_id=user.id,
                description=f"Drag-and-drop {old} -> {enabled} (org {org_id})",
                details={"de": old, "para": enabled, "org_id": org_id}, request=request,
            )
        except Exception as _audit_err:
            logger.warning("drag_toggle sem audit log: %s", _audit_err)
        return JSONResponse({"success": True, "enabled": enabled})
    except Exception as e:
        db.rollback()
        logger.error("Erro ao salvar drag toggle: %s", e)
        return JSONResponse({"error": "Falha ao salvar a configuracao"}, status_code=500)


@router.post("/reordenar")
async def reordenar_prazos(request: Request, db: Session = Depends(get_db)):
    """Persist manual row ordering from the Controladoria table."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada"}, status_code=403)

    org_id = _get_org_id(request)
    try:
        data = await request.json()
        ids = [int(value) for value in data.get("ids", []) if value]
    except (TypeError, ValueError):
        return JSONResponse({"error": "Lista de prazos invalida"}, status_code=400)

    if not ids:
        return JSONResponse({"error": "Nenhum prazo informado"}, status_code=400)

    try:
        for position, prazo_id in enumerate(ids, start=1):
            db.execute(
                text("""
                    UPDATE prazos_processuais
                    SET ordem = :ordem, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id AND org_id = :org_id
                """),
                {"ordem": position, "id": prazo_id, "org_id": org_id},
            )
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("Erro ao reordenar prazos: %s", exc)
        return JSONResponse({"error": "Erro ao salvar ordem dos prazos"}, status_code=500)

    return JSONResponse({"success": True, "count": len(ids)})


@router.post("/{prazo_id}/update")
async def update_prazo(prazo_id: int, request: Request, db: Session = Depends(get_db)):
    """Inline update a single field on a prazo (Asana-style editing)."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada"}, status_code=403)

    org_id = _get_org_id(request)

    try:
        data = await request.json()
    except Exception:
        return JSONResponse({"error": "JSON invalido"}, status_code=400)

    field = data.get("field")
    value = data.get("value")

    # Parte contrária: se há case_id grava em cases.polo_passivo; se avulso
    # (sem case_id) grava em prazos_processuais.parte_contraria_override. 11/06.
    if field == "parte_contraria":
        row = db.execute(
            text("SELECT case_id FROM prazos_processuais WHERE id = :id AND org_id = :org_id"),
            {"id": prazo_id, "org_id": org_id},
        ).fetchone()
        if not row:
            return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)
        novo = (value or "").strip() or None
        if row.case_id:
            db.execute(
                text("UPDATE cases SET polo_passivo = :v WHERE id = :cid AND org_id = :org_id"),
                {"v": novo, "cid": row.case_id, "org_id": org_id},
            )
        else:
            db.execute(
                text(
                    "UPDATE prazos_processuais SET parte_contraria_override = :v, "
                    "updated_at = CURRENT_TIMESTAMP WHERE id = :id AND org_id = :org_id"
                ),
                {"v": novo, "id": prazo_id, "org_id": org_id},
            )
        db.commit()
        return JSONResponse({"success": True, "field": field, "value": novo or "—"})

    # Whitelist of editable fields
    allowed_fields = {
        "descricao": "descricao",
        "status": "status",
        "tipo": "tipo",
        "responsavel": "responsavel",
        "responsavel_user_id": "responsavel_user_id",
        "processo": "processo_override",
        "cliente": "cliente_override",
        "dias_prazo": "dias_prazo",
        "data_inicio": "data_inicio",
        "data_intimacao": "data_intimacao",
        "tipo_peticao": "tipo_peticao",
        "hora_vencimento": "hora_vencimento",
    }

    if field not in allowed_fields:
        return JSONResponse(
            {"error": f"Campo '{field}' nao editavel. Campos permitidos: {', '.join(allowed_fields.keys())}"},
            status_code=400,
        )

    # Validate status values
    if field == "status" and value not in ("pendente", "em_andamento", "concluido", "perdido", "aguarda_correcao"):
        return JSONResponse({"error": f"Status invalido: {value}"}, status_code=400)
    if field == "status" and value == "concluido":
        return JSONResponse(
            {"error": "Use /controladoria/{id}/concluir para gravar tipo_peticao e data_conclusao"},
            status_code=400,
        )

    # UsuarioDemo 03/06: horário do vencimento ("HH:MM"); vazio limpa a hora.
    if field == "hora_vencimento":
        value = (value or "").strip()
        if value and not _HORA_RE.match(value):
            return JSONResponse({"error": "Horario invalido (use HH:MM)"}, status_code=400)
        value = value or None

    if field == "tipo_peticao":
        value = normalizar_tipo_produtividade(value)

    if field == "responsavel_user_id":
        prazo_row = db.execute(
            text("""
                SELECT tipo, data_vencimento, hora_vencimento, responsavel_user_id
                FROM prazos_processuais
                WHERE id = :id AND org_id = :org_id
            """),
            {"id": prazo_id, "org_id": org_id},
        ).fetchone()
        if not prazo_row:
            return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)
        previous_user_id = prazo_row.responsavel_user_id
        try:
            user_id = int(value) if value not in (None, "", "—") else None
        except (TypeError, ValueError):
            return JSONResponse({"error": "Responsavel invalido"}, status_code=400)

        if user_id:
            user_row = db.execute(
                text("""
                    SELECT id, name, color, photo_url
                    FROM users
                    WHERE id = :user_id AND org_id = :org_id AND enabled = TRUE
                """),
                {"user_id": user_id, "org_id": org_id},
            ).fetchone()
            if not user_row:
                return JSONResponse({"error": "Usuario nao encontrado"}, status_code=404)
            responsavel_nome = user_row.name
        else:
            user_row = None
            responsavel_nome = None

        result = db.execute(
            text("""
                UPDATE prazos_processuais
                SET responsavel_user_id = :responsavel_user_id,
                    responsavel = :responsavel,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id AND org_id = :org_id
            """),
            {
                "responsavel_user_id": user_id,
                "responsavel": responsavel_nome,
                "id": prazo_id,
                "org_id": org_id,
            },
        )
        db.commit()
        if result.rowcount == 0:
            return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)
        if user_id and user_id != previous_user_id:
            _notify_prazo_assignee_dm(
                db,
                org_id,
                user.id,
                user_id,
                prazo_row.tipo,
                prazo_row.data_vencimento,
                prazo_row.hora_vencimento,
            )
        payload = {
            "success": True,
            "field": field,
            "value": user_id,
            "responsavel": responsavel_nome or "—",
            "responsavel_user": None,
        }
        if user_row:
            payload["responsavel_user"] = {
                "id": user_row.id,
                "name": user_row.name,
                "initials": _initials(user_row.name),
                "photo_url": user_row.photo_url or "",
                "color": user_row.color or "#1C2447",
            }
        return JSONResponse(payload)

    col = allowed_fields[field]

    # P8 fix + reunião 10/06: ao editar dias_prazo OU data_intimacao (Dia de Início),
    # recalcular data_vencimento e data_inicio — respeitando dias corridos (administrativo).
    extra_updates = ""
    extra_params = {}
    if field in ("dias_prazo", "data_intimacao"):
        prazo_row = db.execute(
            text("""
                SELECT p.data_inicio, p.data_intimacao, p.dias_prazo, p.uf, p.dobro,
                       COALESCE(p.dias_corridos, FALSE) AS dias_corridos, p.processo_override,
                       c.tribunal, c.numero_processo, c.case_number
                FROM prazos_processuais p
                LEFT JOIN cases c ON p.case_id = c.id
                WHERE p.id = :id AND p.org_id = :org_id
            """),
            {"id": prazo_id, "org_id": org_id},
        ).fetchone()
        if prazo_row:
            try:
                if field == "dias_prazo":
                    new_dias = int(value)
                    base_intimacao = prazo_row.data_intimacao
                    value = new_dias
                else:  # data_intimacao (Dia de Início)
                    base_intimacao = date.fromisoformat(value) if value else prazo_row.data_intimacao
                    new_dias = prazo_row.dias_prazo or 15
                if base_intimacao and new_dias and new_dias > 0:
                    uf = prazo_row.uf or "MG"
                    dobro = prazo_row.dobro or False
                    if bool(prazo_row.dias_corridos):
                        new_venc = calcular_prazo_corrido(base_intimacao, new_dias)
                        new_inicio = base_intimacao + timedelta(days=1)
                    else:
                        processo_ref = prazo_row.processo_override or prazo_row.numero_processo or prazo_row.case_number
                        tribunal_inferido = inferir_tribunal(processo_ref)
                        tribunal_codigo = normalizar_tribunal(prazo_row.tribunal or tribunal_inferido)
                        if tribunal_codigo == "CNJ" and not prazo_row.tribunal and tribunal_inferido == "CNJ":
                            tribunal_codigo = "Outro"
                        new_venc = calcular_prazo(base_intimacao, new_dias, uf, dobro, tribunal=tribunal_codigo)
                        new_inicio = proximo_dia_util(base_intimacao + timedelta(days=1), uf, tribunal=tribunal_codigo)
                    extra_updates = ", data_vencimento = :new_venc, data_inicio = :new_inicio"
                    extra_params["new_venc"] = new_venc
                    extra_params["new_inicio"] = new_inicio
            except (TypeError, ValueError):
                pass
    elif field == "responsavel":
        extra_updates = ", responsavel_user_id = NULL"

    try:
        result = db.execute(
            text(
                f"UPDATE prazos_processuais SET {col} = :value{extra_updates}, updated_at = CURRENT_TIMESTAMP "
                "WHERE id = :id AND org_id = :org_id"
            ),
            {"value": value, "id": prazo_id, "org_id": org_id, **extra_params},
        )
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error("Erro ao atualizar prazo %d: %s", prazo_id, e)
        return JSONResponse({"error": f"Erro ao salvar: {str(e)}"}, status_code=500)

    if result.rowcount == 0:
        return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)

    resp = {"success": True, "field": field, "value": value}
    if extra_params.get("new_venc"):
        resp["new_vencimento"] = str(extra_params["new_venc"])
    if extra_params.get("new_inicio"):
        resp["new_inicio"] = str(extra_params["new_inicio"])

    logger.info("Prazo %d atualizado: %s = %s %s", prazo_id, field, value,
                f"(vencimento recalculado: {extra_params.get('new_venc')})" if extra_params.get("new_venc") else "")
    return JSONResponse(resp)


@router.post("/{prazo_id}/excluir")
async def excluir_prazo(prazo_id: int, request: Request, db: Session = Depends(get_db)):
    """Excluir prazo."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    # Ruling 2026-06-03 (CWE-862): exclusao exige cargo com delete (gestor/advogado).
    if not has_permission(user.user_type or "", "cases.delete"):
        return JSONResponse({"error": "Permissao negada para excluir prazo"}, status_code=403)

    org_id = _get_org_id(request)

    result = db.execute(
        text(
            "DELETE FROM prazos_processuais WHERE id = :id AND org_id = :org_id"
        ),
        {"id": prazo_id, "org_id": org_id},
    )
    db.commit()

    if result.rowcount == 0:
        return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)

    logger.info("Prazo %d excluido", prazo_id)

    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        return JSONResponse({"success": True})
    return RedirectResponse(url=f"{PREFIX}/controladoria", status_code=303)


@router.post("/bulk-concluir")
async def bulk_concluir(request: Request, db: Session = Depends(get_db)):
    """Concluir multiplos prazos de uma vez."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada"}, status_code=403)

    org_id = _get_org_id(request)
    try:
        data = await request.json()
    except Exception:
        return JSONResponse({"error": "JSON invalido"}, status_code=400)

    ids = [int(i) for i in data.get("ids", []) if str(i).isdigit()]
    if not ids:
        return JSONResponse({"error": "Nenhum prazo selecionado"}, status_code=400)

    tipo_raw = data.get("tipo_peticao")
    if not (tipo_raw or "").strip():
        return JSONResponse({"error": "Tipo de peticao obrigatorio"}, status_code=400)
    tipo_peticao = normalizar_tipo_produtividade(tipo_raw)
    try:
        data_conclusao = date.fromisoformat(data.get("data_conclusao")) if data.get("data_conclusao") else date.today()
    except ValueError:
        return JSONResponse({"error": "Data de conclusao invalida"}, status_code=400)

    result = db.execute(
        text(
            "UPDATE prazos_processuais SET status = 'concluido', "
            "tipo_peticao = :tipo_peticao, data_conclusao = :data_conclusao, "
            "updated_at = CURRENT_TIMESTAMP "
            "WHERE id IN :ids AND org_id = :org_id AND status != 'concluido'"
        ).bindparams(bindparam("ids", expanding=True)),
        {"ids": ids, "org_id": org_id, "tipo_peticao": tipo_peticao, "data_conclusao": data_conclusao},
    )
    db.commit()
    return JSONResponse({"success": True, "updated": result.rowcount, "tipo_peticao": tipo_peticao})


@router.post("/bulk-excluir")
async def bulk_excluir(request: Request, db: Session = Depends(get_db)):
    """Excluir multiplos prazos de uma vez."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    # Ruling 2026-06-03 (CWE-862): exclusao em massa exige cargo com delete.
    if not has_permission(user.user_type or "", "cases.delete"):
        return JSONResponse({"error": "Permissao negada para excluir prazos"}, status_code=403)

    org_id = _get_org_id(request)
    try:
        data = await request.json()
    except Exception:
        return JSONResponse({"error": "JSON invalido"}, status_code=400)

    ids = [int(i) for i in data.get("ids", []) if str(i).isdigit()]
    if not ids:
        return JSONResponse({"error": "Nenhum prazo selecionado"}, status_code=400)

    result = db.execute(
        text("DELETE FROM prazos_processuais WHERE id IN :ids AND org_id = :org_id")
        .bindparams(bindparam("ids", expanding=True)),
        {"ids": ids, "org_id": org_id},
    )
    db.commit()
    return JSONResponse({"success": True, "deleted": result.rowcount})


@router.post("/{prazo_id}/duplicar")
async def duplicar_prazo(prazo_id: int, request: Request, db: Session = Depends(get_db)):
    """Duplicar um prazo processual."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)
    if not has_permission(user.user_type or "", "cases.edit"):
        return JSONResponse({"error": "Permissao negada"}, status_code=403)

    org_id = _get_org_id(request)

    row = db.execute(
        text(
            "SELECT case_id, tipo, data_intimacao, data_inicio, data_vencimento, "
            "dias_prazo, responsavel, responsavel_user_id, descricao, uf, dobro "
            "FROM prazos_processuais WHERE id = :id AND org_id = :org_id"
        ),
        {"id": prazo_id, "org_id": org_id},
    ).fetchone()

    if not row:
        return JSONResponse({"error": "Prazo nao encontrado"}, status_code=404)

    result = db.execute(
        text(
            "INSERT INTO prazos_processuais "
            "(case_id, org_id, tipo, data_intimacao, data_inicio, data_vencimento, "
            "dias_prazo, responsavel, responsavel_user_id, status, descricao, uf, dobro) "
            "VALUES (:case_id, :org_id, :tipo, :di, :ds, :dv, :dp, :resp, "
            ":resp_user_id, 'pendente', :desc, :uf, :dobro) RETURNING id"
        ),
        {
            "case_id": row.case_id, "org_id": org_id, "tipo": row.tipo,
            "di": row.data_intimacao, "ds": row.data_inicio,
            "dv": row.data_vencimento, "dp": row.dias_prazo,
            "resp": row.responsavel, "resp_user_id": row.responsavel_user_id, "desc": row.descricao,
            "uf": row.uf, "dobro": row.dobro,
        },
    )
    new_id = result.scalar()
    db.commit()
    return JSONResponse({"success": True, "new_id": new_id})


@router.get("/export/excel")
async def export_excel(request: Request, db: Session = Depends(get_db)):
    """Exportar prazos para planilha Excel (.xlsx)."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    org_id = _get_org_id(request)
    prazos = _get_prazos(db, org_id)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JSONResponse(
            {"error": "openpyxl nao instalado. Execute: pip install openpyxl"},
            status_code=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Prazos Processuais"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Status fills
    status_fills = {
        "pendente": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "em_andamento": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "concluido": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "perdido": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }

    # Cabecalho
    headers = [
        "Processo", "Cliente", "Parte Contraria", "Tipo", "Intimacao",
        "Inicio", "Vencimento", "Dias", "Responsavel", "Status", "UF", "Descricao",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dados
    for row_idx, p in enumerate(prazos, 2):
        values = [
            p["processo"],
            p["cliente"],
            p["parte_contraria"],
            p["tipo"],
            p["data_intimacao"].strftime("%d/%m/%Y") if isinstance(p["data_intimacao"], date) else str(p["data_intimacao"]),
            p["data_inicio"].strftime("%d/%m/%Y") if isinstance(p["data_inicio"], date) else str(p["data_inicio"]),
            p["data_vencimento"].strftime("%d/%m/%Y") if isinstance(p["data_vencimento"], date) else str(p["data_vencimento"]),
            p["dias_prazo"],
            p["responsavel"],
            p["status"],
            p["uf"],
            p["descricao"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col == 10:  # Status column
                fill = status_fills.get(p["status"])
                if fill:
                    cell.fill = fill

    # Ajustar larguras
    col_widths = [20, 25, 25, 20, 12, 12, 12, 6, 20, 14, 4, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width

    # Salvar em buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    hoje_str = date.today().strftime("%Y-%m-%d")
    filename = f"prazos_processuais_{hoje_str}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/stats")
async def api_stats(request: Request, db: Session = Depends(get_db)):
    """JSON stats para os cards do dashboard."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)

    org_id = _get_org_id(request)
    stats = _get_stats(db, org_id)
    return JSONResponse(stats)


# ---------------------------------------------------------------------------
# Indices (Analytics Charts)
# ---------------------------------------------------------------------------

# Month names in Portuguese
_MONTH_NAMES_PT = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}

# Statuses considered as "won" for productivity indexes.
_WON_STATUSES = {
    "won", "approved", "granted", "deferido", "procedente", "ganho", "vitoria",
}
# Statuses considered as resolved (won or lost)
_RESOLVED_STATUSES = _WON_STATUSES | {
    "lost", "denied", "improcedente", "indeferido", "perdido", "derrota",
    "closed", "completed", "archived", "encerrado", "finalizado",
}


@router.get("/indices", response_class=HTMLResponse)
async def indices_page(request: Request, db: Session = Depends(get_db)):
    """Render the Indices analytics page."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    org_ctx = inject_org_context(request)
    return templates.TemplateResponse(
        "controladoria/indices.html",
        {
            "request": request,
            "user": user,
            "PREFIX": PREFIX,
            **org_ctx,
        },
    )


@router.get("/api/indices")
async def get_indices(
    request: Request,
    setor: str = None,
    tipo: str = None,
    periodo: str = None,
    db: Session = Depends(get_db),
):
    """Return analytics data for the 6 charts on the Indices page."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"error": "Nao autenticado"}, status_code=401)

    org_id = _get_org_id(request)

    # Build date filter
    date_cutoff = None
    if periodo == "3m":
        date_cutoff = datetime.utcnow() - timedelta(days=90)
    elif periodo == "6m":
        date_cutoff = datetime.utcnow() - timedelta(days=180)
    elif periodo == "1y":
        date_cutoff = datetime.utcnow() - timedelta(days=365)

    # Build dynamic SQL with filters
    conditions = ["org_id = :org_id"]
    params: dict = {"org_id": org_id}

    if setor:
        conditions.append("area_of_practice = :setor")
        params["setor"] = setor
    if tipo:
        conditions.append("tipo_acao = :tipo")
        params["tipo"] = tipo
    if date_cutoff:
        conditions.append("created_at >= :date_cutoff")
        params["date_cutoff"] = date_cutoff

    where_clause = " AND ".join(conditions)

    query = text(f"""
        SELECT id, status, area_of_practice, tipo_acao,
               created_at, updated_at,
               EXTRACT(MONTH FROM created_at)::int AS month,
               EXTRACT(YEAR FROM created_at)::int AS year
        FROM cases
        WHERE {where_clause}
        ORDER BY created_at
    """)

    try:
        rows = db.execute(query, params).fetchall()
    except Exception as e:
        logger.error("Erro ao consultar indices: %s", e)
        rows = []

    # ── Helpers ──
    from collections import defaultdict

    # ── 1. Tempo Medio por Setor (months between created_at and updated_at) ──
    setor_times: dict = defaultdict(list)
    for r in rows:
        if r.area_of_practice and r.updated_at and r.created_at:
            if (r.status or "").lower() in _RESOLVED_STATUSES:
                delta = (r.updated_at - r.created_at).days / 30.0
                if delta >= 0:
                    setor_times[r.area_of_practice].append(delta)

    tempo_medio_setor = {
        "labels": list(setor_times.keys()),
        "data": [round(sum(v) / len(v), 1) if v else 0 for v in setor_times.values()],
    }

    # ── 2. Tempo Medio por Tipo ──
    tipo_times: dict = defaultdict(list)
    for r in rows:
        if r.tipo_acao and r.updated_at and r.created_at:
            if (r.status or "").lower() in _RESOLVED_STATUSES:
                delta = (r.updated_at - r.created_at).days / 30.0
                if delta >= 0:
                    tipo_times[r.tipo_acao].append(delta)

    tempo_medio_tipo = {
        "labels": list(tipo_times.keys()),
        "data": [round(sum(v) / len(v), 1) if v else 0 for v in tipo_times.values()],
    }

    # ── 3. Indice de Vitoria por Setor (%) ──
    setor_won: dict = defaultdict(int)
    setor_resolved: dict = defaultdict(int)
    for r in rows:
        if r.area_of_practice:
            sl = (r.status or "").lower()
            if sl in _RESOLVED_STATUSES:
                setor_resolved[r.area_of_practice] += 1
                if sl in _WON_STATUSES:
                    setor_won[r.area_of_practice] += 1

    vs_labels = list(setor_resolved.keys())
    vitoria_setor = {
        "labels": vs_labels,
        "data": [
            round(setor_won[s] / setor_resolved[s] * 100) if setor_resolved[s] else 0
            for s in vs_labels
        ],
    }

    # ── 4. Indice de Vitoria por Tipo (%) ──
    tipo_won: dict = defaultdict(int)
    tipo_resolved: dict = defaultdict(int)
    for r in rows:
        if r.tipo_acao:
            sl = (r.status or "").lower()
            if sl in _RESOLVED_STATUSES:
                tipo_resolved[r.tipo_acao] += 1
                if sl in _WON_STATUSES:
                    tipo_won[r.tipo_acao] += 1

    vt_labels = list(tipo_resolved.keys())
    vitoria_tipo = {
        "labels": vt_labels,
        "data": [
            round(tipo_won[t] / tipo_resolved[t] * 100) if tipo_resolved[t] else 0
            for t in vt_labels
        ],
    }

    # ── 5. Processos por Mes ──
    month_counts: dict = defaultdict(int)
    for r in rows:
        if r.year and r.month:
            month_counts[(int(r.year), int(r.month))] += 1

    sorted_months = sorted(month_counts.keys())
    processos_mes = {
        "labels": [
            f"{_MONTH_NAMES_PT.get(m, m)}/{str(y)[2:]}" for y, m in sorted_months
        ],
        "data": [month_counts[k] for k in sorted_months],
    }

    # ── 6. Distribuicao por Setor ao Longo do Tempo (%) ──
    all_setores: set = set()
    month_setor_counts: dict = defaultdict(lambda: defaultdict(int))
    for r in rows:
        if r.area_of_practice and r.year and r.month:
            key = (int(r.year), int(r.month))
            month_setor_counts[key][r.area_of_practice] += 1
            all_setores.add(r.area_of_practice)

    all_setores_sorted = sorted(all_setores)
    dist_months = sorted(month_setor_counts.keys())
    dist_labels = [
        f"{_MONTH_NAMES_PT.get(m, m)}/{str(y)[2:]}" for y, m in dist_months
    ]

    dist_datasets = []
    for sname in all_setores_sorted:
        pct_data = []
        for mk in dist_months:
            total_in_month = sum(month_setor_counts[mk].values())
            if total_in_month > 0:
                pct_data.append(
                    round(month_setor_counts[mk][sname] / total_in_month * 100, 1)
                )
            else:
                pct_data.append(0)
        dist_datasets.append({"label": sname, "data": pct_data})

    distribuicao_setor = {
        "labels": dist_labels,
        "datasets": dist_datasets,
    }

    return JSONResponse({
        "tempo_medio_setor": tempo_medio_setor,
        "tempo_medio_tipo": tempo_medio_tipo,
        "vitoria_setor": vitoria_setor,
        "vitoria_tipo": vitoria_tipo,
        "processos_mes": processos_mes,
        "distribuicao_setor": distribuicao_setor,
    })
