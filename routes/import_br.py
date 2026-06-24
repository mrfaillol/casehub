"""
CaseHub Lite - Generic CSV/Excel Import System
Import data from CSV or Excel files with column mapping wizard.

Endpoints:
    GET  /import-br           — Import wizard landing page
    POST /import-br/upload    — Upload file, return preview
    POST /import-br/map       — Accept column mapping, execute import
    GET  /import-br/templates — Download template CSV files
"""
import csv
import io
import json
import logging
from datetime import datetime, date

from fastapi import APIRouter, Depends, Request, Form, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from core.template_config import templates, PREFIX
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional

logger = logging.getLogger(__name__)

from models import get_db, Client, Case
from auth import get_current_user
from models.tenant import tenant_query

router = APIRouter(prefix="/import-br", tags=["import-br"])


# ---------------------------------------------------------------------------
# Entity definitions: field mappings for each importable entity type
# ---------------------------------------------------------------------------
ENTITY_FIELDS = {
    "clientes": {
        "label": "Clientes",
        "icon": "fas fa-users",
        "fields": {
            "nome": {"label": "Nome Completo", "required": True, "description": "Nome completo do cliente"},
            "email": {"label": "E-mail", "required": False, "description": "Endereco de e-mail"},
            "telefone": {"label": "Telefone", "required": False, "description": "Telefone com DDD"},
            "cpf": {"label": "CPF", "required": False, "description": "CPF (apenas numeros ou com pontuacao)"},
            "cnpj": {"label": "CNPJ", "required": False, "description": "CNPJ para pessoa juridica"},
            "endereco": {"label": "Endereco", "required": False, "description": "Endereco completo"},
            "cidade": {"label": "Cidade", "required": False, "description": "Cidade"},
            "estado": {"label": "Estado (UF)", "required": False, "description": "Sigla do estado (ex: MG, SP)"},
        },
    },
    "processos": {
        "label": "Processos",
        "icon": "fas fa-gavel",
        "fields": {
            "numero_processo": {"label": "Numero do Processo", "required": True, "description": "Numero CNJ do processo"},
            "cliente": {"label": "Cliente (nome ou ID)", "required": True, "description": "Nome do cliente vinculado"},
            "parte_contraria": {"label": "Parte Contraria", "required": False, "description": "Nome da parte contraria"},
            "tribunal": {"label": "Tribunal", "required": False, "description": "Ex: TJMG, TRF1, STJ"},
            "vara": {"label": "Vara", "required": False, "description": "Vara ou turma"},
            "comarca": {"label": "Comarca", "required": False, "description": "Comarca do processo"},
            "tipo_acao": {"label": "Tipo de Acao", "required": False, "description": "Ex: Indenizatoria, Trabalhista"},
            "status": {"label": "Status", "required": False, "description": "Ex: ativo, arquivado, suspenso"},
        },
    },
    "prazos": {
        "label": "Prazos Processuais",
        "icon": "fas fa-clock",
        "fields": {
            "processo": {"label": "Numero do Processo", "required": True, "description": "Numero CNJ vinculado"},
            "tipo_prazo": {"label": "Tipo de Prazo", "required": True, "description": "Ex: contestacao, recurso, audiencia"},
            "data_intimacao": {"label": "Data da Intimacao", "required": True, "description": "Formato: DD/MM/AAAA"},
            "dias": {"label": "Dias", "required": True, "description": "Quantidade de dias uteis para o prazo"},
            "responsavel": {"label": "Responsavel", "required": False, "description": "Nome do advogado responsavel"},
            "cliente": {"label": "Cliente (nome)", "required": False, "description": "Nome do cliente — correlacionado ao processo/cadastro"},
            "parte_contraria": {"label": "Parte Contraria", "required": False, "description": "Nome da parte contraria"},
            "data_vencimento": {"label": "Data de Vencimento", "required": False, "description": "Opcional — se informada, e' validada contra o calculo CPC"},
            "status": {"label": "Status", "required": False, "description": "NO PRAZO / Atencao! / Fatal -> pendente; Fora do Prazo -> perdido"},
            "uf": {"label": "UF (estado)", "required": False, "description": "Sigla do estado p/ feriados (default MG)"},
        },
    },
    "contatos": {
        "label": "Contatos",
        "icon": "fas fa-address-book",
        "fields": {
            "nome": {"label": "Nome", "required": True, "description": "Nome completo do contato"},
            "telefone": {"label": "Telefone", "required": False, "description": "Telefone com DDD"},
            "email": {"label": "E-mail", "required": False, "description": "Endereco de e-mail"},
            "cargo": {"label": "Cargo/Titulo", "required": False, "description": "Ex: Advogado, Perito, Juiz"},
            "escritorio": {"label": "Escritorio/Empresa", "required": False, "description": "Nome do escritorio ou empresa"},
        },
    },
}


# ---------------------------------------------------------------------------
# File parsing helpers
# ---------------------------------------------------------------------------
def parse_upload(file_bytes: bytes, filename: str):
    """Parse CSV or Excel file, return (headers, rows) where rows is list of dicts."""
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else f"col_{i}" for i, cell in enumerate(ws[1])]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = str(val).strip() if val is not None else ""
                rows.append(row_dict)
        wb.close()
        return headers, rows
    else:
        # CSV - handle BOM and different encodings
        try:
            text = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = file_bytes.decode('latin-1')

        # Detect delimiter
        first_line = text.split('\n')[0]
        delimiter = ';' if first_line.count(';') > first_line.count(',') else ','

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        headers = list(reader.fieldnames) if reader.fieldnames else []
        return headers, rows


def _get_context(request: Request, db: Session, **kwargs):
    """Build template context."""
    from i18n import get_translations
    lang = request.cookies.get("lang", "pt")
    t = get_translations(lang)
    user = get_current_user(request, db)
    return {
        "request": request,
        "user": user,
        "PREFIX": PREFIX,
        "t": t,
        "lang": lang,
        **kwargs,
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@router.get("", response_class=HTMLResponse)
async def import_wizard(request: Request, db: Session = Depends(get_db)):
    """Import wizard landing page."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    return templates.TemplateResponse("app/import/wizard.html", _get_context(
        request, db,
        entity_types=ENTITY_FIELDS,
        step="upload",
    ))


@router.post("/upload", response_class=HTMLResponse)
async def import_upload(
    request: Request,
    entity_type: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload CSV/Excel file and return preview with column mapping."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    if entity_type not in ENTITY_FIELDS:
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error=f"Tipo de entidade invalido: {entity_type}",
        ))

    # Read file
    content = await file.read()
    filename = file.filename or "upload.csv"

    if not filename.endswith(('.csv', '.xlsx', '.xls')):
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error="Formato de arquivo nao suportado. Use CSV (.csv) ou Excel (.xlsx).",
        ))

    try:
        headers, rows = parse_upload(content, filename)
    except Exception as e:
        logger.error("Failed to parse upload %s: %s", filename, e)
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error=f"Erro ao ler arquivo: {str(e)}",
        ))

    if not rows:
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error="Arquivo vazio ou sem dados.",
        ))

    # Store full data in session-like approach via hidden form field (JSON)
    import base64
    data_json = base64.b64encode(json.dumps(rows, ensure_ascii=False, default=str).encode()).decode()

    # Auto-map: try to match CSV headers to entity fields
    entity_fields = ENTITY_FIELDS[entity_type]["fields"]
    auto_map = {}
    for csv_col in headers:
        col_lower = csv_col.lower().strip()
        for field_key in entity_fields:
            if col_lower == field_key or col_lower == entity_fields[field_key]["label"].lower():
                auto_map[csv_col] = field_key
                break
            # Fuzzy: check if field key is contained in csv column name
            if field_key in col_lower or col_lower in field_key:
                auto_map[csv_col] = field_key
                break

    preview_rows = rows[:5]

    return templates.TemplateResponse("app/import/wizard.html", _get_context(
        request, db,
        entity_types=ENTITY_FIELDS,
        step="map",
        entity_type=entity_type,
        entity_info=ENTITY_FIELDS[entity_type],
        csv_headers=headers,
        preview_rows=preview_rows,
        total_rows=len(rows),
        auto_map=auto_map,
        data_payload=data_json,
        filename=filename,
    ))


@router.post("/map", response_class=HTMLResponse)
async def import_execute(
    request: Request,
    db: Session = Depends(get_db),
):
    """Accept column mapping and execute the import."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url=f"{PREFIX}/login", status_code=302)

    form = await request.form()
    entity_type = form.get("entity_type", "")
    data_payload = form.get("data_payload", "")
    filename = form.get("filename", "upload.csv")

    if entity_type not in ENTITY_FIELDS:
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error="Tipo de entidade invalido.",
        ))

    # Decode data
    import base64
    try:
        rows = json.loads(base64.b64decode(data_payload).decode())
    except Exception:
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error="Dados de importacao corrompidos. Tente novamente.",
        ))

    # Build column mapping from form: mapping_<csv_col> -> entity_field
    column_map = {}
    for key, value in form.items():
        if key.startswith("mapping_") and value and value != "__skip__":
            csv_col = key[len("mapping_"):]
            column_map[csv_col] = value

    if not column_map:
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="upload",
            error="Nenhuma coluna foi mapeada. Selecione pelo menos uma.",
        ))

    # Execute import
    org_id = getattr(request.state, "org_id", None)
    imported = 0
    errors = []
    skipped = 0

    for i, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        try:
            mapped = {}
            for csv_col, field_key in column_map.items():
                mapped[field_key] = row.get(csv_col, "").strip() if row.get(csv_col) else ""

            result = _import_single_row(db, entity_type, mapped, org_id, user.id)
            if result == "imported":
                imported += 1
            elif result == "skipped":
                skipped += 1
            else:
                errors.append(f"Linha {i}: {result}")
        except Exception as e:
            errors.append(f"Linha {i}: {str(e)}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error("Import commit failed: %s", e)
        return templates.TemplateResponse("app/import/wizard.html", _get_context(
            request, db,
            entity_types=ENTITY_FIELDS,
            step="result",
            entity_type=entity_type,
            entity_info=ENTITY_FIELDS[entity_type],
            imported=0,
            errors=[f"Erro ao salvar no banco: {str(e)}"],
            skipped=0,
            filename=filename,
        ))

    return templates.TemplateResponse("app/import/wizard.html", _get_context(
        request, db,
        entity_types=ENTITY_FIELDS,
        step="result",
        entity_type=entity_type,
        entity_info=ENTITY_FIELDS[entity_type],
        imported=imported,
        errors=errors[:50],  # Limit error display
        skipped=skipped,
        total_rows=len(rows),
        filename=filename,
    ))


def _import_single_row(db: Session, entity_type: str, mapped: dict, org_id, user_id) -> str:
    """Import a single mapped row. Returns 'imported', 'skipped', or error string."""

    if entity_type == "clientes":
        return _import_client(db, mapped, org_id)
    elif entity_type == "processos":
        return _import_case(db, mapped, org_id)
    elif entity_type == "prazos":
        return _import_prazo(db, mapped, org_id)
    elif entity_type == "contatos":
        return _import_contact(db, mapped, org_id, user_id)
    return "Tipo desconhecido"


def _import_client(db: Session, mapped: dict, org_id) -> str:
    """Import a client row."""
    nome = mapped.get("nome", "").strip()
    if not nome:
        return "skipped"  # No name = skip

    # Split name into first/last
    parts = nome.split(None, 1)
    first_name = parts[0] if parts else nome
    last_name = parts[1] if len(parts) > 1 else ""

    # Check for duplicate by email or CPF
    email = mapped.get("email", "").strip()
    cpf = mapped.get("cpf", "").strip()

    if email:
        existing = tenant_query(db, Client, org_id).filter(Client.email == email).first()
        if existing:
            return "skipped"

    client = Client(
        first_name=first_name,
        last_name=last_name,
        email=email or None,
        phone=mapped.get("telefone", "").strip() or None,
        cpf=cpf or None,
        cnpj=mapped.get("cnpj", "").strip() or None,
        address=mapped.get("endereco", "").strip() or None,
        city=mapped.get("cidade", "").strip() or None,
        state=mapped.get("estado", "").strip() or None,
        org_id=org_id,
        status="active",
    )
    db.add(client)
    return "imported"


def _import_case(db: Session, mapped: dict, org_id) -> str:
    """Import a case/processo row."""
    numero = mapped.get("numero_processo", "").strip()
    if not numero:
        return "Numero do processo obrigatorio"

    # Find client by name
    cliente_ref = mapped.get("cliente", "").strip()
    client = None
    if cliente_ref:
        # Try by ID first
        if cliente_ref.isdigit():
            client = tenant_query(db, Client, org_id).filter(Client.id == int(cliente_ref)).first()
        if not client:
            # Try by name match
            client = tenant_query(db, Client, org_id).filter(
                (Client.first_name + " " + Client.last_name).ilike(f"%{cliente_ref}%")
            ).first()
    if not client:
        return f"Cliente nao encontrado: {cliente_ref}"

    # Check duplicate
    existing = tenant_query(db, Case, org_id).filter(Case.numero_processo == numero).first()
    if existing:
        return "skipped"

    # Generate case number
    import secrets
    case_number = f"PROC-{secrets.token_hex(3).upper()}"

    case = Case(
        client_id=client.id,
        case_number=case_number,
        numero_processo=numero,
        polo_passivo=mapped.get("parte_contraria", "").strip() or None,
        tribunal=mapped.get("tribunal", "").strip() or None,
        vara=mapped.get("vara", "").strip() or None,
        comarca=mapped.get("comarca", "").strip() or None,
        tipo_acao=mapped.get("tipo_acao", "").strip() or None,
        status=mapped.get("status", "").strip() or "ativo",
        case_name=f"{numero} - {client.full_name}",
        org_id=org_id,
    )
    db.add(case)
    return "imported"


def _parse_date_br(data_str: str):
    """Parse a date string in common BR/ISO formats. Returns date or None."""
    if not data_str:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(data_str, fmt).date()
        except ValueError:
            continue
    return None


# Map the human-facing status labels (as shown in the spreadsheets the firm uses)
# to the canonical prazos_processuais.status values that Controladoria understands.
# NO PRAZO / Atencao! / Fatal -> pendente ; Fora do Prazo -> perdido ; concluido -> concluido
def _map_prazo_status(status_str: str) -> str:
    s = (status_str or "").strip().lower()
    if not s:
        return "pendente"
    # Normalize accents/punctuation minimally
    s = s.replace("!", "").replace("ç", "c").replace("ã", "a").replace("é", "e").strip()
    if "fora do prazo" in s or "perdido" in s or "vencido" in s:
        return "perdido"
    if "concluido" in s or "concluído" in s or "cumprido" in s:
        return "concluido"
    # "no prazo", "atencao", "fatal" e qualquer outro = ativo/pendente
    return "pendente"


def _import_prazo(db: Session, mapped: dict, org_id) -> str:
    """Import a prazo row into prazos_processuais (the table Controladoria reads).

    Correlates client by name (org-scoped) and case by numero_processo (org-scoped),
    creating Processos+Clientes when missing, computes CPC business-day dates, and
    cross-validates any provided data_vencimento. All queries are org-scoped.
    """
    from datetime import timedelta
    from services.prazos_cpc import calcular_prazo, proximo_dia_util

    processo = mapped.get("processo", "").strip()
    tipo_prazo = mapped.get("tipo_prazo", "").strip()
    data_str = mapped.get("data_intimacao", "").strip()
    dias_str = mapped.get("dias", "").strip()

    if not processo or not tipo_prazo:
        return "Processo e tipo de prazo obrigatorios"

    # Intimacao date (required by the wizard; defensive parse here)
    data_intimacao = _parse_date_br(data_str)
    if data_str and not data_intimacao:
        return f"Data de intimacao invalida: {data_str}"
    if not data_intimacao:
        return "Data de intimacao obrigatoria"

    # Business days for the deadline
    dias = 15  # CPC default (procedimento comum)
    if dias_str:
        try:
            dias = int(float(dias_str))
        except (TypeError, ValueError):
            return f"Dias invalido: {dias_str}"
    if dias <= 0:
        dias = 15

    uf = (mapped.get("uf", "").strip().upper() or "MG")[:2]
    responsavel = mapped.get("responsavel", "").strip()
    cliente_nome = mapped.get("cliente", "").strip()
    parte_contraria = mapped.get("parte_contraria", "").strip()
    status = _map_prazo_status(mapped.get("status", ""))

    # --- Correlate client by name (org-scoped), create if missing ---
    client = None
    if cliente_nome:
        if cliente_nome.isdigit():
            client = tenant_query(db, Client, org_id).filter(Client.id == int(cliente_nome)).first()
        if not client:
            client = tenant_query(db, Client, org_id).filter(
                (Client.first_name + " " + Client.last_name).ilike(f"%{cliente_nome}%")
            ).first()
        if not client:
            parts = cliente_nome.split(None, 1)
            client = Client(
                first_name=parts[0] if parts else cliente_nome,
                last_name=parts[1] if len(parts) > 1 else "",
                org_id=org_id,
                status="active",
            )
            db.add(client)
            db.flush()  # need client.id for the case below

    # --- Correlate case by numero_processo (org-scoped), create if missing ---
    case = tenant_query(db, Case, org_id).filter(Case.numero_processo == processo).first()
    if not case:
        # Only create a case when we can attach a client (cases.client_id is NOT NULL).
        if client:
            import secrets
            case = Case(
                client_id=client.id,
                case_number=f"PROC-{secrets.token_hex(3).upper()}",
                numero_processo=processo,
                polo_passivo=parte_contraria or None,
                status="ativo",
                case_name=f"{processo} - {client.full_name}",
                org_id=org_id,
            )
            db.add(case)
            db.flush()  # need case.id for the prazo below
    else:
        # Backfill parte_contraria on existing case when provided and empty.
        if parte_contraria and not (case.polo_passivo or "").strip():
            case.polo_passivo = parte_contraria

    # --- Compute CPC business-day dates ---
    try:
        data_inicio = proximo_dia_util(data_intimacao + timedelta(days=1), uf)
        data_vencimento = calcular_prazo(data_intimacao, dias, uf, False)
    except Exception as e:
        return f"Erro ao calcular prazo: {str(e)}"

    # --- Cross-validate provided data_vencimento against the CPC calculation ---
    venc_warning = None
    venc_informado = _parse_date_br(mapped.get("data_vencimento", "").strip())
    if venc_informado and venc_informado != data_vencimento:
        # Use the CPC-computed value as source of truth; record the divergence.
        venc_warning = (
            f"vencimento informado {venc_informado.isoformat()} difere do calculo CPC "
            f"{data_vencimento.isoformat()} (intimacao {data_intimacao.isoformat()} + {dias} dias uteis/{uf})"
        )

    # When the case could not be resolved/created, keep the raw process number
    # and client name as overrides so Controladoria still shows them.
    processo_override = None if case else processo
    cliente_override = None if (case or not cliente_nome) else cliente_nome

    descricao_parts = [f"Importado via /import-br (intimacao {data_intimacao.isoformat()}, {dias} dias uteis)"]
    if venc_warning:
        descricao_parts.append(f"Aviso: {venc_warning}")
    descricao = " — ".join(descricao_parts)

    db.execute(
        text("""
            INSERT INTO prazos_processuais
                (case_id, org_id, tipo, data_intimacao, data_inicio, data_vencimento,
                 dias_prazo, responsavel, status, descricao, uf, dobro,
                 processo_override, cliente_override)
            VALUES
                (:case_id, :org_id, :tipo, :data_intimacao, :data_inicio, :data_vencimento,
                 :dias_prazo, :responsavel, :status, :descricao, :uf, :dobro,
                 :processo_override, :cliente_override)
        """),
        {
            "case_id": case.id if case else None,
            "org_id": org_id,
            "tipo": tipo_prazo,
            "data_intimacao": data_intimacao,
            "data_inicio": data_inicio,
            "data_vencimento": data_vencimento,
            "dias_prazo": dias,
            "responsavel": responsavel or None,
            "status": status,
            "descricao": descricao,
            "uf": uf,
            "dobro": False,
            "processo_override": processo_override,
            "cliente_override": cliente_override,
        },
    )
    return "imported"


def _import_contact(db: Session, mapped: dict, org_id, user_id) -> str:
    """Import a contact row."""
    nome = mapped.get("nome", "").strip()
    if not nome:
        return "skipped"

    # Use raw SQL since contacts table is not an ORM model
    try:
        db.execute(
            text("""
                INSERT INTO contacts (name, phone, email, title, company, contact_type, org_id, created_by, is_active)
                VALUES (:name, :phone, :email, :title, :company, :contact_type, :org_id, :created_by, true)
            """),
            {
                "name": nome,
                "phone": mapped.get("telefone", "").strip() or None,
                "email": mapped.get("email", "").strip() or None,
                "title": mapped.get("cargo", "").strip() or None,
                "company": mapped.get("escritorio", "").strip() or None,
                "contact_type": "professional",
                "org_id": org_id,
                "created_by": user_id,
            }
        )
        return "imported"
    except Exception as e:
        return f"Erro ao salvar contato: {str(e)}"


# ---------------------------------------------------------------------------
# Template download
# ---------------------------------------------------------------------------
@router.get("/templates/{entity_type}")
async def download_template(entity_type: str):
    """Download a template CSV file for the given entity type."""
    import os

    template_files = {
        "clientes": "modelo_clientes.csv",
        "processos": "modelo_processos.csv",
        "prazos": "modelo_prazos.csv",
        "contatos": "modelo_contatos.csv",
    }

    if entity_type not in template_files:
        raise HTTPException(status_code=404, detail="Template nao encontrado")

    filepath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "templates", template_files[entity_type])

    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
    else:
        # Fallback: generate in-memory
        content = _generate_template_csv(entity_type)

    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={template_files[entity_type]}",
        },
    )


def _generate_template_csv(entity_type: str) -> str:
    """Generate a template CSV string for the entity type."""
    if entity_type == "clientes":
        return "nome,email,telefone,cpf,cnpj,endereco,cidade,estado\nPessoa Demo,pessoa_demo@example.com,(11) 90000-0000,000.000.000-00,,Rua Exemplo 100,Cidade Demo,UF\nEmpresa Demo Ltda,contato@example.com,(11) 90000-0001,,00.000.000/0000-00,Av Exemplo 1000,Cidade Demo,UF\n"
    elif entity_type == "processos":
        return "numero_processo,cliente,parte_contraria,tribunal,vara,comarca,tipo_acao,status\n0001234-56.2024.8.13.0145,PessoaDemo Silva,Banco XYZ S/A,TJMG,1a Vara Civel,Juiz de Fora,Indenizatoria,ativo\n0009876-54.2024.5.03.0036,Joao Santos,Empresa ABC Ltda,TRT3,2a Vara do Trabalho,Belo Horizonte,Trabalhista,ativo\n"
    elif entity_type == "prazos":
        return (
            "processo,tipo_prazo,data_intimacao,dias,responsavel,cliente,parte_contraria,data_vencimento,status,uf\n"
            "0001234-56.2024.8.13.0145,contestacao,15/03/2026,15,Dr. Carlos,PessoaDemo Silva,Banco XYZ S/A,07/04/2026,NO PRAZO,MG\n"
            "0009876-54.2024.5.03.0036,recurso_ordinario,20/03/2026,8,Dra. Ana,Joao Santos,Empresa ABC Ltda,,Fora do Prazo,MG\n"
        )
    elif entity_type == "contatos":
        return "nome,telefone,email,cargo,escritorio\nDr. Pedro Almeida,(11) 99999-0001,pedro@example.com,Advogado,Almeida & Associados\nDra. Lucia Ferreira,(11) 99999-0002,lucia@example.com,Juiza,TJMG\n"
    return ""
